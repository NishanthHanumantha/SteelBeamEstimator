"""
excel_exporter.py — GroundTruth_Benchmark_Report.xlsx (15 sheets).
MODEL_VERSION: 8.9.1
"""
from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

MODEL_VERSION = "8.9.1"

_HDR = PatternFill("solid", fgColor="1F4E79")
_HFONT = Font(bold=True, color="FFFFFF", size=11)
_TITLE = Font(bold=True, size=14, color="1F4E79")
_THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)


def _header(ws, row: int, n: int) -> None:
    for c in range(1, n + 1):
        cell = ws.cell(row, c)
        cell.fill = _HDR
        cell.font = _HFONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = _THIN


def _table(ws, start: int, headers: List[str], rows: List[List[Any]]) -> int:
    for i, h in enumerate(headers, 1):
        ws.cell(start, i, h)
    _header(ws, start, len(headers))
    r = start + 1
    for row in rows:
        for i, v in enumerate(row, 1):
            cell = ws.cell(r, i, v)
            cell.border = _THIN
        r += 1
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        ws.column_dimensions[letter].width = min(
            36, max((len(str(c.value or "")) for c in col), default=8) + 2
        )
    return r


class ExcelExporter:
    def export(
        self,
        path: Path,
        compiled: Dict[str, Any],
        results: List[Dict[str, Any]],
    ) -> Path:
        wb = Workbook()
        self._exec(wb.active, compiled)
        self._drawing_summary(wb.create_sheet("Drawing Set Summary"), results)
        self._runtime(wb.create_sheet("Pipeline Runtime"), results)
        self._beam_det(wb.create_sheet("Beam Detection"), results)
        self._beam_match(wb.create_sheet("Beam Matching"), results)
        self._bar_det(wb.create_sheet("Bar Detection"), results)
        self._bar_match(wb.create_sheet("Bar Matching"), results)
        self._missing(wb.create_sheet("Missing Bars"), results)
        self._dia(wb.create_sheet("Diameter Comparison"), results)
        self._dia_steel(wb.create_sheet("Diameter Steel Quantity"), results)
        self._steel(wb.create_sheet("Steel Quantity Comparison"), results)
        self._errors(wb.create_sheet("Error Classification"), results, compiled)
        self._overall(wb.create_sheet("Overall Benchmark"), compiled)
        self._stats(wb.create_sheet("Compiled Statistics"), compiled)
        self._rank(wb.create_sheet("Drawing Set Ranking"), compiled)

        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)
        return path

    def _exec(self, ws, compiled: Dict[str, Any]) -> None:
        ws.title = "Executive Summary"
        b = compiled.get("benchmark") or {}
        ws["A1"] = "Ground Truth Benchmark Report — Phase QA.2A"
        ws["A1"].font = _TITLE
        ws["A2"] = f"MODEL_VERSION: {MODEL_VERSION}"
        rows = [[k, b.get(k)] for k in (
            "total_drawing_sets", "production_runs_completed", "model_workbooks_generated",
            "total_beams", "detected_beams", "correct_beams", "beam_detection_pct",
            "total_bars", "detected_bars", "correct_bars", "missing_bars",
            "bar_detection_pct", "bar_accuracy_pct", "steel_accuracy_pct",
            "overall_accuracy_pct", "average_pipeline_runtime_s",
            "validation_status", "recommendation",
        )]
        _table(ws, 4, ["Metric", "Value"], rows)

    def _drawing_summary(self, ws, results: List[Dict[str, Any]]) -> None:
        headers = [
            "Drawing Set", "Pipe OK", "Runtime s", "Beam Det %", "Beam Match %",
            "Bar Det %", "Bar Acc %", "Steel Acc %", "Errors",
        ]
        rows = []
        for r in results:
            s = r.get("drawing_summary") or {}
            rows.append([
                r.get("drawing_set"), s.get("pipeline_success"), s.get("pipeline_elapsed_s"),
                s.get("beam_detection_pct"), s.get("beam_matching_pct"),
                s.get("bar_detection_pct"), s.get("bar_accuracy_pct"),
                s.get("steel_accuracy_pct"), s.get("error_count"),
            ])
        _table(ws, 1, headers, rows)

    def _runtime(self, ws, results: List[Dict[str, Any]]) -> None:
        headers = ["Drawing Set", "Success", "Elapsed s", "Model Excel", "Error"]
        rows = []
        for r in results:
            p = r.get("pipeline") or {}
            rows.append([
                r.get("drawing_set"), p.get("success"), p.get("elapsed_s"),
                p.get("model_excel"), p.get("error"),
            ])
        _table(ws, 1, headers, rows)

    def _beam_det(self, ws, results: List[Dict[str, Any]]) -> None:
        headers = ["Drawing Set", "Est Beams", "Detected", "Undetected", "Extra", "Det %"]
        rows = []
        for r in results:
            b = r.get("beam_matching") or {}
            rows.append([
                r.get("drawing_set"), b.get("estimator_beams"), b.get("detected_beams"),
                b.get("undetected_beams"), b.get("extra_beams"), b.get("detection_pct"),
            ])
        _table(ws, 1, headers, rows)

    def _beam_match(self, ws, results: List[Dict[str, Any]]) -> None:
        headers = ["Drawing Set", "Est Beam", "Model Beam", "Matched", "Method", "Status"]
        rows = []
        for r in results:
            for p in (r.get("beam_matching") or {}).get("pairs") or []:
                rows.append([
                    r.get("drawing_set"), p.get("estimator_beam_id"), p.get("model_beam_id"),
                    p.get("matched"), p.get("method"), p.get("status"),
                ])
        _table(ws, 1, headers, rows)

    def _bar_det(self, ws, results: List[Dict[str, Any]]) -> None:
        headers = ["Drawing Set", "Est Bars", "Detected", "Det %"]
        rows = []
        for r in results:
            b = r.get("bar_matching") or {}
            rows.append([
                r.get("drawing_set"), b.get("estimator_bars"), b.get("detected_bars"),
                b.get("detection_pct"),
            ])
        _table(ws, 1, headers, rows)

    def _bar_match(self, ws, results: List[Dict[str, Any]]) -> None:
        headers = [
            "Drawing Set", "Beam", "Role", "Dia", "Est Qty", "Model Qty",
            "Status", "Diff Qty",
        ]
        rows = []
        for r in results:
            for row in (r.get("bar_matching") or {}).get("rows") or []:
                rows.append([
                    row.get("drawing_set"), row.get("beam_id"), row.get("bar_role"),
                    row.get("diameter"), row.get("estimator_qty"), row.get("model_qty"),
                    row.get("status"), row.get("difference_qty"),
                ])
        _table(ws, 1, headers, rows)

    def _missing(self, ws, results: List[Dict[str, Any]]) -> None:
        headers = ["Drawing Set", "Beam", "Role", "Diameter", "Est Qty"]
        rows = []
        for r in results:
            for row in (r.get("bar_matching") or {}).get("missing_detail") or []:
                rows.append([
                    row.get("drawing_set"), row.get("beam_id"), row.get("bar_role"),
                    row.get("diameter"), row.get("estimator_qty"),
                ])
        _table(ws, 1, headers, rows)

    def _dia(self, ws, results: List[Dict[str, Any]]) -> None:
        headers = [
            "Drawing Set", "Diameter", "Est Qty", "Model Qty", "Diff", "Diff %",
            "Missing %", "Extra %",
        ]
        rows = []
        for r in results:
            for row in ((r.get("metrics") or {}).get("metric6_diameter_accuracy") or []):
                rows.append([
                    r.get("drawing_set"), row.get("diameter_label"), row.get("estimator_quantity"),
                    row.get("model_quantity"), row.get("difference"), row.get("difference_pct"),
                    row.get("missing_pct"), row.get("extra_pct"),
                ])
        _table(ws, 1, headers, rows)

    def _dia_steel(self, ws, results: List[Dict[str, Any]]) -> None:
        headers = ["Drawing Set", "Diameter", "Est KG", "Model KG", "Diff KG", "Diff %"]
        rows = []
        for r in results:
            for row in ((r.get("metrics") or {}).get("metric7_diameter_steel") or []):
                rows.append([
                    r.get("drawing_set"), row.get("diameter_label"), row.get("estimator_kg"),
                    row.get("model_kg"), row.get("difference_kg"), row.get("difference_pct"),
                ])
        _table(ws, 1, headers, rows)

    def _steel(self, ws, results: List[Dict[str, Any]]) -> None:
        headers = [
            "Drawing Set", "Est KG", "Model KG", "Diff KG", "Diff %", "Acc %",
            "Est MT", "Model MT", "Diff MT",
        ]
        rows = []
        for r in results:
            s = ((r.get("metrics") or {}).get("metric8_overall_steel") or {})
            rows.append([
                r.get("drawing_set"), s.get("estimator_total_kg"), s.get("model_total_kg"),
                s.get("difference_kg"), s.get("difference_pct"), s.get("accuracy_pct"),
                s.get("estimator_total_mt"), s.get("model_total_mt"), s.get("difference_mt"),
            ])
        _table(ws, 1, headers, rows)

    def _errors(self, ws, results: List[Dict[str, Any]], compiled: Dict[str, Any]) -> None:
        headers = ["Drawing Set", "Error Type", "Beam", "Detail"]
        rows = []
        for r in results:
            for e in (r.get("errors") or {}).get("items") or []:
                rows.append([e.get("drawing_set"), e.get("error_type"), e.get("beam_id"), e.get("detail")])
        r = _table(ws, 1, headers, rows)
        freq = (compiled.get("errors") or {}).get("frequency") or {}
        _table(ws, r + 2, ["Error Type", "Count"], [[k, v] for k, v in freq.items()])

    def _overall(self, ws, compiled: Dict[str, Any]) -> None:
        b = compiled.get("benchmark") or {}
        _table(ws, 1, ["Field", "Value"], [[k, v] for k, v in b.items()])

    def _stats(self, ws, compiled: Dict[str, Any]) -> None:
        s = compiled.get("statistics") or {}
        _table(ws, 1, ["Statistic", "Value"], [[k, v] for k, v in s.items()])

    def _rank(self, ws, compiled: Dict[str, Any]) -> None:
        headers = [
            "Rank", "Drawing Set", "Overall %", "Beam Det %", "Bar Det %",
            "Bar Acc %", "Steel Acc %",
        ]
        rows = []
        for r in compiled.get("ranking") or []:
            rows.append([
                r.get("rank"), r.get("drawing_set"), r.get("overall_accuracy_pct"),
                r.get("beam_detection_pct"), r.get("bar_detection_pct"),
                r.get("bar_accuracy_pct"), r.get("steel_accuracy_pct"),
            ])
        _table(ws, 1, headers, rows)
