"""
excel_exporter.py — Engineering_Benchmark_Report.xlsx (13 sheets).
MODEL_VERSION: 8.9.0
"""
from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

MODEL_VERSION = "8.9.0"

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
_PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
_FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
_THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)


def _style_header(ws, row: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row, c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = _THIN


def _autosize(ws, max_width: int = 36) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = min(max_width, max((len(str(c.value or "")) for c in col), default=8) + 2)
        ws.column_dimensions[letter].width = max(10, width)


def _write_table(ws, start_row: int, headers: List[str], rows: List[List[Any]]) -> int:
    for i, h in enumerate(headers, 1):
        ws.cell(start_row, i, h)
    _style_header(ws, start_row, len(headers))
    r = start_row + 1
    for row in rows:
        for i, val in enumerate(row, 1):
            cell = ws.cell(r, i, val)
            cell.border = _THIN
            cell.alignment = Alignment(vertical="center")
        r += 1
    return r


class ExcelExporter:
    def export(
        self,
        out_path: Path,
        compiled: Dict[str, Any],
        comparisons: List[Dict[str, Any]],
    ) -> Path:
        wb = Workbook()
        # Remove default after creating sheets
        self._sheet_executive(wb.active, compiled)
        self._sheet_drawing_summary(wb.create_sheet("Drawing Set Summary"), comparisons)
        self._sheet_beam_detection(wb.create_sheet("Beam Detection"), comparisons)
        self._sheet_beam_accuracy(wb.create_sheet("Beam Accuracy"), comparisons)
        self._sheet_bar_detection(wb.create_sheet("Bar Detection"), comparisons)
        self._sheet_bar_accuracy(wb.create_sheet("Bar Accuracy"), comparisons)
        self._sheet_missing_bars(wb.create_sheet("Missing Bars"), comparisons)
        self._sheet_diameter_comparison(wb.create_sheet("Diameter Comparison"), comparisons)
        self._sheet_diameter_steel(wb.create_sheet("Diameter Steel Quantity"), comparisons)
        self._sheet_steel(wb.create_sheet("Steel Quantity Comparison"), comparisons)
        self._sheet_errors(wb.create_sheet("Error Classification"), comparisons, compiled)
        self._sheet_compiled_stats(wb.create_sheet("Compiled Statistics"), compiled)
        self._sheet_overall(wb.create_sheet("Overall Benchmark"), compiled)

        out_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(out_path)
        return out_path

    def _sheet_executive(self, ws, compiled: Dict[str, Any]) -> None:
        ws.title = "Executive Summary"
        bench = compiled.get("benchmark") or {}
        ws["A1"] = "Engineering Benchmark Report — Phase QA.2"
        ws["A1"].font = _TITLE_FONT
        ws["A2"] = f"MODEL_VERSION: {MODEL_VERSION}"
        rows = [
            ("Total Drawing Sets", bench.get("total_drawing_sets")),
            ("Total Beams", bench.get("total_beams")),
            ("Detected Beams", bench.get("detected_beams")),
            ("Correct Beams", bench.get("correct_beams")),
            ("Beam Detection %", bench.get("beam_detection_pct")),
            ("Total Bars", bench.get("total_bars")),
            ("Detected Bars", bench.get("detected_bars")),
            ("Correct Bars", bench.get("correct_bars")),
            ("Missing Bars", bench.get("missing_bars")),
            ("Bar Detection %", bench.get("bar_detection_pct")),
            ("Bar Accuracy %", bench.get("bar_accuracy_pct")),
            ("Steel Accuracy %", bench.get("steel_accuracy_pct")),
            ("Overall Accuracy %", bench.get("overall_accuracy_pct")),
            ("Validation Status", bench.get("validation_status")),
            ("Recommendation", bench.get("recommendation")),
        ]
        _write_table(ws, 4, ["Metric", "Value"], [[a, b] for a, b in rows])
        _autosize(ws)

    def _sheet_drawing_summary(self, ws, comparisons: List[Dict[str, Any]]) -> None:
        headers = [
            "Drawing Set", "Beam Det %", "Beam Acc %", "Bar Det %", "Bar Acc %",
            "Steel Acc %", "Est Beams", "Det Beams", "Est Bars", "Det Bars",
            "Missing Bars", "Est KG", "Model KG", "Errors",
        ]
        rows = []
        for c in comparisons:
            s = c.get("summary") or {}
            rows.append([
                s.get("drawing_set"), s.get("beam_detection_pct"), s.get("beam_accuracy_pct"),
                s.get("bar_detection_pct"), s.get("bar_accuracy_pct"), s.get("steel_accuracy_pct"),
                s.get("total_estimator_beams"), s.get("detected_beams"),
                s.get("estimator_bars"), s.get("detected_bars"), s.get("missing_bars"),
                s.get("estimator_kg"), s.get("model_kg"), s.get("error_count"),
            ])
        _write_table(ws, 1, headers, rows)
        _autosize(ws)

    def _sheet_beam_detection(self, ws, comparisons: List[Dict[str, Any]]) -> None:
        headers = ["Drawing Set", "Total Est Beams", "Detected", "Undetected", "Extra", "Detection %", "Missing IDs"]
        rows = []
        for c in comparisons:
            d = c.get("beam_detection") or {}
            rows.append([
                c.get("drawing_set"), d.get("total_estimator_beams"), d.get("detected_beams"),
                d.get("undetected_beams"), d.get("extra_beams"), d.get("detection_pct"),
                ", ".join(d.get("missing_ids") or []),
            ])
        _write_table(ws, 1, headers, rows)
        _autosize(ws)

    def _sheet_beam_accuracy(self, ws, comparisons: List[Dict[str, Any]]) -> None:
        headers = [
            "Drawing Set", "Beam ID", "Detected", "Matched", "Est Bars", "Det Bars",
            "Correct Bars", "Missing Bars", "Detection %", "Accuracy %", "Steel Diff", "Status",
        ]
        rows = []
        for c in comparisons:
            for r in c.get("beam_level") or []:
                rows.append([
                    r.get("drawing_set"), r.get("beam_id"), r.get("detected"), r.get("matched"),
                    r.get("estimator_bars"), r.get("detected_bars"), r.get("correct_bars"),
                    r.get("missing_bars"), r.get("detection_pct"), r.get("accuracy_pct"),
                    r.get("steel_difference_kg"), r.get("status"),
                ])
        _write_table(ws, 1, headers, rows)
        _autosize(ws)

    def _sheet_bar_detection(self, ws, comparisons: List[Dict[str, Any]]) -> None:
        headers = ["Drawing Set", "Est Bars", "Detected Bars", "Detection %"]
        rows = []
        for c in comparisons:
            d = c.get("bar_detection") or {}
            rows.append([c.get("drawing_set"), d.get("estimator_bars"), d.get("detected_bars"), d.get("detection_pct")])
        _write_table(ws, 1, headers, rows)
        _autosize(ws)

    def _sheet_bar_accuracy(self, ws, comparisons: List[Dict[str, Any]]) -> None:
        headers = [
            "Drawing Set", "Beam", "Bar Role", "Diameter", "Est Qty", "Model Qty",
            "Matched", "Difference", "Status",
        ]
        rows = []
        for c in comparisons:
            for r in c.get("bar_rows") or []:
                rows.append([
                    r.get("drawing_set"), r.get("beam_id"), r.get("bar_role"), r.get("diameter"),
                    r.get("estimator_qty"), r.get("model_qty"), r.get("matched"),
                    r.get("difference"), r.get("status"),
                ])
        _write_table(ws, 1, headers, rows)
        _autosize(ws)

    def _sheet_missing_bars(self, ws, comparisons: List[Dict[str, Any]]) -> None:
        headers = ["Drawing Set", "Beam", "Bar Role", "Diameter", "Est Qty", "Status"]
        rows = []
        for c in comparisons:
            for r in c.get("missing_bars") or []:
                rows.append([
                    r.get("drawing_set"), r.get("beam_id"), r.get("bar_role"),
                    r.get("diameter"), r.get("estimator_qty"), r.get("status"),
                ])
        _write_table(ws, 1, headers, rows)
        _autosize(ws)

    def _sheet_diameter_comparison(self, ws, comparisons: List[Dict[str, Any]]) -> None:
        headers = [
            "Drawing Set", "Diameter", "Est Bars", "Model Bars", "Diff", "Diff %",
            "Correct", "Incorrect",
        ]
        rows = []
        for c in comparisons:
            for r in c.get("diameter_comparison") or []:
                rows.append([
                    c.get("drawing_set"), r.get("diameter_label"), r.get("estimator_qty_bars"),
                    r.get("model_qty_bars"), r.get("difference_bars"), r.get("difference_pct_bars"),
                    r.get("correct_detection"), r.get("incorrect_detection"),
                ])
        _write_table(ws, 1, headers, rows)
        _autosize(ws)

    def _sheet_diameter_steel(self, ws, comparisons: List[Dict[str, Any]]) -> None:
        headers = [
            "Drawing Set", "Diameter", "Est KG", "Model KG", "Difference", "Diff %",
            "Abs Error", "Correct Count", "Incorrect Count",
        ]
        rows = []
        for c in comparisons:
            for r in c.get("diameter_steel") or []:
                rows.append([
                    c.get("drawing_set"), r.get("diameter_label"), r.get("estimator_kg"),
                    r.get("model_kg"), r.get("difference_kg"), r.get("difference_pct"),
                    r.get("absolute_error_kg"), r.get("correct_count"), r.get("incorrect_count"),
                ])
        _write_table(ws, 1, headers, rows)
        _autosize(ws)

    def _sheet_steel(self, ws, comparisons: List[Dict[str, Any]]) -> None:
        headers = [
            "Drawing Set", "Est KG", "Model KG", "Diff KG", "Diff %", "Accuracy %",
            "Est MT", "Model MT", "Diff MT",
        ]
        rows = []
        for c in comparisons:
            s = c.get("steel_quantity") or {}
            rows.append([
                c.get("drawing_set"), s.get("estimator_total_kg"), s.get("model_total_kg"),
                s.get("difference_kg"), s.get("difference_pct"), s.get("accuracy_pct"),
                s.get("estimator_total_mt"), s.get("model_total_mt"), s.get("difference_mt"),
            ])
        _write_table(ws, 1, headers, rows)
        _autosize(ws)

    def _sheet_errors(self, ws, comparisons: List[Dict[str, Any]], compiled: Dict[str, Any]) -> None:
        headers = ["Drawing Set", "Error Type", "Beam ID", "Detail"]
        rows = []
        for c in comparisons:
            for e in (c.get("errors") or {}).get("items") or []:
                rows.append([c.get("drawing_set"), e.get("error_type"), e.get("beam_id"), e.get("detail")])
        r = _write_table(ws, 1, headers, rows)
        ws.cell(r + 1, 1, "Compiled Frequency")
        ws.cell(r + 1, 1).font = Font(bold=True)
        freq = (compiled.get("errors") or {}).get("frequency") or {}
        _write_table(ws, r + 2, ["Error Type", "Count"], [[k, v] for k, v in freq.items()])
        _autosize(ws)

    def _sheet_compiled_stats(self, ws, compiled: Dict[str, Any]) -> None:
        stats = compiled.get("statistics") or {}
        rows = [[k, v] for k, v in stats.items()]
        _write_table(ws, 1, ["Statistic", "Value"], rows)
        _autosize(ws)

    def _sheet_overall(self, ws, compiled: Dict[str, Any]) -> None:
        bench = compiled.get("benchmark") or {}
        rows = [[k, v] for k, v in bench.items()]
        _write_table(ws, 1, ["Field", "Value"], rows)
        _autosize(ws)
