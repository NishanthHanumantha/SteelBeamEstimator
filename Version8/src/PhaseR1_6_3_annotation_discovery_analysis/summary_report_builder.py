"""
Export JSON, Markdown, and Excel engineering review package.
MODEL_VERSION: 8.8.3
"""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, List

from beam_analysis_model import BeamAnalysisRecord, MODEL_VERSION, PHASE_ID
from input_loader import natural_beam_key

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:  # pragma: no cover
    Workbook = None  # type: ignore


class SummaryReportBuilder:
    def __init__(self, output_dir: Path):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def export_all(self, payload: Dict[str, Any]) -> Dict[str, str]:
        records: List[BeamAnalysisRecord] = payload["records"]
        detected_ids = set(payload["detected_ids"])
        missing_ids = set(payload["missing_ids"])
        paths: Dict[str, str] = {}

        def dump(name: str, data: Any) -> None:
            p = self.output_dir / name
            p.write_text(json.dumps(data, indent=2, default=str), encoding="utf-8")
            paths[name] = str(p)

        inventory = [r.inventory.to_dict() for r in records]
        evidence = [r.drawing_evidence.to_dict() for r in records]
        traces = [
            {
                "beam_id": r.inventory.beam_id,
                "rule012_status": r.rule012_status,
                "stages": [s.to_dict() for s in r.pipeline_trace],
            }
            for r in records
        ]
        detected = [
            {
                "beam_id": r.inventory.beam_id,
                **r.stirrup_status.to_dict(),
                "drawing_name": r.inventory.drawing_name,
            }
            for r in records if r.inventory.beam_id in detected_ids
        ]
        missing = [
            {
                "beam_id": r.inventory.beam_id,
                **r.stirrup_status.to_dict(),
                "drawing_name": r.inventory.drawing_name,
                "unknown_annotation_texts": r.drawing_evidence.unknown_annotation_texts,
                "role_counts": r.drawing_evidence.role_counts,
            }
            for r in records if r.inventory.beam_id in missing_ids
        ]

        dump("beam_inventory.json", {"model_version": MODEL_VERSION, "beams": inventory})
        dump("beam_annotation_evidence.json", {"model_version": MODEL_VERSION, "beams": evidence})
        dump("beam_pipeline_trace.json", {"model_version": MODEL_VERSION, "beams": traces})
        dump("detected_beams.json", {"model_version": MODEL_VERSION, "count": len(detected), "beams": detected})
        dump("missing_beams.json", {"model_version": MODEL_VERSION, "count": len(missing), "beams": missing})
        dump("beam_comparison_statistics.json", payload["comparison"])
        dump("pattern_analysis.json", payload["pattern"])
        dump("engineering_review_dataset.json", payload["review_dataset"])
        dump("engineering_questions.json", payload["questions"])
        dump("annotation_discovery_summary.json", {
            "model_version": MODEL_VERSION,
            "phase": PHASE_ID,
            "statistics": payload["statistics"],
            "validation": payload["validation"],
            "regression": payload["regression"],
            "recommendation": payload["recommendation"],
            "pattern_conclusion": payload["pattern"].get("pattern_conclusion"),
            "elapsed_s": payload.get("elapsed_s"),
        })

        md = self._markdown(payload, detected, missing)
        md_path = self.output_dir / "phase_r163_summary.md"
        md_path.write_text(md, encoding="utf-8")
        paths["phase_r163_summary.md"] = str(md_path)

        xlsx_path = self.output_dir / "engineering_review_report.xlsx"
        self._write_excel(xlsx_path, payload, records, detected_ids, missing_ids)
        paths["engineering_review_report.xlsx"] = str(xlsx_path)
        return paths

    def _markdown(self, payload: Dict[str, Any], detected: List[Dict], missing: List[Dict]) -> str:
        stats = payload["statistics"]
        pattern = payload["pattern"]
        lines = [
            "# Phase R.1.6.3 — Annotation Discovery Analysis & Engineering Review",
            "",
            f"**MODEL_VERSION:** `{MODEL_VERSION}`",
            f"**Recommendation:** `{payload['recommendation']}`",
            "",
            "## Executive Summary",
            "",
            "This phase collects evidence only. It does not correct annotations, intents, details, or EngineeringBars.",
            "",
            f"- Total beams: `{stats.get('total_beams')}`",
            f"- Detected stirrup beams (RULE-012 PASS): `{stats.get('detected_beams')}`",
            f"- Missing stirrup beams (RULE-012 FAIL): `{stats.get('missing_beams')}`",
            f"- Coverage %: `{stats.get('coverage_pct')}`",
            f"- Pattern conclusion: **{pattern.get('pattern_conclusion')}**",
            "",
            "## Pattern Observations",
            "",
        ]
        for obs in pattern.get("observations") or []:
            lines.append(f"- {obs}")
        lines.extend([
            "",
            "## Detected Beams",
            "",
            ", ".join(d["beam_id"] for d in detected) or "(none)",
            "",
            "## Missing Beams (sample)",
            "",
            ", ".join(m["beam_id"] for m in missing[:25]) + (" …" if len(missing) > 25 else ""),
            "",
            "## Meeting Checklist",
            "",
            "- [ ] Review detected vs missing lists with Estimation Team",
            "- [ ] Inspect sample missing beams on DXF",
            "- [ ] Confirm typical / schedule / shared-annotation conventions",
            "- [ ] Fill Engineering Comments / Estimator Comments / Root Cause columns",
            "- [ ] Decide whether deterministic annotation improvement or LLM-assisted recovery is next",
            "",
            "## Recommended Next Actions",
            "",
            "1. Hold Estimation Team review using `engineering_review_report.xlsx`.",
            "2. Do not implement corrections until root-cause fields are completed.",
            "3. After review, prioritise annotation discovery improvements for missing stirrups.",
            "",
        ])
        return "\n".join(lines)

    def _write_excel(
        self,
        path: Path,
        payload: Dict[str, Any],
        records: List[BeamAnalysisRecord],
        detected_ids: set,
        missing_ids: set,
    ) -> None:
        if Workbook is None:
            path.write_text("openpyxl not available", encoding="utf-8")
            return
        wb = Workbook()
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(color="FFFFFF", bold=True)

        def sheet(title: str):
            if title == "Executive Summary":
                ws = wb.active
                ws.title = title
            else:
                ws = wb.create_sheet(title)
            return ws

        def write_table(ws, headers: List[str], rows: List[List[Any]]) -> None:
            ws.append(headers)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(wrap_text=True)
            for row in rows:
                ws.append(row)

        stats = payload["statistics"]
        ws = sheet("Executive Summary")
        write_table(ws, ["Item", "Value"], [
            ["MODEL_VERSION", MODEL_VERSION],
            ["Total Beams", stats.get("total_beams")],
            ["Detected Beams", stats.get("detected_beams")],
            ["Missing Beams", stats.get("missing_beams")],
            ["Coverage %", stats.get("coverage_pct")],
            ["Pattern Conclusion", payload["pattern"].get("pattern_conclusion")],
            ["Recommendation", payload["recommendation"]],
            ["Detection Only", "YES"],
            ["LLM Used", "NO"],
            ["Production Modified", "NO"],
        ])

        ws = sheet("Beam Inventory")
        write_table(ws, [
            "Beam ID", "Length mm", "Width mm", "Depth mm", "Orientation",
            "Drawing", "Registry Status", "RULE-012",
        ], [
            [
                r.inventory.beam_id,
                r.inventory.beam_length_mm,
                r.inventory.beam_width_mm,
                r.inventory.beam_depth_mm,
                r.inventory.orientation,
                r.inventory.drawing_name,
                r.inventory.registry_status,
                r.rule012_status,
            ]
            for r in sorted(records, key=lambda x: natural_beam_key(x.inventory.beam_id))
        ])

        ws = sheet("Detected Beams")
        write_table(ws, [
            "Beam ID", "Notation", "Diameter", "Spacing", "Legs", "EngineeringBar Count", "Drawing",
        ], [
            [
                r.inventory.beam_id,
                r.stirrup_status.detected_notation,
                r.stirrup_status.detected_diameter_mm,
                r.stirrup_status.spacing_mm,
                r.stirrup_status.leg_count,
                r.stirrup_status.engineeringbar_count,
                r.inventory.drawing_name,
            ]
            for r in records if r.inventory.beam_id in detected_ids
        ])

        ws = sheet("Missing Beams")
        write_table(ws, [
            "Beam ID", "Note", "Annotation Count", "Roles", "Unknown Texts", "Drawing",
        ], [
            [
                r.inventory.beam_id,
                r.stirrup_status.note or "No Stirrup Representation",
                r.drawing_evidence.annotation_count,
                json.dumps(r.drawing_evidence.role_counts),
                "; ".join(r.drawing_evidence.unknown_annotation_texts[:8]),
                r.inventory.drawing_name,
            ]
            for r in records if r.inventory.beam_id in missing_ids
        ])

        ws = sheet("Pipeline Traceability")
        write_table(ws, [
            "Beam ID", "Annotation", "Intent", "Detail", "Piece", "EngineeringBars", "RULE-012",
        ], [
            [
                r.inventory.beam_id,
                *[s.status for s in r.pipeline_trace],
                r.rule012_status,
            ]
            for r in records
        ])

        ws = sheet("Drawing Evidence")
        write_table(ws, [
            "Beam ID", "Annotation Count", "Nearest Dist", "Leaders Near",
            "Relationships", "Layers", "Labels", "Section Refs", "Coords",
        ], [
            [
                r.inventory.beam_id,
                r.drawing_evidence.annotation_count,
                r.drawing_evidence.nearest_annotation_distance,
                r.drawing_evidence.leader_count_near_beam,
                r.drawing_evidence.relationship_count,
                ", ".join(r.drawing_evidence.layer_names),
                ", ".join(r.drawing_evidence.associated_reinforcement_labels[:8]),
                ", ".join(r.drawing_evidence.section_references[:5]),
                json.dumps(r.drawing_evidence.coordinates),
            ]
            for r in records
        ])

        ws = sheet("Pattern Analysis")
        write_table(ws, ["Type", "Text"], [
            *[["Observation", o] for o in payload["pattern"].get("observations") or []],
            ["Conclusion", payload["pattern"].get("pattern_conclusion")],
            ["Disclaimer", payload["pattern"].get("disclaimer")],
        ])

        ws = sheet("Engineering Review")
        write_table(ws, [
            "Beam ID", "Detected", "Notation", "Drawing File", "Evidence Summary",
            "Engineering Comments", "Estimator Comments", "Root Cause", "Status",
        ], [
            [
                row["beam_id"], row["detected"], row.get("detected_notation"), row.get("drawing_file"),
                row.get("evidence_summary"), row.get("engineering_comments"),
                row.get("estimator_comments"), row.get("root_cause"), row.get("status"),
            ]
            for row in payload["review_dataset"].get("rows") or []
        ])

        ws = sheet("Questions for Estimation Team")
        write_table(ws, ["ID", "Focus", "Question"], [
            [q["id"], q["focus"], q["question"]]
            for q in payload["questions"].get("questions") or []
        ])

        wb.save(path)
