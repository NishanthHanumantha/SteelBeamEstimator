"""
Load estimator workbooks into sheet grids (values only).
MODEL_VERSION: 8.6.0
"""
from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Optional, Union

import openpyxl

MODEL_VERSION = "8.6.0"


class EstimatorWorkbookLoader:
    """Read-only workbook loader — no sheet-name assumptions."""

    def __init__(self, path: Union[str, Path]):
        self.path = Path(path)
        if not self.path.exists():
            raise FileNotFoundError(self.path)
        self._wb = openpyxl.load_workbook(self.path, data_only=True)
        self.sheet_names = list(self._wb.sheetnames)
        self.grids: Dict[str, List[List[Any]]] = {}
        for name in self.sheet_names:
            self.grids[name] = self._sheet_to_grid(self._wb[name])

    def close(self) -> None:
        self._wb.close()

    @staticmethod
    def _sheet_to_grid(ws) -> List[List[Any]]:
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        grid: List[List[Any]] = []
        for r in range(1, max_row + 1):
            row = [ws.cell(r, c).value for c in range(1, max_col + 1)]
            grid.append(row)
        return grid

    def info(self) -> Dict[str, Any]:
        return {
            "path": str(self.path),
            "filename": self.path.name,
            "sheet_names": self.sheet_names,
            "size_bytes": self.path.stat().st_size,
            "sheet_dims": {
                name: {"rows": len(grid), "cols": max((len(r) for r in grid), default=0)}
                for name, grid in self.grids.items()
            },
        }


def discover_estimator_workbook(folder: Union[str, Path]) -> Optional[Path]:
    folder = Path(folder)
    if not folder.exists():
        return None
    candidates = sorted(
        [p for p in folder.glob("*.xlsx") if not p.name.startswith("~$")],
        key=lambda p: p.stat().st_size,
        reverse=True,
    )
    return candidates[0] if candidates else None
