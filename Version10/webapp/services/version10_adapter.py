"""
Version10 production pipeline adapter (Phase W.2).

Application layer only. Invokes existing Run_PY runners under RunContext.
Does not reimplement discovery, T1, hybrid, steel calculation, or Excel logic.
"""
from __future__ import annotations

import json
import logging
import os
import subprocess
import sys
import time
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional

import config

logger = logging.getLogger("steel_webapp.v10_adapter")

StageCallback = Optional[Callable[[str, str], None]]


class AdapterError(Exception):
    """User-facing adapter/pipeline failure."""


@dataclass
class AdapterResult:
    success: bool
    run_id: str
    output_path: Optional[str] = None
    output_filename: Optional[str] = None
    summary: Dict[str, Any] = field(default_factory=dict)
    warnings: List[str] = field(default_factory=list)
    error: Optional[str] = None
    stages_run: List[str] = field(default_factory=list)
    duration_s: float = 0.0
    engine_root: str = ""
    t1_executed: bool = False


def pipeline_mode() -> str:
    return (os.environ.get("STEEL_WEB_PIPELINE_MODE") or "live").strip().lower()


def stage_env(staging: Path) -> Dict[str, str]:
    env = os.environ.copy()
    env["STEEL_ENGINE_ROOT"] = str(config.ENGINE_ROOT.resolve())
    env["STEEL_RUN_ROOT"] = str(staging.resolve())
    env["STEEL_OUTPUT_ROOT"] = str((staging / "data" / "output").resolve())
    return env


def write_r2a_gn_pointer(gn_path: Path) -> None:
    config.R2A_GN_POINTER.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "general_notes_dxf": str(gn_path.resolve()),
        "project_id": "WEBAPP_UPLOAD",
        "source": config.GN_POINTER_SOURCE,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
    }
    config.R2A_GN_POINTER.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def clear_r2a_gn_pointer() -> None:
    try:
        pointer = config.R2A_GN_POINTER
        if not pointer.exists():
            return
        data = json.loads(pointer.read_text(encoding="utf-8"))
        if data.get("source") == config.GN_POINTER_SOURCE:
            pointer.unlink(missing_ok=True)
    except Exception:
        logger.warning("Could not clear R.2A GN pointer file")


def load_estimator_summary(staging: Path) -> Dict[str, Any]:
    """Only fields actually written by VB.1 JSON exports."""
    summary: Dict[str, Any] = {}
    steel_path = staging / config.STEEL_SUMMARY_REL
    totals_path = staging / config.ENGINEERING_TOTALS_REL
    if steel_path.exists():
        try:
            data = json.loads(steel_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            data = {}
        if isinstance(data, dict):
            if data.get("total_beams") is not None:
                summary["total_beams"] = data.get("total_beams")
            if data.get("total_weight_kg") is not None:
                summary["total_steel_kg"] = data.get("total_weight_kg")
            if data.get("total_bars") is not None:
                summary["total_bars"] = data.get("total_bars")
    if totals_path.exists() and "total_steel_kg" not in summary:
        try:
            totals = json.loads(totals_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            totals = {}
        if isinstance(totals, dict):
            if totals.get("total_beams") is not None:
                summary.setdefault("total_beams", totals.get("total_beams"))
            if totals.get("total_steel_kg") is not None:
                summary["total_steel_kg"] = totals.get("total_steel_kg")
    return summary


def _write_stub_xlsx(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    try:
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Project_Header"
        ws["A1"] = "Steel Beam Estimation"
        ws["A2"] = "STEEL_WEB_PIPELINE_MODE=stub"
        ws["A3"] = "Not a production engineering workbook"
        wb.save(path)
    except Exception:
        path.write_bytes(
            b"PK\x03\x04stub-xlsx-placeholder-not-openpyxl\n"
        )


def _write_stub_hybrid(staging: Path, run_id: str) -> None:
    """Stub-pipeline Hybrid artefact. Does not invoke Claude."""
    try:
        from PhaseW5_production_hybrid_shadow.settings import load_settings

        cfg = load_settings()
        mode = cfg.mode
        key = cfg.api_key_status
    except Exception:
        mode = "off"
        key = "ABSENT"
    if mode == "off":
        return
    if key != "PRESENT":
        classification = "HYBRID_UNAVAILABLE"
        status = "KEY_ABSENT"
        reason = "ANTHROPIC_API_KEY_ABSENT" if key == "ABSENT" else "ANTHROPIC_API_KEY_EMPTY"
    else:
        classification = "HYBRID_UNAVAILABLE"
        status = "NO_ENGINEERING_CONTEXT"
        reason = "STUB_PIPELINE_NO_R13"
    rel = getattr(config, "W6_OBSERVABILITY_REL", None) or (
        "data/output/PhaseW6_hybrid_semantic_resolution/hybrid_observability.json"
    )
    path = staging / rel
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(
            {
                "phase_id": "W.6",
                "run_id": run_id,
                "hybrid_mode": mode,
                "hybrid_enabled": True,
                "hybrid_invocation_attempted": False,
                "api_key_configured": key == "PRESENT",
                "classification": classification,
                "hybrid_status": status,
                "reason": reason,
                "production_authority": "semantic_only" if mode == "production" else "none",
                "production_authority_applied": False,
                "claude_invocation_count": 0,
                "request_count": 0,
                "successful_invocation_count": 0,
                "failed_invocation_count": 0,
                "fallback_used": True,
                "stub": True,
            },
            indent=2,
        ),
        encoding="utf-8",
    )


def _run_stub_pipeline(run_id: str, staging: Path) -> None:
    fail_stage = (os.environ.get("STEEL_WEB_FAIL_STAGE") or "").strip()
    for stage in config.PRODUCTION_STAGES:
        if fail_stage and stage["id"] == fail_stage:
            raise AdapterError(
                f"Engineering pipeline failed during stage {stage['id']}. "
                "Check webapp/logs/webapp.log for details, then try again."
            )
        rel = config.SOFT_ARTEFACTS.get(stage["id"])
        if stage["id"] == "HYBRID":
            _write_stub_hybrid(staging, run_id)
            continue
        if rel:
            artefact = staging / rel
            artefact.parent.mkdir(parents=True, exist_ok=True)
            if stage["id"] == "VB1":
                _write_stub_xlsx(artefact)
            else:
                artefact.write_text(
                    json.dumps({"stub": True, "stage": stage["id"], "run_id": run_id}),
                    encoding="utf-8",
                )
    excel = staging / config.VB1_EXCEL_REL
    if not excel.exists():
        _write_stub_xlsx(excel)
    summary_path = staging / config.STEEL_SUMMARY_REL
    summary_path.parent.mkdir(parents=True, exist_ok=True)
    summary_path.write_text(
        json.dumps(
            {
                "total_weight_kg": 0.0,
                "total_beams": 0,
                "total_bars": 0,
                "calculation_method": "STUB",
            },
            indent=2,
        ),
        encoding="utf-8",
    )


def _run_live_stage(stage: Dict[str, Any], staging: Path) -> None:
    script = config.ENGINE_ROOT / stage["script"]
    if not script.exists():
        raise AdapterError(f"Production runner not found: {stage['script']}")

    cmd = [sys.executable, str(script), str(staging.resolve())]
    logger.info("Runner start stage=%s cmd=%s", stage["id"], " ".join(cmd))
    t0 = time.perf_counter()
    proc = subprocess.run(
        cmd,
        cwd=str(config.ENGINE_ROOT),
        env=stage_env(staging),
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        timeout=int(stage.get("timeout_s") or 600),
    )
    elapsed = round(time.perf_counter() - t0, 2)
    logger.info(
        "Runner finish stage=%s exit=%s duration_s=%s",
        stage["id"],
        proc.returncode,
        elapsed,
    )
    if proc.returncode == 0:
        return

    err_tail = (proc.stderr or proc.stdout or "").strip()
    if err_tail:
        logger.error(
            "Stage %s stderr/stdout (tail):\n%s",
            stage["id"],
            "\n".join(err_tail.splitlines()[-40:]),
        )
    rel = config.SOFT_ARTEFACTS.get(stage["id"])
    if rel and (staging / rel).exists():
        logger.warning(
            "Stage %s exit=%s with artefact present — soft success",
            stage["id"],
            proc.returncode,
        )
        return
    raise AdapterError(
        f"Engineering pipeline failed during stage {stage['id']}. "
        "Check webapp/logs/webapp.log for details, then try again."
    )


def invoke_version10_pipeline(
    run_id: str,
    staging: Path,
    gn_path: Path,
    on_stage: StageCallback = None,
) -> AdapterResult:
    """
    Execute the canonical Version10 production stages for one isolated run.

    Returns a web-friendly AdapterResult. Excel is left in the run tree;
    the Flask service copies it to a unique download name.
    """
    t0 = time.perf_counter()
    warnings: List[str] = []
    stages_run: List[str] = []
    engine_root = str(config.ENGINE_ROOT.resolve())
    mode = pipeline_mode()

    if "version10" not in engine_root.lower().replace("\\", "/"):
        warnings.append(
            f"ENGINE_ROOT does not look like Version10: {engine_root}"
        )
    if not config.t1_is_configured():
        raise AdapterError("T1 is missing from the Version10 production stage list.")
    if not config.t1_runner_path().exists():
        raise AdapterError("T1 production runner is not present in Version10/Run_PY.")

    manifest = {
        "run_id": run_id,
        "app_release": config.APP_RELEASE,
        "engine_label": config.ENGINE_LABEL,
        "engine_root": engine_root,
        "pipeline_mode": mode,
        "stages": [s["id"] for s in config.PRODUCTION_STAGES],
        "t1_included": True,
        "started_at": datetime.now(timezone.utc).isoformat(),
    }
    staging.mkdir(parents=True, exist_ok=True)
    (staging / "logs").mkdir(parents=True, exist_ok=True)
    (staging / "data" / "output").mkdir(parents=True, exist_ok=True)
    (staging / "run_manifest.json").write_text(
        json.dumps(manifest, indent=2), encoding="utf-8"
    )

    logger.info(
        "Pipeline invoke run_id=%s mode=%s engine_root=%s staging=%s gn=%s",
        run_id,
        mode,
        engine_root,
        staging,
        gn_path,
    )

    try:
        write_r2a_gn_pointer(gn_path)
        if mode == "stub":
            if on_stage:
                on_stage("VROOT1", "Preparing estimation...")
            _run_stub_pipeline(run_id, staging)
            stages_run = [s["id"] for s in config.PRODUCTION_STAGES]
        else:
            for stage in config.PRODUCTION_STAGES:
                if on_stage:
                    on_stage(stage["id"], stage["label"])
                _run_live_stage(stage, staging)
                stages_run.append(stage["id"])
                if stage["id"] == "HYBRID":
                    w6 = staging / getattr(
                        config,
                        "W6_OBSERVABILITY_REL",
                        "data/output/PhaseW6_hybrid_semantic_resolution/hybrid_observability.json",
                    )
                    if w6.is_file():
                        try:
                            obs = json.loads(w6.read_text(encoding="utf-8"))
                        except (OSError, json.JSONDecodeError):
                            obs = {}
                        if obs.get("fallback_used"):
                            warnings.append(
                                "Hybrid semantic fallback used: "
                                f"{obs.get('classification') or obs.get('reason')}"
                            )

        required = {
            "EngineeringFacts.json": staging / config.R21D_FACTS_REL,
            "geometry_registry.json": staging / config.L22_REGISTRY_REL,
            "GeometryContexts.json": staging / config.R3_CONTEXTS_REL,
            "Estimation_Output.xlsx": staging / config.VB1_EXCEL_REL,
        }
        for label, path in required.items():
            if not path.exists():
                raise AdapterError(
                    f"Production pipeline completed but {label} was not generated "
                    f"at {path}."
                )
        t1_artefact = staging / config.T1_EVIDENCE_REL
        t1_executed = "T1" in stages_run
        if t1_executed and not t1_artefact.exists():
            warnings.append("T1 ran but stirrup_geometry_evidence.json was not found.")

        summary = load_estimator_summary(staging)
        excel = staging / config.VB1_EXCEL_REL
        duration = round(time.perf_counter() - t0, 2)
        logger.info(
            "Pipeline complete run_id=%s excel=%s duration_s=%s t1=%s summary=%s",
            run_id,
            excel,
            duration,
            t1_executed,
            summary,
        )
        return AdapterResult(
            success=True,
            run_id=run_id,
            output_path=str(excel.resolve()),
            output_filename="Estimation_Output.xlsx",
            summary=summary,
            warnings=warnings,
            stages_run=stages_run,
            duration_s=duration,
            engine_root=engine_root,
            t1_executed=t1_executed,
        )
    except subprocess.TimeoutExpired as exc:
        logger.exception("Pipeline timeout run_id=%s", run_id)
        raise AdapterError(
            "Engineering pipeline timed out. Please try again with a smaller drawing set."
        ) from exc
    finally:
        clear_r2a_gn_pointer()
        logger.info("Run artefacts retained staging=%s", staging)
