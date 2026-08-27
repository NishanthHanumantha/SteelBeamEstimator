#!/usr/bin/env python3
"""W.14 Galera GF production upload/poll/download. Does not print secrets."""
from __future__ import annotations

import json
import os
import sys
import time
import urllib.error
import urllib.request
import zipfile
from io import BytesIO
from pathlib import Path

BASE = os.environ.get("STEEL_W14_BASE", "http://13.127.104.99").rstrip("/")
GALERA = Path(r"C:\Users\nishanth.h\SteelBeamEstimator\Test_Input\2nd Set Drawings-Galera_GF")
OUT = Path(os.environ.get("STEEL_W14_DOWNLOAD_DIR", r"C:\Users\nishanth.h\AppData\Local\Temp\w14_downloads"))
POLL_S = float(os.environ.get("STEEL_W14_POLL_S", "15"))
TIMEOUT_S = float(os.environ.get("STEEL_W14_TIMEOUT_S", "15000"))


def get(path: str, timeout: float = 60):
    try:
        with urllib.request.urlopen(BASE + path, timeout=timeout) as resp:
            return resp.status, resp.read(), dict(resp.headers)
    except urllib.error.HTTPError as exc:
        return exc.code, exc.read(), dict(exc.headers)


def post_multipart(files: dict[str, tuple[str, bytes]], timeout: float = 300):
    boundary = "----w14Boundary"
    chunks = []
    for field, (filename, data) in files.items():
        chunks.append(f"--{boundary}\r\n".encode("ascii"))
        chunks.append(
            (
                f'Content-Disposition: form-data; name="{field}"; filename="{filename}"\r\n'
                "Content-Type: application/octet-stream\r\n\r\n"
            ).encode("ascii")
        )
        chunks.append(data)
        chunks.append(b"\r\n")
    chunks.append(f"--{boundary}--\r\n".encode("ascii"))
    body = b"".join(chunks)
    req = urllib.request.Request(
        BASE + "/api/estimate",
        data=body,
        method="POST",
        headers={"Content-Type": f"multipart/form-data; boundary={boundary}"},
    )
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            return resp.status, json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        raw = exc.read()
        try:
            payload = json.loads(raw.decode("utf-8"))
        except Exception:
            payload = {"raw": raw.decode("utf-8", "replace")}
        return exc.code, payload


def galera_files() -> dict[str, tuple[str, bytes]]:
    gn = GALERA / "general_notes" / "SE-100-R0-SH-01&SH-02(GENERAL NOTES).dxf"
    fr = GALERA / "framing" / "Galera_GF_FramingPlan.dxf"
    reinf = GALERA / "reinforcement" / "Galera_GF_BeamReinforcementDetails.dxf"
    missing = [str(p) for p in (gn, fr, reinf) if not p.is_file()]
    if missing:
        raise SystemExit("missing Galera DXF: " + "; ".join(missing))
    return {
        "general_notes": (gn.name, gn.read_bytes()),
        "framing": (fr.name, fr.read_bytes()),
        "reinforcement": (reinf.name, reinf.read_bytes()),
    }


def validate_xlsx(raw: bytes, dest: Path) -> dict:
    dest.parent.mkdir(parents=True, exist_ok=True)
    dest.write_bytes(raw)
    info = {
        "bytes": len(raw),
        "pk": raw[:2] == b"PK",
        "zip_ok": False,
        "openpyxl_ok": False,
        "path": str(dest),
    }
    try:
        with zipfile.ZipFile(BytesIO(raw)) as zf:
            names = zf.namelist()
        info["zip_ok"] = any(n.startswith("xl/") for n in names)
    except Exception as exc:
        info["zip_error"] = str(exc)
    try:
        from openpyxl import load_workbook

        wb = load_workbook(dest, read_only=True, data_only=True)
        info["openpyxl_ok"] = True
        info["sheets"] = list(wb.sheetnames)
        wb.close()
    except Exception as exc:
        info["openpyxl_error"] = str(type(exc).__name__)
    return info


def main() -> int:
    print("BASE", BASE)
    st, raw, _ = get("/health", timeout=20)
    health = json.loads(raw.decode("utf-8"))
    print("HEALTH", st, health.get("phase"), health.get("hybrid", {}).get("mode"))
    blob = json.dumps(health).lower()
    if "sk-ant-" in blob:
        print("SECRET_LEAK")
        return 2
    if health.get("phase") != "W.14":
        print("PHASE_UNEXPECTED", health.get("phase"))
        return 1
    if str((health.get("hybrid") or {}).get("mode") or "").lower() != "production":
        print("HYBRID_MODE_UNEXPECTED", (health.get("hybrid") or {}).get("mode"))
        return 1

    restore = os.environ.get("STEEL_W14_RESTORE_RUN")
    if restore:
        print("RESTORE", restore)
        st, raw, _ = get(f"/api/status/{restore}", timeout=30)
        print("RESTORE_STATUS", st, raw[:400])
        st, raw, hdr = get(f"/api/download/{restore}", timeout=120)
        print("RESTORE_DOWNLOAD", st, hdr.get("Content-Disposition"), len(raw), raw[:2])
        if st == 200:
            print("RESTORE_XLSX", validate_xlsx(raw, OUT / f"restore_{restore}.xlsx"))
        return 0 if st == 200 and raw[:2] == b"PK" else 1

    files = galera_files()
    print("GALERA_FILES", {k: (v[0], len(v[1])) for k, v in files.items()})
    st, payload = post_multipart(files)
    print("ESTIMATE", st, {k: payload.get(k) for k in ("ok", "run_id", "error", "code")})
    if st != 200 or not payload.get("ok"):
        return 1
    run_id = payload["run_id"]
    print("RUN_ID", run_id)
    deadline = time.time() + TIMEOUT_S
    last = {}
    while time.time() < deadline:
        st, raw, _ = get(f"/api/status/{run_id}", timeout=30)
        try:
            last = json.loads(raw.decode("utf-8"))
        except Exception:
            last = {"status_http": st, "raw": raw[:200].decode("utf-8", "replace")}
        status = last.get("status")
        print(
            "POLL",
            status,
            last.get("result_lifecycle"),
            last.get("download_ready"),
            last.get("message"),
            last.get("elapsed_s"),
            last.get("hybrid"),
        )
        if status in {"success", "error"}:
            break
        time.sleep(POLL_S)
    print(
        "FINAL_STATUS",
        json.dumps(
            {
                k: last.get(k)
                for k in (
                    "status",
                    "run_id",
                    "result_lifecycle",
                    "download_ready",
                    "result_registered",
                    "excel_generated",
                    "excel_exists",
                    "workbook_name",
                    "duration_s",
                    "summary",
                    "hybrid",
                    "error",
                )
            },
            default=str,
        ),
    )
    if last.get("status") != "success":
        return 1
    results = []
    for i in range(2):
        st, raw, hdr = get(f"/api/download/{run_id}", timeout=180)
        info = {
            "http": st,
            "disposition": hdr.get("Content-Disposition"),
            "content_type": hdr.get("Content-Type"),
        }
        if st == 200:
            info.update(validate_xlsx(raw, OUT / f"galera_{run_id}_{i+1}.xlsx"))
        results.append(info)
        print("DOWNLOAD", i + 1, info)
    ok = all(r.get("http") == 200 and r.get("pk") and r.get("zip_ok") for r in results)
    print("USER_DELIVERY", "PASS" if ok else "FAIL")
    print("RUN_ID", run_id)
    return 0 if ok else 1


if __name__ == "__main__":
    sys.exit(main())
