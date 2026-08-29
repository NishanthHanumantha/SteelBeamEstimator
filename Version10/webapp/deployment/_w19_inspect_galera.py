"""Inspect W.19 Galera run: real PieceGenerator extents, M.2, L.2, BBS, Excel.

Does not reconstruct 0.25L. Reads artefacts written by the live pipeline.
"""
from __future__ import annotations

import json
import math
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional

ROOT = Path(r"C:\Users\nishanth.h\SteelBeamEstimator")
FOCUS = ("B1", "B10", "B23")
OUT = ROOT / "Version10" / "webapp" / "deployment" / "W19_GALERA_LOCAL_VALIDATION.json"


def _load(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def _find_pieces(out: Path) -> List[Dict[str, Any]]:
    candidates = list(out.glob("**/reinforcement_pieces*.json")) + list(
        out.glob("**/PhaseR1_3*/**/*.json")
    )
    pieces: List[Dict[str, Any]] = []
    for p in candidates:
        try:
            data = _load(p)
        except Exception:
            continue
        rows = data if isinstance(data, list) else (
            data.get("pieces") or data.get("reinforcement_pieces") or []
        )
        if isinstance(rows, list) and rows and isinstance(rows[0], dict):
            if "piece_start_mm" in rows[0] or "piece_type" in rows[0]:
                pieces.extend(rows)
    return pieces


def _bars_for_beam(l2_model: Dict[str, Any]) -> List[Dict[str, Any]]:
    out = []
    for key in (
        "top_main_bars", "top_extra_bars", "bottom_main_bars",
        "bottom_extra_bars", "spacer_bars",
    ):
        for bar in l2_model.get(key) or []:
            if isinstance(bar, dict):
                rec = dict(bar)
                rec["_l2_key"] = key
                out.append(rec)
    return out


def _role_token(bar_id: str) -> str:
    parts = str(bar_id or "").split("-")
    if len(parts) >= 4 and parts[0] == "R13":
        return parts[2] if len(parts) == 4 else "-".join(parts[2:-1])
    return ""


def inspect_run(staging: Path) -> Dict[str, Any]:
    out = staging / "data" / "output"
    l2_path = out / "PhaseR1.3_pipeline_integration" / "beam_reinforcement_models_production.json"
    spacer_report_path = out / "PhaseR1.3_pipeline_integration" / "spacer_rule_report.json"
    excel = out / "Production_Output" / "Estimation_Output.xlsx"
    pieces_dump = out / "PhaseR1.3_pipeline_integration" / "engineering_bar_models.json"

    l2 = _load(l2_path) if l2_path.is_file() else {}
    models = {m.get("beam_id"): m for m in (l2.get("models") or []) if isinstance(m, dict)}
    spacer_report = _load(spacer_report_path) if spacer_report_path.is_file() else {}

    piece_rows = []
    if pieces_dump.is_file():
        dump = _load(pieces_dump)
        for bm in dump if isinstance(dump, list) else (dump.get("beams") or dump.get("models") or []):
            if not isinstance(bm, dict):
                continue
            bid = bm.get("beam_id")
            for bar in bm.get("bars") or []:
                meta = bar.get("engineering_metadata") or {}
                piece_rows.append({
                    "beam_id": bid,
                    "bar_role": bar.get("bar_role"),
                    "bar_label": bar.get("bar_label"),
                    "piece_id": meta.get("piece_id"),
                    "piece_type": meta.get("piece_type"),
                    "piece_start_mm": meta.get("piece_start_mm"),
                    "piece_end_mm": meta.get("piece_end_mm"),
                    "cut_length_mm": meta.get("cut_length_mm"),
                })

    bbs_info: Dict[str, Any] = {}
    excel_sheets: List[str] = []
    project_totals: Dict[str, Any] = {}
    b27 = None
    if excel.is_file():
        from openpyxl import load_workbook
        wb = load_workbook(excel, data_only=True, read_only=True)
        excel_sheets = list(wb.sheetnames)
        if "Bar Bending Schedule" in wb.sheetnames:
            ws = wb["Bar Bending Schedule"]
            rows = list(ws.iter_rows(values_only=True))
            header = None
            current_beam = None
            for row in rows:
                vals = [c for c in row]
                text = " ".join(str(c) for c in vals if c is not None)
                if not text.strip():
                    continue
                joined = " | ".join(str(c) if c is not None else "" for c in vals[:8])
                if str(vals[0] or "").startswith("Beam") or (vals[2] and str(vals[2]).startswith("B") and vals[1] is None):
                    # beam header patterns vary; catch Beam ID in description-ish cells
                    for c in vals:
                        s = str(c or "")
                        if s in FOCUS or s.split()[-1] in FOCUS if s else False:
                            current_beam = s.split()[-1] if s.split()[-1] in FOCUS else s
                for bid in FOCUS:
                    if f" {bid} " in f" {text} " or text.strip() == bid or text.startswith(f"{bid} "):
                        if "Beam" in text or text.strip() == bid:
                            current_beam = bid
                if current_beam in FOCUS and "Spacer" in text:
                    bbs_info.setdefault(current_beam, []).append({
                        "row": joined,
                        "qty": vals[5] if len(vals) > 5 else None,
                        "cut": vals[7] if len(vals) > 7 else None,
                    })
        if "Project Totals" in wb.sheetnames:
            ws = wb["Project Totals"]
            for row in ws.iter_rows(values_only=True):
                if row and row[0]:
                    project_totals[str(row[0])] = row[1] if len(row) > 1 else None
        if "Beam Summary" in wb.sheetnames:
            ws = wb["Beam Summary"]
            for row in ws.iter_rows(values_only=True):
                if row and str(row[0]) == "B27":
                    b27 = list(row)
                    break
        wb.close()

    beams = []
    hooked = {
        "B1": (5918.3, 2799.6),
        "B10": (2424.2,),
        "B23": (3710.1,),
    }
    for bid in FOCUS:
        model = models.get(bid) or {}
        bars = _bars_for_beam(model)
        spacers = [b for b in bars if b.get("_l2_key") == "spacer_bars" or _role_token(b.get("bar_id") or "") == "SPACER_BAR"]
        longs = [b for b in bars if b not in spacers]
        zones = []
        for sp in spacers:
            for z in (sp.get("zones") or []):
                zones.append(z)
            if not sp.get("zones"):
                zones.append({
                    "zone_start_mm": sp.get("zone_start_mm"),
                    "zone_end_mm": sp.get("zone_end_mm"),
                    "zone_length_mm": sp.get("zone_length_mm"),
                    "raw_quantity": sp.get("raw_quantity"),
                    "quantity": sp.get("quantity"),
                })
        beam_pieces = [p for p in piece_rows if p.get("beam_id") == bid]
        overlap_from_cut = False
        for z in zones:
            zl = z.get("zone_length_mm")
            if zl is None:
                continue
            for cut in hooked.get(bid, ()):
                if abs(float(zl) - float(cut)) < 1.0:
                    overlap_from_cut = True
        qtys = [int(s.get("quantity") or 0) for s in spacers]
        beams.append({
            "beam_id": bid,
            "geometry": model.get("geometry") or {},
            "pieces": beam_pieces,
            "longitudinal": [{
                "bar_id": b.get("bar_id"),
                "piece_id": b.get("source_bar_id"),
                "piece_type": b.get("piece_type"),
                "bar_label": b.get("bar_label"),
                "role_token": _role_token(b.get("bar_id") or ""),
                "piece_start_mm": b.get("piece_start_mm"),
                "piece_end_mm": b.get("piece_end_mm"),
                "fabrication_cut_length_mm": b.get("cut_length_mm"),
            } for b in longs],
            "l2_spacers": [{
                "quantity": s.get("quantity"),
                "cut_length_mm": s.get("cut_length_mm"),
                "diameter_mm": s.get("diameter_mm"),
                "spacing_mm": s.get("spacing_mm"),
                "zone_start_mm": s.get("zone_start_mm"),
                "zone_end_mm": s.get("zone_end_mm"),
                "zone_length_mm": s.get("zone_length_mm"),
                "raw_quantity": s.get("raw_quantity"),
                "extent_fallback": s.get("extent_fallback"),
                "zones": s.get("zones"),
                "piece_start_mm": s.get("piece_start_mm"),
                "piece_end_mm": s.get("piece_end_mm"),
            } for s in spacers],
            "overlap_zones": zones,
            "l2_spacer_quantities": qtys,
            "l2_spacer_row_count": len(spacers),
            "bbs_spacer_rows": bbs_info.get(bid) or [],
            "fabrication_cut_used_as_overlap": overlap_from_cut,
            "must_not_be_3_7_3": qtys != [3, 7, 3] and sorted(qtys) != [3, 3, 7],
        })

    return {
        "staging": str(staging),
        "l2_path": str(l2_path),
        "excel_path": str(excel) if excel.is_file() else None,
        "excel_sheets": excel_sheets,
        "project_totals": project_totals,
        "b27_beam_summary_row": b27,
        "spacer_report": spacer_report,
        "piece_extent_present_count": sum(
            1 for p in piece_rows
            if p.get("piece_start_mm") is not None and p.get("piece_end_mm") is not None
        ),
        "piece_row_count": len(piece_rows),
        "beams": beams,
    }


def main() -> int:
    if len(sys.argv) < 2:
        print("usage: _w19_inspect_galera.py <staging_dir>")
        return 2
    staging = Path(sys.argv[1])
    data = inspect_run(staging)
    OUT.write_text(json.dumps(data, indent=2, default=str), encoding="utf-8")
    print("WROTE", OUT)
    for b in data["beams"]:
        print(b["beam_id"], "l2_qty", b["l2_spacer_quantities"], "cut_as_overlap", b["fabrication_cut_used_as_overlap"], "pieces", len(b["pieces"]))
    print("SHEETS", data["excel_sheets"])
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
