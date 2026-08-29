#!/usr/bin/env python3
"""Poll production status and download XLSX. Does not print secrets."""
from __future__ import annotations

import json
import os
import re
import sys
import time
import urllib.error
import urllib.request
import zipfile
from io import BytesIO
from pathlib import Path

BASE = os.environ.get("STEEL_W19_BASE", "http://13.127.104.99").rstrip("/")
OUT = Path(os.environ.get("STEEL_W19_DOWNLOAD_DIR", r"C:\Users\nishanth.h\AppData\Local\Temp\w19_downloads"))
POLL_S = float(os.environ.get("STEEL_W19_POLL_S", "20"))
TIMEOUT_S = float(os.environ.get("STEEL_W19_TIMEOUT_S", "15000"))


def get(path: str, timeout: float = 60):
    try:
        with urllib.request.urlopen(BASE + path, timeout=timeout) as resp:
            return resp.status, resp.read(), dict(resp.headers)
    except urllib.error.HTTPError as exc:
        return exc.code, exc.read(), dict(exc.headers)


def validate_xlsx(raw: bytes, dest: Path) -> dict:
    dest.parent.mkdir(parents=True, exist_ok=True)
    dest.write_bytes(raw)
    info = {
        "bytes": len(raw),
        "pk": raw[:2] == b"PK",
        "zip_ok": False,
        "openpyxl_ok": False,
        "path": str(dest),
        "sheets": [],
    }
    try:
        with zipfile.ZipFile(BytesIO(raw)) as zf:
            names = zf.namelist()
        info["zip_ok"] = any(n.startswith("xl/") for n in names)
    except Exception as exc:
        info["zip_error"] = str(exc)
        return info
    try:
        from openpyxl import load_workbook

        wb = load_workbook(dest, read_only=True, data_only=True)
        info["openpyxl_ok"] = True
        info["sheets"] = list(wb.sheetnames)
        expected = {
            "Project Totals",
            "Steel Summary",
            "Bar Bending Schedule",
            "Diameter Summary",
            "Beam Summary",
        }
        info["five_sheets"] = expected.issubset(set(info["sheets"]))
        project = {}
        if "Project Totals" in wb.sheetnames:
            for row in wb["Project Totals"].iter_rows(values_only=True):
                if row and row[0]:
                    project[str(row[0])] = row[1] if len(row) > 1 else None
        info["project_totals"] = project
        bbs = {}
        frames = set()
        if "Bar Bending Schedule" in wb.sheetnames:
            current = None
            for row in wb["Bar Bending Schedule"].iter_rows(values_only=True):
                vals = list(row)
                desc = str(vals[2] or "").strip() if len(vals) > 2 else ""
                frame = str(vals[1] or "").strip() if len(vals) > 1 else ""
                if re.match(r"^B\d+[A-Z]?$", desc):
                    current = desc
                    if frame:
                        frames.add(frame)
                    continue
                if current and "Spacer" in desc:
                    bbs.setdefault(current, []).append({
                        "frame": frame,
                        "description": desc,
                        "dia": vals[3] if len(vals) > 3 else None,
                        "spacing": vals[4] if len(vals) > 4 else None,
                        "qty": vals[5] if len(vals) > 5 else None,
                        "cut_m": vals[7] if len(vals) > 7 else None,
                    })
        info["bbs_spacers"] = {k: bbs[k] for k in ("B1", "B10", "B23") if k in bbs}
        info["bbs_all_spacer_beams"] = sorted(bbs)
        info["frames"] = sorted(frames)
        b27 = None
        if "Steel Summary" in wb.sheetnames:
            ws = wb["Steel Summary"]
            headers = None
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 1:
                    headers = [str(c or "") for c in row]
                if row and str(row[0] or "").strip() == "B27":
                    b27 = {headers[j] if headers and j < len(headers) else str(j): row[j] for j in range(len(row))}
                    break
        if b27 is None and "Beam Summary" in wb.sheetnames:
            ws = wb["Beam Summary"]
            headers = None
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 1:
                    headers = [str(c or "") for c in row]
                if row and str(row[0] or "").strip() == "B27":
                    b27 = {headers[j] if headers and j < len(headers) else str(j): row[j] for j in range(len(row))}
                    break
        info["b27"] = b27
        wb.close()
    except Exception as exc:
        info["openpyxl_error"] = f"{type(exc).__name__}: {exc}"
    return info


def main() -> int:
    if len(sys.argv) < 2:
        print("usage: _w19_poll_download.py <run_id> [label]")
        return 2
    run_id = sys.argv[1]
    label = sys.argv[2] if len(sys.argv) > 2 else "run"
    print("BASE", BASE)
    print("RUN_ID", run_id)
    st, raw, _ = get("/health", timeout=20)
    health = json.loads(raw.decode("utf-8"))
    print("HEALTH", st, health.get("phase"), health.get("app_release"), (health.get("hybrid") or {}).get("mode"))
    if "sk-ant-" in json.dumps(health).lower():
        print("SECRET_LEAK")
        return 2
    deadline = time.time() + TIMEOUT_S
    last = {}
    while time.time() < deadline:
        st, raw, _ = get(f"/api/status/{run_id}", timeout=45)
        try:
            last = json.loads(raw.decode("utf-8"))
        except Exception:
            last = {"status_http": st, "raw": raw[:300].decode("utf-8", "replace")}
        print(
            "POLL",
            last.get("status"),
            last.get("result_lifecycle"),
            last.get("download_ready"),
            last.get("message"),
            last.get("elapsed_s"),
            last.get("hybrid"),
        )
        if last.get("status") in {"success", "error"}:
            break
        time.sleep(POLL_S)
    print("FINAL_STATUS", json.dumps({
        k: last.get(k)
        for k in (
            "status", "run_id", "result_lifecycle", "download_ready",
            "result_registered", "excel_generated", "excel_exists",
            "workbook_name", "duration_s", "summary", "hybrid", "error",
        )
    }, default=str))
    if last.get("status") != "success":
        return 1
    results = []
    for i in range(2):
        st, raw, hdr = get(f"/api/download/{run_id}", timeout=180)
        info = {
            "http": st,
            "disposition": hdr.get("Content-Disposition"),
            "content_type": hdr.get("Content-Type"),
        }
        if st == 200:
            info.update(validate_xlsx(raw, OUT / f"{label}_{run_id}_{i+1}.xlsx"))
        results.append(info)
        print("DOWNLOAD", i + 1, json.dumps(info, default=str)[:4000])
    ok = all(r.get("http") == 200 and r.get("pk") and r.get("zip_ok") and r.get("openpyxl_ok") for r in results)
    print("USER_DELIVERY", "PASS" if ok else "FAIL")
    print("RUN_ID", run_id)
    return 0 if ok else 1


if __name__ == "__main__":
    sys.exit(main())
