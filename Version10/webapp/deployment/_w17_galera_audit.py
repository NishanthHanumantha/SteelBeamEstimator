"""W.17 read-only Galera B1/B10/B23 calculation trace. Does not mutate engineering logic."""
from __future__ import annotations

import importlib.util as ilu
import json
import math
import sys
import types
from pathlib import Path

ENGINE = Path(r"C:\Users\nishanth.h\SteelBeamEstimator\Version10\src")
R2A = ENGINE / "PhaseR.2A_engineering_context"
VB1 = ENGINE / "PhaseVB.1_production_output_completion"
GN = Path(r"C:\Users\nishanth.h\AppData\Local\Temp\w16_gn\galera_gn.dxf")
L2 = Path(r"C:\Users\nishanth.h\AppData\Local\Temp\w16_gn\galera_l2.json")
XLSX = Path(r"C:\Users\nishanth.h\SteelBeamEstimator\Version10\Downloaded_Output\W16_Galera_GF_Estimation_Output.xlsx")
OUT_JSON = Path(r"C:\Users\nishanth.h\SteelBeamEstimator\Version10\webapp\deployment\W17_GALERA_B1_B10_B23_CALCULATION_TRACE.json")
BEAMS = ("B1", "B10", "B23")
DENSITY = 7850.0


def bootstrap():
    pkg = types.ModuleType("PhaseR2A")
    pkg.__path__ = [str(R2A)]
    sys.modules["PhaseR2A"] = pkg
    for sub in [
        "engineering_context_model",
        "engineering_context_cache",
        "engineering_context_loader",
        "general_notes_text_extractor",
        "development_length_parser",
        "cover_parser",
        "steel_grade_parser",
        "concrete_grade_parser",
        "hook_rule_parser",
        "lap_rule_parser",
        "general_notes_classifier",
        "engineering_context_builder",
        "engineering_context_validator",
        "engineering_context_factory",
    ]:
        spec = ilu.spec_from_file_location(f"PhaseR2A.{sub}", R2A / f"{sub}.py")
        mod = ilu.module_from_spec(spec)
        mod.__package__ = "PhaseR2A"
        sys.modules[f"PhaseR2A.{sub}"] = mod
        spec.loader.exec_module(mod)
    sys.path.insert(0, str(VB1))


def ld_lookup(loader, dia, steel, conc):
    table = loader.context.development_length_table
    key = (steel, int(dia), conc)
    if key in table:
        ld = int(table[key])
        return {
            "used": True,
            "table_hit": True,
            "key": list(key),
            "ld_mm": ld,
            "ld_m": round(ld / 1000.0, 3),
            "ld_over_d": round(ld / float(dia), 4) if dia else None,
            "formula": "EngineeringContextLoader.get_development_length_mm: development_length_table[(steel, dia, conc)]",
            "fallback": False,
        }
    factor = loader.context.fallback_dev_length_factor or 40
    ld = int(factor * dia)
    return {
        "used": True,
        "table_hit": False,
        "key": list(key),
        "ld_mm": ld,
        "ld_m": round(ld / 1000.0, 3),
        "ld_over_d": factor,
        "formula": f"fallback: fallback_dev_length_factor ({factor}) * diameter_mm",
        "fallback": True,
    }


def unit_weight_kg_m(dia):
    area = math.pi * dia ** 2 / 4.0
    # kg/m = area_mm2 * 1m * density / 1e9
    return area * 1000.0 * DENSITY / 1e9


def main():
    bootstrap()
    from PhaseR2A.engineering_context_factory import EngineeringContextFactory
    from PhaseR2A.engineering_context_loader import EngineeringContextLoader
    from steel_weight_completion import SteelWeightCompletion
    from bbs_completion_engine import BBSCompletionEngine
    import openpyxl

    ctx, passed, warns = EngineeringContextFactory.create(GN, "GALERA", force_rebuild=True)
    loader = EngineeringContextLoader(ctx)
    summary_meta = loader.summary()
    steel = loader.get_steel_grade()
    conc = loader.get_concrete_grade("BEAM")
    cover = loader.get_cover("BEAM")
    hook = loader.get_hook_multiple(135)

    l2 = json.loads(L2.read_text(encoding="utf-8"))
    models = {m["beam_id"]: m for m in l2["models"]}
    sw = SteelWeightCompletion(L2, loader=loader)
    proj = sw.compute()
    bbs = BBSCompletionEngine(proj, frame_type="GF").generate()

    wb = openpyxl.load_workbook(XLSX, data_only=True)
    excel_bbs = []
    cur = None
    for row in wb["Bar Bending Schedule"].iter_rows(values_only=True):
        if not row:
            continue
        desc = row[2]
        if isinstance(desc, str) and desc in BEAMS and row[3] == 1:
            cur = desc
            excel_bbs.append({"beam_id": cur, "kind": "header", "row": list(row[:17])})
            continue
        if cur and isinstance(desc, str) and desc.startswith("B") and row[3] == 1:
            cur = None
            if desc in BEAMS:
                cur = desc
                excel_bbs.append({"beam_id": cur, "kind": "header", "row": list(row[:17])})
            continue
        if cur in BEAMS:
            excel_bbs.append({
                "beam_id": cur,
                "kind": "line",
                "frame": row[1],
                "description": row[2],
                "dia_mm": row[3],
                "spacing_m": row[4],
                "qty": row[5],
                "dvlp_m": row[6],
                "cut_m": row[7],
                "total_m": row[8],
                "y8": row[9], "y10": row[10], "y12": row[11], "y16": row[12],
                "y20": row[13], "y25": row[14], "y32": row[15],
                "total_kg": row[16],
            })
    excel_ss = {}
    for row in wb["Steel Summary"].iter_rows(values_only=True):
        if row and row[0] in BEAMS:
            dias = [float(row[i] or 0) for i in range(1, 8)]
            excel_ss[row[0]] = {
                "y8": dias[0], "y10": dias[1], "y12": dias[2], "y16": dias[3],
                "y20": dias[4], "y25": dias[5], "y32": dias[6],
                "diameter_sum": round(sum(dias), 3),
                "total": float(row[8] or 0),
            }
    wb.close()

    beams_out = {}
    for bid in BEAMS:
        model = models[bid]
        geom = model.get("geometry") or {}
        bw = next(b for b in proj.beam_weights if b.beam_id == bid)
        l2_bars = []
        for key in (
            "top_main_bars", "top_extra_bars", "bottom_main_bars", "bottom_extra_bars",
            "side_face_reinforcement", "stirrups", "spacer_bars", "supplementary_bars",
            "chair_bars", "development_length_regions", "continuity_regions",
        ):
            for bar in model.get(key) or []:
                if not isinstance(bar, dict):
                    continue
                l2_bars.append({
                    "l2_key": key,
                    "bar_id": bar.get("bar_id"),
                    "bar_label": bar.get("bar_label"),
                    "diameter_mm": bar.get("diameter_mm"),
                    "quantity": bar.get("quantity"),
                    "spacing_mm": bar.get("spacing_mm"),
                    "cut_length_mm": bar.get("cut_length_mm"),
                    "steel_grade": bar.get("steel_grade"),
                    "development_length_mm": bar.get("development_length_mm"),
                    "cover_mm": bar.get("cover_mm"),
                    "zone": bar.get("zone"),
                    "source_phase": bar.get("source_phase"),
                    "engineering_metadata": bar.get("engineering_metadata") or {},
                })

        computed = []
        for bar in bw.bar_weights:
            dia = float(bar.diameter_mm)
            role = bar.role
            ld = None
            family = role
            cut_formula = bar.cut_length_source
            inputs = {}
            span = float(bw.span_mm or 0)
            width = float(bw.width_mm or 0)
            depth = float(bw.depth_mm or 0)
            if role in ("TOP_MAIN", "BOTTOM_MAIN", "TOP_EXTRA", "BOTTOM_EXTRA", "SIDE_FACE", "BENT", "CRANKED", "DEVELOPMENT", "LAP"):
                ld = ld_lookup(loader, dia, steel, conc)
                cut_formula = "cut_mm = clear_span_mm + 2 * Ld_mm"
                inputs = {
                    "clear_span_mm": span,
                    "Ld_mm": ld["ld_mm"],
                    "two_Ld_mm": 2 * ld["ld_mm"],
                }
                expected_cut = span + 2 * ld["ld_mm"] if span > 0 else 2 * ld["ld_mm"]
            elif role == "SPACER":
                ld = {"used": False, "reason": "spacer uses width - 2*cover, not Ld"}
                expected_cut = width - 2 * cover
                cut_formula = "cut_mm = width_mm - 2 * cover_mm  (or provided_cut_length_mm if set)"
                inputs = {"width_mm": width, "cover_mm": cover, "provided_cut_length_mm": None}
            elif role == "STIRRUP":
                ld = {"used": False, "reason": "stirrup Dvlp.L in Excel is hook allowance 2*N*d, not TABLE 1 Ld"}
                perim = 2 * (width - 2 * cover) + 2 * (depth - 2 * cover)
                hook_mm = 2 * hook * dia
                expected_cut = perim + hook_mm
                cut_formula = "cut_mm = 2*(width-2*cover)+2*(depth-2*cover)+2*hook_multiple*d"
                inputs = {
                    "width_mm": width, "depth_mm": depth, "cover_mm": cover,
                    "hook_multiple": hook, "hook_mm": hook_mm, "perimeter_mm": perim,
                    "defaults_if_missing": "depth default 600 if None; width default 200 if None",
                }
            else:
                ld = ld_lookup(loader, dia, steel, conc)
                expected_cut = span + 2 * ld["ld_mm"] if span > 0 else 1000.0
                cut_formula = "default longitudinal: span + 2*Ld or 1000mm DEFAULT_FALLBACK"
                inputs = {"clear_span_mm": span}

            uw = unit_weight_kg_m(dia)
            cut_m = bar.cut_length_mm / 1000.0
            total_m = cut_m * bar.quantity
            wt = uw * total_m
            excel_dvlp = None
            if role == "STIRRUP":
                excel_dvlp_formula = "SI.1: dvlp_m = 2 * hook_multiple * d / 1000  (hook allowance, NOT Ld)"
                excel_dvlp = 2 * hook * dia / 1000.0
            else:
                excel_dvlp_formula = "BBSCompletionEngine: dvlp_m = cut_m - span_m   (= 2*Ld/1000 for longitudinal; NEGATIVE for spacers)"
                excel_dvlp = cut_m - span / 1000.0 if span else None

            computed.append({
                "bar_id": bar.bar_id,
                "bar_label": bar.bar_label,
                "role": role,
                "family": family,
                "diameter_mm": dia,
                "quantity": bar.quantity,
                "steel_grade_on_bar": bar.steel_grade,
                "cut_length_mm": bar.cut_length_mm,
                "cut_length_source": bar.cut_length_source,
                "area_mm2": bar.area_mm2,
                "weight_per_bar_kg": bar.weight_per_bar_kg,
                "total_weight_kg": bar.total_weight_kg,
                "formula_used": bar.formula_used,
                "development_length": ld,
                "cut_length_formula": cut_formula,
                "cut_length_inputs": inputs,
                "expected_cut_mm_from_current_formula": expected_cut,
                "cut_matches_formula": abs(bar.cut_length_mm - expected_cut) <= 0.6,
                "geometry": {
                    "field_used": "geometry.clear_span_mm" if role != "SPACER" and role != "STIRRUP" else (
                        "geometry.width_mm and geometry.depth_mm" if role == "STIRRUP" else "geometry.width_mm"
                    ),
                    "clear_span_mm": span,
                    "width_mm": width,
                    "depth_mm": depth,
                    "cover_mm": cover,
                },
                "total_length": {
                    "formula": "total_m = quantity * cut_length_mm / 1000",
                    "quantity": bar.quantity,
                    "cut_m": round(cut_m, 6),
                    "total_m": round(total_m, 6),
                },
                "weight": {
                    "unit_weight_formula": "uw_kg_m = (pi * d^2 / 4) * 1000 * 7850 / 1e9",
                    "unit_weight_kg_m": uw,
                    "total_m": round(total_m, 6),
                    "formula": "W = total_m * uw_kg_m   equivalently (pi*d^2/4)*cut_mm*qty*7850/1e9",
                    "recomputed_kg": wt,
                    "engine_kg": bar.total_weight_kg,
                    "match": abs(wt - bar.total_weight_kg) < 1e-6,
                },
                "excel_dvlp_m_display": excel_dvlp,
                "excel_dvlp_m_formula": excel_dvlp_formula,
                "code_path": {
                    "weight": "SteelWeightCompletion._compute_bar / _derive_cut_length" if role != "STIRRUP" else "StirrupImprover.compute_beam + StirrupWeightEngine.cut_length_mm then SteelWeightCompletion SI.1 branch",
                    "bbs": "BBSCompletionEngine.generate" if role != "STIRRUP" else "BBSCompletionEngine.generate re-invokes StirrupImprover.compute_beam per STIRRUP BarSteelWeight",
                    "quantity": "L.2 bar.quantity (directly extracted)" if role != "STIRRUP" else "StirrupQuantityEngine.calculate (spacing-derived) or L.2 quantity for legacy",
                },
            })

        bbs_lines = [r for r in bbs if r.beam_id == bid and not r.is_beam_header]
        excel_lines = [r for r in excel_bbs if r["beam_id"] == bid and r["kind"] == "line"]
        bbs_sum = sum(float(r.total_weight_kg or 0) for r in bbs_lines)
        steel_sum = sum(bw.bar_weights[i].total_weight_kg for i in range(len(bw.bar_weights)))
        ss = excel_ss.get(bid, {})

        beams_out[bid] = {
            "geometry": {
                "l2_geometry": geom,
                "field_used_by_vb1": "geometry.clear_span_mm  ->  SteelWeightCompletion._compute_beam span_mm",
                "clear_span_mm": geom.get("clear_span_mm"),
                "width_mm": geom.get("width_mm"),
                "depth_mm": geom.get("depth_mm"),
                "source_on_model": geom.get("source") or geom.get("geometry_source") or "UNTRACEABLE_beyond_L2_geometry_object",
            },
            "l2_bar_records": l2_bars,
            "computed_bar_weights": computed,
            "bbs_engine_rows": [
                {
                    "description": r.description,
                    "frame": r.frame_type,
                    "diameter_mm": r.diameter_mm,
                    "spacing_m": r.spacing_m,
                    "quantity": r.quantity,
                    "dvlp_length_m": r.dvlp_length_m,
                    "cut_length_m": r.cut_length_m,
                    "total_length_m": r.total_length_m,
                    "total_weight_kg": r.total_weight_kg,
                    "weight_d8": r.weight_d8, "weight_d10": r.weight_d10,
                    "weight_d12": r.weight_d12, "weight_d16": r.weight_d16,
                    "weight_d20": r.weight_d20, "weight_d25": r.weight_d25,
                    "weight_d32": r.weight_d32,
                }
                for r in bbs_lines
            ],
            "excel_rows": excel_lines,
            "reconciliation": {
                "n_l2_bars": len(l2_bars),
                "n_computed": len(computed),
                "n_bbs_lines": len(bbs_lines),
                "n_excel_lines": len(excel_lines),
                "computed_weight_sum_kg": round(steel_sum, 6),
                "bbs_weight_sum_kg": round(bbs_sum, 6),
                "steel_summary_diameter_sum_kg": ss.get("diameter_sum"),
                "steel_summary_total_kg": ss.get("total"),
                "weight_by_diameter": {str(k): round(v, 6) for k, v in bw.weight_by_diameter.items() if v},
                "beam_total_kg": round(bw.total_weight_kg, 6),
                "bbs_vs_computed_diff": round(bbs_sum - steel_sum, 6),
                "ss_total_vs_diameter_diff": round((ss.get("total") or 0) - (ss.get("diameter_sum") or 0), 6),
            },
        }

    # Ld table snapshot for used diameters
    dias_used = sorted({int(b["diameter_mm"]) for bid in BEAMS for b in beams_out[bid]["computed_bar_weights"]})
    ld_grid = {}
    for d in dias_used:
        info = ld_lookup(loader, d, steel, conc)
        ld_grid[str(d)] = info

    out = {
        "dataset": "2nd Set — Galera GF",
        "beams": list(BEAMS),
        "phase": "W.17 READ-ONLY CALCULATION AUDIT",
        "production_mutation": "NO",
        "project_engineering_context": {
            "frame": "GF",
            "cover_mm": cover,
            "cover_source": summary_meta.get("cover_source"),
            "steel_grade": steel,
            "concrete_grade_beam": conc,
            "hook_multiple_135": hook,
            "dev_length_factor_representative_dia12": summary_meta.get("dev_length_factor"),
            "dev_length_source": summary_meta.get("dev_length_source"),
            "gn_dxf_path": summary_meta.get("gn_dxf_path"),
            "parse_confidence": summary_meta.get("parse_confidence"),
            "loader_validation_passed": passed,
            "density_kg_m3": DENSITY,
            "ld_by_diameter_for_project_grades": ld_grid,
            "note": "Excel Project Totals prints representative Ld/d at dia=12 only. Per-bar Ld is table lookup (steel, dia, conc).",
        },
        "pipeline_map": [
            "DXF (framing + reinforcement + GN)",
            "PhaseR1_2A GeometryProvider -> L.2 geometry.clear_span_mm / width_mm / depth_mm",
            "PhaseR1.3 EngineeringBarBuilder -> L.2 bar records",
            "SteelWeightCompletion._compute_beam/_compute_bar/_derive_cut_length",
            "StirrupImprover.compute_beam (SI.1) for stirrups",
            "BBSCompletionEngine.generate",
            "EstimatorExcelGenerator (Bar Bending Schedule + Steel Summary)",
        ],
        "bbs_header_row_note": "Beam header reuses Dia=1 (member count), Spacing=span_m, No.of Bars=width_mm, Dvlp.L=depth_m. These are presentation fields, not bar calculations.",
        "beams_detail": beams_out,
    }
    OUT_JSON.write_text(json.dumps(out, indent=2, default=str), encoding="utf-8")
    print("wrote", OUT_JSON)
    for bid in BEAMS:
        r = beams_out[bid]["reconciliation"]
        print(bid, "lines excel", r["n_excel_lines"], "bbs", r["n_bbs_lines"], "computed", r["n_computed"],
              "ss", r["steel_summary_total_kg"], "bbs_sum", r["bbs_weight_sum_kg"],
              "diff_ss", r["ss_total_vs_diameter_diff"], "bbs_vs_comp", r["bbs_vs_computed_diff"])


if __name__ == "__main__":
    main()
