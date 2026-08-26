"""Public W.12 post-deploy checks. No secrets."""
from __future__ import annotations

import io
import json
import urllib.error
import urllib.request
import zipfile
from pathlib import Path

BASE = "http://13.127.104.99"
OUT = Path(r"C:\Users\nishanth.h\AppData\Local\Temp\w12_downloads")
RID = "20260826_084708_f74912b8"


def get(path: str, timeout: float = 60):
    try:
        with urllib.request.urlopen(BASE + path, timeout=timeout) as resp:
            return resp.status, resp.read(), dict(resp.headers)
    except urllib.error.HTTPError as exc:
        return exc.code, exc.read(), dict(exc.headers)


def main() -> None:
    st, raw, _ = get("/health")
    health = json.loads(raw.decode("utf-8"))
    print("PUBLIC_HEALTH", st, health.get("phase"), health.get("app_release"))
    print("HYBRID", health.get("hybrid", {}).get("mode"))
    print("RESULT_DELIVERY", health.get("result_delivery"))
    blob = json.dumps(health).lower()
    print("SECRET_LEAK", "sk-ant-" in blob)

    st, html, _ = get("/")
    text = html.decode("utf-8", "replace")
    print("HOME", st)
    print("HAS_BTN", 'id="btn-download"' in text)
    print("HAS_BUTTON_TYPE", 'id="btn-download">Download Excel</button>' in text)
    print("HAS_HREF_HASH", 'href="#"' in text and "btn-download" in text)
    print("HAS_895", "8.9.5" in text)

    st, raw, _ = get(f"/api/status/{RID}")
    print("OLD_STATUS_HTTP", st)
    print("OLD_STATUS", raw[:800].decode("utf-8", "replace"))

    st, raw, hdr = get(f"/api/download/{RID}", timeout=120)
    print("OLD_DL", st, hdr.get("Content-Disposition"), len(raw), raw[:2])
    st2, raw2, _ = get(f"/api/download/{RID}", timeout=120)
    print("OLD_DL2", st2, len(raw2), raw == raw2)

    if st == 200 and raw[:2] == b"PK":
        OUT.mkdir(parents=True, exist_ok=True)
        dest = OUT / f"restore_{RID}.xlsx"
        dest.write_bytes(raw)
        with zipfile.ZipFile(io.BytesIO(raw)) as zf:
            names = zf.namelist()
        print("ZIP_OK", any(n.startswith("xl/") for n in names), "BYTES", len(raw))
        try:
            from openpyxl import load_workbook

            wb = load_workbook(dest, read_only=True, data_only=True)
            print("OPENPYXL", True, wb.sheetnames[:10])
            wb.close()
        except Exception as exc:
            print("OPENPYXL_FAIL", type(exc).__name__)

    st, raw, _ = get("/api/download/does-not-exist")
    print("INVALID", st, raw.decode("utf-8", "replace")[:200])
    st, raw, _ = get("/api/download/../../../etc/passwd")
    print("TRAVERSAL", st, raw[:80])


if __name__ == "__main__":
    main()
