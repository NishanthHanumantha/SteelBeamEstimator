"""Phase W.19.1 — bind R.2A artefacts into Excel Project Totals."""
from __future__ import annotations

import importlib.util as _ilu
import json
import sys
import tempfile
import types
import unittest
from pathlib import Path

WEBAPP_ROOT = Path(__file__).resolve().parents[1]
ENGINE_SRC = WEBAPP_ROOT.parent / "src"
VB1 = ENGINE_SRC / "PhaseVB.1_production_output_completion"


def _bootstrap_vb1() -> None:
    if str(VB1) not in sys.path:
        sys.path.insert(0, str(VB1))
    if "PhaseVB1" not in sys.modules:
        pkg = types.ModuleType("PhaseVB1")
        pkg.__path__ = [str(VB1)]
        sys.modules["PhaseVB1"] = pkg
    spec = _ilu.spec_from_file_location(
        "PhaseVB1.phase_vb1_orchestrator", VB1 / "phase_vb1_orchestrator.py"
    )
    mod = _ilu.module_from_spec(spec)
    mod.__package__ = "PhaseVB1"
    sys.modules["PhaseVB1.phase_vb1_orchestrator"] = mod
    spec.loader.exec_module(mod)


_bootstrap_vb1()
from PhaseVB1.phase_vb1_orchestrator import loader_summary_from_r2a_artefacts  # noqa: E402
from production_output_models import (  # noqa: E402
    BeamSteelWeight,
    DiameterSummary,
    ProjectSteelSummary,
)
from estimator_excel_generator import EstimatorExcelGenerator  # noqa: E402


def _tiny_summary() -> ProjectSteelSummary:
    bw = BeamSteelWeight(
        beam_id="B1",
        beam_name="B1",
        span_mm=4000,
        depth_mm=450,
        width_mm=230,
        bar_weights=[],
        total_weight_kg=10.0,
        weight_by_diameter={16: 10.0},
    )
    return ProjectSteelSummary(
        total_weight_kg=10.0,
        beam_weights=[bw],
        diameter_summary=[
            DiameterSummary(
                diameter_mm=16,
                total_bars=1,
                total_length_mm=1000,
                total_weight_kg=10.0,
                weight_fraction=1.0,
            )
        ],
        total_bars=1,
        total_beams=1,
    )


def _project_totals(path: Path) -> dict:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    values = {
        str(row[0]): row[1]
        for row in wb["Project Totals"].iter_rows(values_only=True)
        if row and row[0]
    }
    sheets = list(wb.sheetnames)
    wb.close()
    return values, sheets


def _write_r2a_artefacts(output_dir: Path, **overrides) -> Path:
    r2a = output_dir.parent / "PhaseR.2A_engineering_context"
    r2a.mkdir(parents=True, exist_ok=True)
    summary = {
        "cover_beam_mm": 30,
        "primary_steel_grade": "Fe550",
        "dev_length_factor_d": 50,
        "concrete_grade_beam": "M30",
        "steel_density_kg_m3": 7850.0,
        "gn_dxf": "SE-100.dxf",
    }
    summary.update(overrides.pop("summary", {}))
    ctx = {
        "cover_rules": [
            {
                "element": "BEAM IN SUPERSTRUCTURE",
                "cover_mm": 30,
                "concrete_grade": "M30",
                "steel_grade": "Fe550",
                "source": "GN_DXF_TABLE_2",
            }
        ],
        "development_length_table": {"Fe550_dia12__M30": 600},
        "primary_steel_grade": "Fe550",
    }
    ctx.update(overrides.pop("context", {}))
    (r2a / "engineering_context_summary.json").write_text(
        json.dumps(summary), encoding="utf-8"
    )
    (r2a / "engineering_context.json").write_text(json.dumps(ctx), encoding="utf-8")
    return r2a


class W191ExcelMetadataBindingTests(unittest.TestCase):
    def test_artefact_maps_cover_and_ld_summary(self):
        with tempfile.TemporaryDirectory() as td:
            out = Path(td) / "Production_Output"
            out.mkdir()
            _write_r2a_artefacts(out)
            mapped = loader_summary_from_r2a_artefacts(out)
        self.assertIsNotNone(mapped)
        self.assertEqual(mapped["cover_beam_mm"], 30)
        self.assertEqual(mapped["cover_source"], "GN_DXF_TABLE_2")
        self.assertEqual(mapped["dev_length_factor"], 50)
        self.assertEqual(mapped["dev_length_source"], "GN_DXF_TABLE_1")
        self.assertEqual(mapped["primary_steel_grade"], "Fe550")

    def test_excel_project_totals_consume_r2a_artefacts(self):
        with tempfile.TemporaryDirectory() as td:
            out = Path(td) / "Production_Output"
            out.mkdir()
            _write_r2a_artefacts(out)
            mapped = loader_summary_from_r2a_artefacts(out)
            gen = EstimatorExcelGenerator(
                bbs_rows=[],
                steel_summary=_tiny_summary(),
                output_dir=out,
                loader_summary=mapped,
            )
            paths = gen.generate()
            values, sheets = _project_totals(paths["production"])
        self.assertEqual(values.get("Cover"), 30)
        self.assertEqual(values.get("Development Length"), "GN table (Fe550, ~50d)")
        self.assertNotIn("UNRESOLVED", str(values.get("Cover", "")))
        self.assertNotIn("UNRESOLVED", str(values.get("Development Length", "")))
        self.assertEqual(
            sheets,
            [
                "Beam Summary",
                "Bar Bending Schedule",
                "Steel Summary",
                "Diameter Summary",
                "Project Totals",
            ],
        )

    def test_missing_artefacts_keep_unresolved_fallback(self):
        with tempfile.TemporaryDirectory() as td:
            out = Path(td) / "Production_Output"
            out.mkdir()
            self.assertIsNone(loader_summary_from_r2a_artefacts(out))
            gen = EstimatorExcelGenerator(
                bbs_rows=[],
                steel_summary=_tiny_summary(),
                output_dir=out,
                loader_summary=None,
            )
            paths = gen.generate()
            values, _ = _project_totals(paths["production"])
        self.assertIn("UNRESOLVED", str(values.get("Cover", "")))
        self.assertIn("IS456 fallback", str(values.get("Cover", "")))
        self.assertIn("UNRESOLVED", str(values.get("Development Length", "")))
        self.assertNotEqual(values.get("Cover"), 30)

    def test_does_not_hardcode_galera_or_inizio(self):
        src = (VB1 / "phase_vb1_orchestrator.py").read_text(encoding="utf-8")
        lower = src.lower()
        self.assertNotIn("if galera", lower)
        self.assertNotIn("if inizio", lower)
        self.assertNotIn("cover = 30", lower)


if __name__ == "__main__":
    unittest.main()
