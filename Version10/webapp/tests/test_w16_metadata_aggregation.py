"""Phase W.16 — project metadata, B27 aggregation, drawing-specific frame."""
from __future__ import annotations

import importlib.util as _ilu
import json
import math
import os
import sys
import tempfile
import types
import unittest
from pathlib import Path

WEBAPP_ROOT = Path(__file__).resolve().parents[1]
ENGINE_SRC = WEBAPP_ROOT.parent / "src"
VB1 = ENGINE_SRC / "PhaseVB.1_production_output_completion"
R2A = ENGINE_SRC / "PhaseR.2A_engineering_context"
VROOT = ENGINE_SRC / "PhaseVROOT.1_dynamic_pipeline_initialization"
HYBRID = ENGINE_SRC / "PhaseP2610D1_vision_semantic_contract_hybrid_foundation"


def _bootstrap_package(pkg_name: str, directory: Path, subs: list[str]) -> None:
    if pkg_name not in sys.modules:
        pkg = types.ModuleType(pkg_name)
        pkg.__path__ = [str(directory)]
        sys.modules[pkg_name] = pkg
    for sub in subs:
        key = f"{pkg_name}.{sub}"
        if key in sys.modules:
            continue
        spec = _ilu.spec_from_file_location(key, directory / f"{sub}.py")
        mod = _ilu.module_from_spec(spec)
        mod.__package__ = pkg_name
        sys.modules[key] = mod
        spec.loader.exec_module(mod)


_bootstrap_package(
    "PhaseR2A",
    R2A,
    [
        "engineering_context_model",
        "engineering_context_cache",
        "engineering_context_loader",
        "general_notes_text_extractor",
        "development_length_parser",
        "cover_parser",
        "steel_grade_parser",
        "concrete_grade_parser",
        "hook_rule_parser",
        "lap_rule_parser",
        "general_notes_classifier",
        "engineering_context_builder",
        "engineering_context_validator",
        "engineering_context_factory",
    ],
)
_bootstrap_package("w16_hybrid", HYBRID, ["config", "normalize"])
if str(VB1) not in sys.path:
    sys.path.insert(0, str(VB1))
if str(VROOT) not in sys.path:
    sys.path.insert(0, str(VROOT))

from PhaseR2A.engineering_context_factory import EngineeringContextFactory  # noqa: E402
from PhaseR2A.engineering_context_loader import EngineeringContextLoader  # noqa: E402
from PhaseR2A.engineering_context_model import CoverRule, EngineeringContext  # noqa: E402
from w16_hybrid.normalize import parse_diameter  # noqa: E402
from project_discovery import ProjectDiscovery  # noqa: E402
from production_output_models import (  # noqa: E402
    BeamSteelWeight,
    DiameterSummary,
    ProjectSteelSummary,
)
from bbs_completion_engine import BBSCompletionEngine  # noqa: E402
from estimator_excel_generator import EstimatorExcelGenerator  # noqa: E402
from steel_weight_completion import (  # noqa: E402
    SteelWeightCompletion,
    canonicalize_bar_diameter_mm,
    reconcile_beam_weight,
)
from workbook_validator import WorkbookValidator  # noqa: E402

PREVIOUS_B27_TOTAL = 42173.254
PREVIOUS_CORRUPT_ROW_KG = 20915.324
_DENSITY = 7850.0


def _ctx(**kwargs) -> EngineeringContext:
    defaults = dict(
        gn_dxf_path="gn.dxf",
        project_id="TEST",
        parsed_at="2026-08-28T00:00:00",
        steel_grades=("Fe550",),
        primary_steel_grade="Fe550",
        concrete_grades=("M30",),
        development_length_table={("Fe550", 12, "M30"): 480},
        cover_rules=(),
        hook_rules=(),
        lap_rules=(),
        spacer_rules=(),
        code_references=(),
        warnings=(),
        parse_confidence=0.9,
        fallback_dev_length_factor=40,
        fallback_cover_mm=40,
        fallback_steel_grade="Fe415",
        fallback_concrete_grade="M25",
    )
    defaults.update(kwargs)
    return EngineeringContext(**defaults)


def _longitudinal_weight(dia_mm: float, qty: int, span_mm: float, ld_factor: int = 40) -> float:
    ld = ld_factor * dia_mm
    cut = span_mm + 2.0 * ld
    area = math.pi * dia_mm ** 2 / 4.0
    return area * cut * _DENSITY / 1e9 * qty


def _l2_b27_anomaly(path: Path, extra_valid: bool = True) -> Path:
    """Reproduce the Inizio B27 252 mm extra-top pattern (span 6550 mm, qty 2)."""
    extra = {
        "bar_id": "b27-xt",
        "diameter_mm": 252,
        "quantity": 2,
        "bar_label": "2-Y25",
        "steel_grade": "Y",
    }
    model = {
        "beam_id": "B27",
        "beam_name": "B27",
        "geometry": {"clear_span_mm": 6550, "depth_mm": 450, "width_mm": 230},
        "top_main_bars": [],
        "bottom_main_bars": [],
        "top_extra_bars": [dict(extra, bar_id="b27-xt-a"), dict(extra, bar_id="b27-xt-b")],
        "bottom_extra_bars": [],
        "side_face_reinforcement": [],
        "stirrups": [],
        "spacer_bars": [],
        "supplementary_bars": [],
        "chair_bars": [],
        "development_length_regions": [],
        "continuity_regions": [],
    }
    if extra_valid:
        model["top_main_bars"] = [
            {"bar_id": "tm", "diameter_mm": 16, "quantity": 3, "bar_label": "3Y16", "steel_grade": "Y"}
        ]
        model["bottom_main_bars"] = [
            {"bar_id": "bm", "diameter_mm": 20, "quantity": 2, "bar_label": "2Y20", "steel_grade": "Y"}
        ]
    path.write_text(json.dumps({"models": [model]}), encoding="utf-8")
    return path


def _tiny_summary(beam_id: str = "B1", total: float = 10.0) -> ProjectSteelSummary:
    bw = BeamSteelWeight(
        beam_id=beam_id,
        beam_name=beam_id,
        span_mm=6550,
        depth_mm=450,
        width_mm=230,
        bar_weights=[],
        total_weight_kg=total,
        weight_by_diameter={16: total},
    )
    return ProjectSteelSummary(
        total_weight_kg=total,
        beam_weights=[bw],
        diameter_summary=[
            DiameterSummary(diameter_mm=16, total_bars=1, total_length_mm=1000, total_weight_kg=total, weight_fraction=1.0)
        ],
        total_bars=1,
        total_beams=1,
    )


class W16GeneralNotesTests(unittest.TestCase):
    def test_cover_and_ld_from_project_gn_table_not_universal_40(self):
        ctx = _ctx(
            cover_rules=(
                CoverRule(
                    element_type="BEAM IN SUPERSTRUCTURE",
                    cover_mm=30,
                    concrete_grade="M30",
                    steel_grade="Fe550",
                    source="GN_DXF_TABLE_2",
                ),
            ),
            development_length_table={("Fe550", 12, "M30"): 456},
            primary_steel_grade="Fe550",
            fallback_concrete_grade="M30",
        )
        loader = EngineeringContextLoader(ctx)
        self.assertEqual(loader.get_cover("BEAM"), 30)
        self.assertEqual(loader.get_development_length_factor(), 38)
        summary = loader.summary()
        self.assertEqual(summary["cover_beam_mm"], 30)
        self.assertEqual(summary["cover_source"], "GN_DXF_TABLE_2")
        self.assertEqual(summary["dev_length_source"], "GN_DXF_TABLE_1")
        self.assertNotEqual(summary["cover_beam_mm"], 40)
        self.assertFalse(str(summary["cover_source"]).startswith("FALLBACK"))

    def test_fallback_cover_and_ld_are_labelled_unresolved(self):
        ctx = _ctx(cover_rules=(), development_length_table={}, primary_steel_grade="")
        loader = EngineeringContextLoader(ctx)
        summary = loader.summary()
        self.assertTrue(str(summary["cover_source"]).startswith("FALLBACK"))
        self.assertTrue(str(summary["dev_length_source"]).startswith("FALLBACK"))

        with tempfile.TemporaryDirectory() as td:
            out = Path(td)
            gen = EstimatorExcelGenerator(
                bbs_rows=[],
                steel_summary=_tiny_summary(),
                output_dir=out,
                loader_summary=None,
            )
            paths = gen.generate()
            import openpyxl
            wb = openpyxl.load_workbook(paths["production"], data_only=True)
            values = {
                str(row[0]): row[1]
                for row in wb["Project Totals"].iter_rows(values_only=True)
                if row and row[0]
            }
            wb.close()
            self.assertIn("UNRESOLVED", str(values.get("Cover", "")))
            self.assertIn("UNRESOLVED", str(values.get("Development Length", "")))
            self.assertNotEqual(values.get("Cover"), 40)

    def test_excel_project_totals_show_resolved_gn_cover(self):
        with tempfile.TemporaryDirectory() as td:
            out = Path(td)
            gen = EstimatorExcelGenerator(
                bbs_rows=[],
                steel_summary=_tiny_summary(),
                output_dir=out,
                loader_summary={
                    "cover_beam_mm": 30,
                    "cover_source": "GN_DXF_TABLE_2",
                    "dev_length_factor": 38,
                    "dev_length_source": "GN_DXF_TABLE_1",
                    "primary_steel_grade": "Fe550",
                    "steel_density": 7850.0,
                },
            )
            paths = gen.generate()
            import openpyxl
            wb = openpyxl.load_workbook(paths["production"], data_only=True)
            values = {
                str(row[0]): row[1]
                for row in wb["Project Totals"].iter_rows(values_only=True)
                if row and row[0]
            }
            names = wb.sheetnames
            wb.close()
            self.assertEqual(values.get("Cover"), 30)
            self.assertIn("Fe550", str(values.get("Development Length", "")))
            self.assertNotIn("UNRESOLVED", str(values.get("Development Length", "")))
            self.assertEqual(
                names,
                [
                    "Beam Summary",
                    "Bar Bending Schedule",
                    "Steel Summary",
                    "Diameter Summary",
                    "Project Totals",
                ],
            )

    def test_gn_discovery_prefers_steel_run_root(self):
        with tempfile.TemporaryDirectory() as td:
            run = Path(td) / "run"
            gn_dir = run / "general_notes"
            gn_dir.mkdir(parents=True)
            gn = gn_dir / "project_notes.dxf"
            gn.write_text("0\nSECTION\n2\nENTITIES\n0\nENDSEC\n0\nEOF\n", encoding="ascii")
            old = os.environ.get("STEEL_RUN_ROOT")
            os.environ["STEEL_RUN_ROOT"] = str(run)
            try:
                found = EngineeringContextFactory._discover_gn_path(Path(td) / "missing_engine")
            finally:
                if old is None:
                    os.environ.pop("STEEL_RUN_ROOT", None)
                else:
                    os.environ["STEEL_RUN_ROOT"] = old
            self.assertEqual(found, gn)


class W16FrameTests(unittest.TestCase):
    def test_galera_filenames_resolve_gf_not_hardcoded_project(self):
        with tempfile.TemporaryDirectory() as td:
            folder = Path(td) / "2nd Set Drawings-Galera_GF"
            (folder / "framing").mkdir(parents=True)
            (folder / "reinforcement").mkdir()
            (folder / "framing" / "Galera_GF_FramingPlan.dxf").write_text("0\nEOF\n")
            (folder / "reinforcement" / "Galera_GF_BeamReinforcementDetails.dxf").write_text("0\nEOF\n")
            floor = ProjectDiscovery().discover(folder)["floor"]
            self.assertEqual(floor, "GF")

    def test_inizio_range_resolves_11_18f_before_typical_floor(self):
        with tempfile.TemporaryDirectory() as td:
            folder = Path(td) / "6th Set Drawings-Inizio_11-18F"
            (folder / "framing").mkdir(parents=True)
            (folder / "reinforcement").mkdir()
            (folder / "framing" / "11-18TH FLOOR.dxf FRAMINIG.dxf").write_text("0\nEOF\n")
            (folder / "reinforcement" / "TYPICAL FLOOR BEAM (11-18).dxf").write_text("0\nEOF\n")
            floor = ProjectDiscovery().discover(folder)["floor"]
            self.assertEqual(floor, "11-18F")

    def test_production_sanitized_filenames_resolve_frame(self):
        with tempfile.TemporaryDirectory() as td:
            galera = Path(td) / "galera_run"
            (galera / "framing").mkdir(parents=True)
            (galera / "reinforcement").mkdir()
            (galera / "framing" / "Galera_GF_FramingPlan.dxf").write_text("0\nEOF\n")
            (galera / "reinforcement" / "Galera_GF_BeamReinforcementDetails_SpreadOut.dxf").write_text("0\nEOF\n")
            self.assertEqual(ProjectDiscovery().discover(galera)["floor"], "GF")
            inizio = Path(td) / "inizio_run"
            (inizio / "framing").mkdir(parents=True)
            (inizio / "reinforcement").mkdir()
            (inizio / "framing" / "11-18TH_FLOOR.dxf_FRAMINIG.dxf").write_text("0\nEOF\n")
            (inizio / "reinforcement" / "479_SE-228_TYPICAL_FLOOR_BEAM_REINFORCEMENT_DETAILS11-18_R0_SH-01_TO_SH-03.dxf").write_text("0\nEOF\n")
            self.assertEqual(ProjectDiscovery().discover(inizio)["floor"], "11-18F")

    def test_unresolved_floor_is_explicit_not_silent_tf(self):
        with tempfile.TemporaryDirectory() as td:
            folder = Path(td) / "unknown_set"
            folder.mkdir()
            (folder / "notes.txt").write_text("no floor token")
            floor = ProjectDiscovery().discover(folder)["floor"]
            self.assertEqual(floor, "UNKNOWN_FLOOR")
        engine = BBSCompletionEngine(_tiny_summary(), frame_type=floor)
        rows = engine.generate()
        self.assertTrue(rows)
        self.assertEqual(rows[0].frame_type, "UNRESOLVED")
        self.assertNotEqual(rows[0].frame_type, "TF")

    def test_bbs_rows_use_resolved_frame(self):
        for frame in ("GF", "11-18F"):
            rows = BBSCompletionEngine(_tiny_summary(), frame_type=frame).generate()
            self.assertTrue(all(r.frame_type == frame for r in rows))


_GN_FIXTURE = Path(os.environ.get("TEMP") or os.environ.get("TMP") or ".") / "w16_gn"


@unittest.skipUnless((_GN_FIXTURE / "galera_gn.dxf").is_file(), "Galera GN DXF not staged")
class W16LiveGeneralNotesTests(unittest.TestCase):
    def test_galera_cover_from_table_2_not_universal_40(self):
        ctx, passed, _ = EngineeringContextFactory.create(
            _GN_FIXTURE / "galera_gn.dxf", "GALERA", force_rebuild=True
        )
        self.assertTrue(passed)
        loader = EngineeringContextLoader(ctx)
        summary = loader.summary()
        self.assertEqual(summary["cover_beam_mm"], 30)
        self.assertEqual(summary["cover_source"], "GN_DXF_TABLE_2")
        self.assertEqual(summary["primary_steel_grade"], "Fe550")
        self.assertEqual(summary["dev_length_source"], "GN_DXF_TABLE_1")
        self.assertNotEqual(summary["dev_length_factor"], 40)

    def test_inizio_cover_from_table_2(self):
        ctx, passed, _ = EngineeringContextFactory.create(
            _GN_FIXTURE / "inizio_gn.dxf", "INIZIO", force_rebuild=True
        )
        self.assertTrue(passed)
        loader = EngineeringContextLoader(ctx)
        summary = loader.summary()
        self.assertEqual(summary["cover_beam_mm"], 30)
        self.assertEqual(summary["cover_source"], "GN_DXF_TABLE_2")
        self.assertEqual(summary["dev_length_source"], "GN_DXF_TABLE_1")


class W16B27AggregationTests(unittest.TestCase):
    def test_canonicalize_y252_to_y25(self):
        self.assertEqual(canonicalize_bar_diameter_mm(252, "2-Y25"), 25)
        self.assertEqual(canonicalize_bar_diameter_mm(252, "2Y252"), 25)
        self.assertEqual(parse_diameter("2-Y25"), 25)
        self.assertEqual(parse_diameter("2Y252"), 25)
        self.assertEqual(parse_diameter("Y12"), 12)
        self.assertIsNone(canonicalize_bar_diameter_mm(7, ""))

    def test_b27_252_pattern_cannot_recreate_42173(self):
        with tempfile.TemporaryDirectory() as td:
            l2 = _l2_b27_anomaly(Path(td) / "models.json")
            summary = SteelWeightCompletion(l2).compute()
        b27 = next(bw for bw in summary.beam_weights if bw.beam_id == "B27")
        ok, diam_sum = reconcile_beam_weight(b27.total_weight_kg, b27.weight_by_diameter)
        self.assertTrue(ok)
        self.assertAlmostEqual(b27.total_weight_kg, diam_sum, places=3)

        extra_w = [
            bar.total_weight_kg
            for bar in b27.bar_weights
            if bar.role == "TOP_EXTRA"
        ]
        self.assertEqual(len(extra_w), 2)
        for w in extra_w:
            self.assertLess(w, 200.0)
            self.assertGreater(w, 50.0)
            self.assertNotAlmostEqual(w, PREVIOUS_CORRUPT_ROW_KG, delta=100.0)

        expected_extra = _longitudinal_weight(25, 2, 6550.0) * 2
        expected_main = _longitudinal_weight(16, 3, 6550.0) + _longitudinal_weight(20, 2, 6550.0)
        expected_total = expected_extra + expected_main
        self.assertAlmostEqual(b27.total_weight_kg, expected_total, delta=0.05)
        self.assertNotAlmostEqual(b27.total_weight_kg, PREVIOUS_B27_TOTAL, delta=1000.0)
        self.assertLess(b27.total_weight_kg, 1000.0)
        self.assertAlmostEqual(summary.total_weight_kg, b27.total_weight_kg, places=3)
        self.assertAlmostEqual(
            summary.total_weight_kg,
            sum(ds.total_weight_kg for ds in summary.diameter_summary),
            places=3,
        )
        dias = {int(b.diameter_mm) for b in b27.bar_weights}
        self.assertNotIn(252, dias)
        self.assertIn(25, dias)

    def test_inizio_workbook_validated_rows_plus_repaired_extra(self):
        """342.605 is the sum of already-valid diameters; repaired extras add Y25."""
        previous_diameter_subtotal = 147.661 + 49.434 + 40.198 + 105.312
        self.assertAlmostEqual(previous_diameter_subtotal, 342.605, places=3)
        previous_difference = PREVIOUS_B27_TOTAL - previous_diameter_subtotal
        self.assertAlmostEqual(previous_difference, 41830.649, delta=0.01)
        repaired_extra = _longitudinal_weight(25, 2, 6550.0) * 2
        corrected_total = previous_diameter_subtotal + repaired_extra
        self.assertAlmostEqual(corrected_total - previous_diameter_subtotal, repaired_extra, places=6)
        self.assertLess(abs(corrected_total - (previous_diameter_subtotal + repaired_extra)), 1e-9)

    def test_unsupported_y28_is_excluded_and_beam_still_reconciles(self):
        """Inizio B137 pattern: dia 28 has no Y28 column; must not inflate beam total."""
        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / "models.json"
            path.write_text(json.dumps({
                "models": [{
                    "beam_id": "B137",
                    "beam_name": "B137",
                    "geometry": {"clear_span_mm": 2645, "depth_mm": 800, "width_mm": 200},
                    "top_main_bars": [
                        {"bar_id": "t", "diameter_mm": 20, "quantity": 2, "bar_label": "2Y20"}
                    ],
                    "bottom_main_bars": [
                        {"bar_id": "b", "diameter_mm": 28, "quantity": 2, "bar_label": "2Y28"}
                    ],
                    "top_extra_bars": [],
                    "bottom_extra_bars": [],
                    "side_face_reinforcement": [],
                    "stirrups": [],
                    "spacer_bars": [],
                    "supplementary_bars": [],
                    "chair_bars": [],
                    "development_length_regions": [],
                    "continuity_regions": [],
                }]
            }), encoding="utf-8")
            bw = SteelWeightCompletion(path).compute().beam_weights[0]
        ok, diam_sum = reconcile_beam_weight(bw.total_weight_kg, bw.weight_by_diameter)
        self.assertTrue(ok)
        self.assertAlmostEqual(bw.total_weight_kg, diam_sum, places=3)
        self.assertNotIn(28, [int(b.diameter_mm) for b in bw.bar_weights])
        self.assertIn(20, [int(b.diameter_mm) for b in bw.bar_weights])

    def test_workbook_aggregation_invariants_and_b27_not_42173(self):
        with tempfile.TemporaryDirectory() as td:
            l2 = _l2_b27_anomaly(Path(td) / "models.json")
            summary = SteelWeightCompletion(l2).compute()
            engine = BBSCompletionEngine(summary, frame_type="11-18F")
            rows = engine.generate()
            self.assertTrue(all(r.frame_type == "11-18F" for r in rows))
            gen = EstimatorExcelGenerator(
                bbs_rows=rows,
                steel_summary=summary,
                output_dir=Path(td) / "out",
                loader_summary={
                    "cover_beam_mm": 30,
                    "cover_source": "GN_DXF_TABLE_2",
                    "dev_length_factor": 40,
                    "dev_length_source": "GN_DXF_TABLE_1",
                    "primary_steel_grade": "Fe550",
                },
            )
            paths = gen.generate()
            result = WorkbookValidator(paths["production"], expected_steel_total_kg=summary.total_weight_kg).validate()
            self.assertTrue(result.validation_passed, result.validation_errors)
            import openpyxl
            wb = openpyxl.load_workbook(paths["production"], data_only=True)
            ss = list(wb["Steel Summary"].iter_rows(values_only=True))
            b27_row = next(r for r in ss if r and r[0] == "B27")
            diam_sum = sum(float(b27_row[i] or 0) for i in range(1, 8))
            total = float(b27_row[8])
            frames = {
                r[1]
                for r in wb["Bar Bending Schedule"].iter_rows(values_only=True)
                if r and r[1] in ("11-18F", "GF", "TF", "UNRESOLVED")
            }
            wb.close()
            self.assertAlmostEqual(total, diam_sum, places=2)
            self.assertNotAlmostEqual(total, PREVIOUS_B27_TOTAL, delta=1000)
            self.assertEqual(frames, {"11-18F"})


if __name__ == "__main__":
    unittest.main()
