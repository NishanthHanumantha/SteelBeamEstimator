"""
Worksheet Formatter — Phase V.B.1 MODULE 5

Applies estimator-matching formatting:
  - Merged headers, engineering titles, column widths
  - Borders, alignment, number formatting, steel units, notes
"""
from typing import Optional

try:
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# ── Colour palette (matching estimator dark-blue header style) ───────────────
COLOUR_TITLE_BG    = "1F3864"   # dark navy
COLOUR_TITLE_FG    = "FFFFFF"   # white
COLOUR_HEADER_BG   = "2E75B6"   # estimator blue
COLOUR_HEADER_FG   = "FFFFFF"   # white
COLOUR_BEAM_BG     = "D6E4F0"   # light blue (beam header rows)
COLOUR_BEAM_FG     = "1F3864"   # dark navy text
COLOUR_ALT_BG      = "EBF3FB"   # alternate row light
COLOUR_TOTAL_BG    = "FFF2CC"   # yellow total rows
COLOUR_TOTAL_FG    = "1F3864"
COLOUR_DIAM_BG     = "E2EFDA"   # green tint for diameter block
COLOUR_BORDER      = "9DC3E6"   # soft border


def _make_fill(hex_colour: str):
    if not OPENPYXL_AVAILABLE:
        return None
    return PatternFill(
        start_color=hex_colour, end_color=hex_colour, fill_type="solid"
    )


def _make_border(style: str = "thin"):
    if not OPENPYXL_AVAILABLE:
        return None
    s = Side(style=style, color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def _apply_cell(ws, row: int, col: int, value=None,
                bold=False, italic=False,
                bg: Optional[str] = None, fg: Optional[str] = None,
                align_h: str = "center", align_v: str = "center",
                wrap: bool = False, size: int = 10,
                number_format: Optional[str] = None,
                border: bool = True):
    if not OPENPYXL_AVAILABLE:
        return
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(
        bold=bold, italic=italic, color=fg or "000000", size=size
    )
    cell.alignment = Alignment(
        horizontal=align_h, vertical=align_v, wrap_text=wrap
    )
    if bg:
        cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    if border:
        cell.border = _make_border("thin")
    if number_format:
        cell.number_format = number_format


class WorksheetFormatter:
    """Applies consistent formatting to each worksheet."""

    # ── Title row ────────────────────────────────────────────────────────────

    def write_title(self, ws, text: str, ncols: int, row: int = 1) -> None:
        if not OPENPYXL_AVAILABLE:
            return
        ws.merge_cells(
            start_row=row, start_column=1, end_row=row, end_column=ncols
        )
        _apply_cell(ws, row, 1, value=text, bold=True, size=12,
                    bg=COLOUR_TITLE_BG, fg=COLOUR_TITLE_FG,
                    align_h="center", border=False)

    # ── Header row ───────────────────────────────────────────────────────────

    def write_header_row(self, ws, headers, row: int, col_widths=None) -> None:
        if not OPENPYXL_AVAILABLE:
            return
        for ci, header in enumerate(headers, start=1):
            _apply_cell(ws, row, ci, value=header, bold=True, size=10,
                        bg=COLOUR_HEADER_BG, fg=COLOUR_HEADER_FG,
                        align_h="center", border=True)
        if col_widths:
            for ci, w in enumerate(col_widths, start=1):
                ws.column_dimensions[get_column_letter(ci)].width = w

    # ── Beam header row ──────────────────────────────────────────────────────

    def write_beam_header(self, ws, row: int, ncols: int, values: list) -> None:
        if not OPENPYXL_AVAILABLE:
            return
        for ci, v in enumerate(values, start=1):
            _apply_cell(ws, row, ci, value=v, bold=True, size=10,
                        bg=COLOUR_BEAM_BG, fg=COLOUR_BEAM_FG,
                        align_h="center", border=True)

    # ── Data row ─────────────────────────────────────────────────────────────

    def write_data_row(self, ws, row: int, values: list, alt: bool = False) -> None:
        if not OPENPYXL_AVAILABLE:
            return
        bg = COLOUR_ALT_BG if alt else None
        for ci, v in enumerate(values, start=1):
            fmt = None
            if isinstance(v, float):
                fmt = "#,##0.000"
            _apply_cell(ws, row, ci, value=v, bg=bg,
                        align_h="center", number_format=fmt)

    # ── Total row ────────────────────────────────────────────────────────────

    def write_total_row(self, ws, row: int, label: str,
                        values: list, label_col: int = 1,
                        ncols: int = None) -> None:
        if not OPENPYXL_AVAILABLE:
            return
        _apply_cell(ws, row, label_col, value=label, bold=True,
                    bg=COLOUR_TOTAL_BG, fg=COLOUR_TOTAL_FG)
        for ci, v in enumerate(values, start=label_col + 1):
            fmt = "#,##0.000" if isinstance(v, float) else None
            _apply_cell(ws, row, ci, value=v, bold=True,
                        bg=COLOUR_TOTAL_BG, fg=COLOUR_TOTAL_FG,
                        number_format=fmt)

    # ── Freeze panes & row height ─────────────────────────────────────────

    def freeze_header(self, ws, freeze_row: int = 3, freeze_col: int = 1) -> None:
        if not OPENPYXL_AVAILABLE:
            return
        from openpyxl.utils import get_column_letter
        ws.freeze_panes = ws.cell(row=freeze_row, column=freeze_col)

    def set_row_heights(self, ws, title_row: int = 1, header_row: int = 2) -> None:
        if not OPENPYXL_AVAILABLE:
            return
        ws.row_dimensions[title_row].height = 24
        ws.row_dimensions[header_row].height = 20

    # ── Engineering notes block ──────────────────────────────────────────────

    def write_notes_block(self, ws, notes: list, start_row: int,
                          ncols: int = 3) -> None:
        if not OPENPYXL_AVAILABLE:
            return
        for ri, note in enumerate(notes, start=start_row):
            ws.merge_cells(
                start_row=ri, start_column=1,
                end_row=ri, end_column=ncols
            )
            _apply_cell(ws, ri, 1, value=note,
                        align_h="left", italic=True,
                        size=9, border=False)
