"""
Estimator Excel Generator — Phase V.B.1 MODULE 6

Generates the FINAL ESTIMATION OUTPUT workbook:
  - Estimation_Output.xlsx    (production)
  - Engineering_Review.xlsx   (engineering review)
  - Archive/Estimation_Output_YYYYMMDD_HHMMSS.xlsx

Worksheet layout matches the estimator reference workbook as closely as possible.
"""
import math
import pathlib
import shutil
from datetime import datetime
from typing import List, Dict, Optional

try:
    import openpyxl
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

from production_output_models import BBSRow, ProjectSteelSummary, BeamSteelWeight
from excel_structure_builder import (
    WS_PROJECT_HEADER, WS_GENERAL_NOTES, WS_BEAM_SUMMARY,
    WS_BBS, WS_STEEL_SUMMARY, WS_DIAM_SUMMARY, WS_PROJECT_TOTALS,
    BBS_COLUMNS, BEAM_SUMMARY_COLUMNS, DIAM_SUMMARY_COLUMNS,
    STEEL_SUMMARY_COLUMNS, PROJECT_TOTALS_COLUMNS,
    PROJECT_HEADER_DATA, GENERAL_NOTES,
    build_project_header_data, build_general_notes,
    ALL_WORKSHEETS,
)
from worksheet_formatter import (
    WorksheetFormatter,
    COLOUR_TITLE_BG, COLOUR_HEADER_BG, COLOUR_BEAM_BG,
    COLOUR_TOTAL_BG, COLOUR_DIAM_BG,
    _apply_cell,
)

_SUPPORTED_DIAMETERS = [8, 10, 12, 16, 20, 25, 32]


class EstimatorExcelGenerator:
    """
    Generates all three workbook variants from BBS rows + steel summary.
    """

    def __init__(
        self,
        bbs_rows: List[BBSRow],
        steel_summary: ProjectSteelSummary,
        output_dir: pathlib.Path,
        loader_summary: Optional[dict] = None,
    ) -> None:
        self.bbs_rows = bbs_rows
        self.steel_summary = steel_summary
        self.output_dir = output_dir
        self.archive_dir = output_dir / "Archive"
        self.fmt = WorksheetFormatter()
        self._loader_summary = loader_summary
        self._header_data = build_project_header_data(loader_summary)
        self._general_notes = build_general_notes(loader_summary)

    # ── public ──────────────────────────────────────────────────────────────

    def generate(self) -> Dict[str, pathlib.Path]:
        """
        Returns {
            'production': Path,
            'engineering_review': Path,
            'archive': Path
        }
        """
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.archive_dir.mkdir(parents=True, exist_ok=True)

        prod_path = self._build_workbook_safe(
            self.output_dir / "Estimation_Output.xlsx", review_mode=False
        )

        review_path = self._build_workbook_safe(
            self.output_dir / "Engineering_Review.xlsx", review_mode=True
        )

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        archive_path = self.archive_dir / f"Estimation_Output_{ts}.xlsx"
        try:
            shutil.copy2(str(prod_path), str(archive_path))
        except (PermissionError, OSError):
            archive_path = prod_path

        return {
            "production": prod_path,
            "engineering_review": review_path,
            "archive": archive_path,
        }

    # ── workbook builder ─────────────────────────────────────────────────────

    def _build_workbook_safe(self, path: pathlib.Path, review_mode: bool) -> pathlib.Path:
        """
        Tries to write to `path`. If the file is locked, writes to a
        timestamped fallback and tries to replace the original silently.
        Returns the path where the file was actually written.
        """
        try:
            self._build_workbook(path, review_mode)
            return path
        except (PermissionError, OSError):
            ts = datetime.now().strftime("%H%M%S")
            fallback = path.with_name(f"{path.stem}_{ts}{path.suffix}")
            print(f"      Note: {path.name} is open in Excel; SI.1 output -> {fallback.name}")
            self._build_workbook(fallback, review_mode)
            try:
                import os
                os.replace(str(fallback), str(path))
                return path
            except (PermissionError, OSError):
                return fallback

    def _build_workbook(self, path: pathlib.Path, review_mode: bool) -> None:
        if not OPENPYXL_OK:
            raise ImportError("openpyxl is required. Install with: pip install openpyxl")

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # remove default sheet

        # Revised production format: omit Project Header + General Notes
        self._ws_beam_summary(wb)
        self._ws_bbs(wb)
        self._ws_steel_summary(wb)
        self._ws_diameter_summary(wb)
        self._ws_project_totals(wb, review_mode=review_mode)

        wb.save(str(path))

    # ── WORKSHEET 1 — Project Header ─────────────────────────────────────────

    def _ws_project_header(self, wb) -> None:
        ws = wb.create_sheet(WS_PROJECT_HEADER)
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 40
        ncols = 2
        self.fmt.write_title(ws, "ESTIMATION OUTPUT — PROJECT HEADER", ncols, row=1)
        self.fmt.write_header_row(ws, ["Field", "Value"], row=2)
        for ri, (k, v) in enumerate(self._header_data.items(), start=3):
            _apply_cell(ws, ri, 1, value=k, bold=True, align_h="left")
            _apply_cell(ws, ri, 2, value=v, align_h="left")
        ts_row = len(self._header_data) + 3
        _apply_cell(ws, ts_row, 1, value="Generated At", bold=True, align_h="left")
        _apply_cell(ws, ts_row, 2,
                    value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"), align_h="left")

    # ── WORKSHEET 2 — General Notes ──────────────────────────────────────────

    def _ws_general_notes(self, wb) -> None:
        ws = wb.create_sheet(WS_GENERAL_NOTES)
        ws.column_dimensions["A"].width = 80
        self.fmt.write_title(ws, "GENERAL NOTES — ENGINEERING ASSUMPTIONS", 1, row=1)
        self.fmt.write_notes_block(ws, self._general_notes, start_row=2, ncols=1)

    # ── WORKSHEET 3 — Beam Summary ───────────────────────────────────────────

    def _ws_beam_summary(self, wb) -> None:
        ws = wb.create_sheet(WS_BEAM_SUMMARY)
        headers = [h for h, _ in BEAM_SUMMARY_COLUMNS]
        widths  = [w for _, w in BEAM_SUMMARY_COLUMNS]
        ncols = len(headers)
        self.fmt.write_title(ws, "BEAM SUMMARY", ncols, row=1)
        self.fmt.write_header_row(ws, headers, row=2, col_widths=widths)
        self.fmt.freeze_header(ws, freeze_row=3, freeze_col=1)

        for ri, bw in enumerate(self.steel_summary.beam_weights, start=3):
            dw = bw.weight_by_diameter
            row_vals = [
                bw.beam_id,
                round(bw.span_mm / 1000, 3) if bw.span_mm else None,
                round(bw.width_mm / 1000, 3) if bw.width_mm else None,
                round(bw.depth_mm / 1000, 3) if bw.depth_mm else None,
                len(bw.bar_weights),
                round(bw.total_weight_kg, 3),
            ] + [round(dw.get(d, 0.0), 3) for d in _SUPPORTED_DIAMETERS]
            self.fmt.write_data_row(ws, ri, row_vals, alt=(ri % 2 == 0))

        # Totals row
        total_row = len(self.steel_summary.beam_weights) + 3
        diam_totals = [
            round(sum(bw.weight_by_diameter.get(d, 0.0)
                      for bw in self.steel_summary.beam_weights), 3)
            for d in _SUPPORTED_DIAMETERS
        ]
        self.fmt.write_total_row(
            ws, total_row, "PROJECT TOTAL",
            [None, None, None, self.steel_summary.total_beams,
             round(self.steel_summary.total_weight_kg, 3)] + diam_totals,
            label_col=1,
        )

    # ── WORKSHEET 4 — Bar Bending Schedule ──────────────────────────────────

    def _ws_bbs(self, wb) -> None:
        ws = wb.create_sheet(WS_BBS)
        headers   = [h for h, _, _ in BBS_COLUMNS]
        widths    = [w for _, _, w in BBS_COLUMNS]
        attr_keys = [a for _, a, _ in BBS_COLUMNS]
        ncols = len(headers)

        self.fmt.write_title(ws, "BAR BENDING SCHEDULE", ncols, row=1)
        self.fmt.write_header_row(ws, headers, row=2, col_widths=widths)
        self.fmt.freeze_header(ws, freeze_row=3, freeze_col=3)

        ri = 3
        alt = False
        for bbs_row in self.bbs_rows:
            vals = [getattr(bbs_row, k, None) for k in attr_keys]
            if bbs_row.is_beam_header:
                self.fmt.write_beam_header(ws, ri, ncols, vals)
            else:
                self.fmt.write_data_row(ws, ri, vals, alt=alt)
                alt = not alt
            ri += 1

        # Diameter totals footer
        self._write_bbs_totals(ws, ri)

    def _write_bbs_totals(self, ws, start_row: int) -> None:
        diam_totals = {d: 0.0 for d in _SUPPORTED_DIAMETERS}
        for row in self.bbs_rows:
            if not row.is_beam_header:
                for d, attr in zip(
                    _SUPPORTED_DIAMETERS,
                    ["weight_d8","weight_d10","weight_d12","weight_d16",
                     "weight_d20","weight_d25","weight_d32"]
                ):
                    v = getattr(row, attr, None)
                    if v:
                        diam_totals[d] += v

        total_kg = sum(diam_totals.values())

        # Header
        _apply_cell(ws, start_row, 1, value="DIAMETER TOTALS",
                    bold=True, bg=COLOUR_DIAM_BG, fg="1F3864")
        _apply_cell(ws, start_row + 1, 1, value="Diameter (mm)",
                    bold=True, bg=COLOUR_DIAM_BG)
        _apply_cell(ws, start_row + 1, 2, value="Total Weight (kg)",
                    bold=True, bg=COLOUR_DIAM_BG)

        for offset, d in enumerate(_SUPPORTED_DIAMETERS, start=2):
            _apply_cell(ws, start_row + offset, 1, value=f"Y{d}", bg=COLOUR_DIAM_BG)
            _apply_cell(ws, start_row + offset, 2,
                        value=round(diam_totals[d], 3),
                        number_format="#,##0.000", bg=COLOUR_DIAM_BG)

        total_offset = 2 + len(_SUPPORTED_DIAMETERS)
        _apply_cell(ws, start_row + total_offset, 1, value="TOTAL",
                    bold=True, bg=COLOUR_TOTAL_BG)
        _apply_cell(ws, start_row + total_offset, 2,
                    value=round(total_kg, 3),
                    bold=True, number_format="#,##0.000", bg=COLOUR_TOTAL_BG)

    # ── WORKSHEET 5 — Steel Summary ──────────────────────────────────────────

    def _ws_steel_summary(self, wb) -> None:
        ws = wb.create_sheet(WS_STEEL_SUMMARY)
        headers = [h for h, _ in STEEL_SUMMARY_COLUMNS]
        widths  = [w for _, w in STEEL_SUMMARY_COLUMNS]
        ncols = len(headers)
        self.fmt.write_title(ws, "STEEL WEIGHT SUMMARY — PER BEAM", ncols, row=1)
        self.fmt.write_header_row(ws, headers, row=2, col_widths=widths)
        self.fmt.freeze_header(ws, freeze_row=3, freeze_col=1)

        for ri, bw in enumerate(self.steel_summary.beam_weights, start=3):
            dw = bw.weight_by_diameter
            row_vals = (
                [bw.beam_id]
                + [round(dw.get(d, 0.0), 3) for d in _SUPPORTED_DIAMETERS]
                + [round(bw.total_weight_kg, 3)]
            )
            self.fmt.write_data_row(ws, ri, row_vals, alt=(ri % 2 == 0))

        # Totals
        total_row = len(self.steel_summary.beam_weights) + 3
        diam_totals = [
            round(sum(bw.weight_by_diameter.get(d, 0.0)
                      for bw in self.steel_summary.beam_weights), 3)
            for d in _SUPPORTED_DIAMETERS
        ]
        self.fmt.write_total_row(
            ws, total_row, "PROJECT TOTAL",
            diam_totals + [round(self.steel_summary.total_weight_kg, 3)],
            label_col=1,
        )

    # ── WORKSHEET 6 — Diameter Summary ──────────────────────────────────────

    def _ws_diameter_summary(self, wb) -> None:
        ws = wb.create_sheet(WS_DIAM_SUMMARY)
        headers = [h for h, _ in DIAM_SUMMARY_COLUMNS]
        widths  = [w for _, w in DIAM_SUMMARY_COLUMNS]
        ncols = len(headers)
        self.fmt.write_title(ws, "STEEL SUMMARY — PER DIAMETER", ncols, row=1)
        self.fmt.write_header_row(ws, headers, row=2, col_widths=widths)

        for ri, ds in enumerate(self.steel_summary.diameter_summary, start=3):
            self.fmt.write_data_row(ws, ri, [
                f"Y{ds.diameter_mm}",
                ds.total_bars,
                round(ds.total_length_mm, 1),
                round(ds.total_weight_kg, 3),
                round(ds.weight_fraction * 100, 2),
            ], alt=(ri % 2 == 0))

        total_row = len(self.steel_summary.diameter_summary) + 3
        self.fmt.write_total_row(ws, total_row, "TOTAL", [
            self.steel_summary.total_bars,
            None,
            round(self.steel_summary.total_weight_kg, 3),
            100.0,
        ], label_col=1)

    # ── WORKSHEET 7 — Project Totals ─────────────────────────────────────────

    def _ws_project_totals(self, wb, review_mode: bool = False) -> None:
        ws = wb.create_sheet(WS_PROJECT_TOTALS)
        headers = [h for h, _ in PROJECT_TOTALS_COLUMNS]
        widths  = [w for _, w in PROJECT_TOTALS_COLUMNS]
        ncols = len(headers)
        title = ("ENGINEERING REVIEW — PROJECT TOTALS"
                 if review_mode else "FINAL ESTIMATION OUTPUT — PROJECT TOTALS")
        self.fmt.write_title(ws, title, ncols, row=1)
        self.fmt.write_header_row(ws, headers, row=2, col_widths=widths)

        ls = self._loader_summary or {}
        density = ls.get("steel_density", 7850.0)
        cover = ls.get("cover_beam_mm", 40)
        dl_factor = ls.get("dev_length_factor", 40)
        sg = ls.get("primary_steel_grade", "Fe415")
        cover_source = ls.get("cover_source") or "UNRESOLVED"
        dl_source = ls.get("dev_length_source") or "UNRESOLVED"
        if not ls:
            cover_value = f"UNRESOLVED (IS456 fallback {cover} mm)"
            dl_value = f"UNRESOLVED (IS456 fallback {sg}, ~{dl_factor}d)"
        elif str(cover_source).startswith("FALLBACK") or str(cover_source) == "UNRESOLVED":
            cover_value = f"UNRESOLVED ({cover_source} {cover} mm)"
            dl_value = (
                f"UNRESOLVED ({dl_source} {sg}, ~{dl_factor}d)"
                if str(dl_source).startswith("FALLBACK") or str(dl_source) == "UNRESOLVED"
                else f"GN table ({sg}, ~{dl_factor}d)"
            )
        else:
            cover_value = cover
            dl_value = f"GN table ({sg}, ~{dl_factor}d)"
        totals_data = [
            ("Total Beams",            self.steel_summary.total_beams,        "beams"),
            ("Total Bars",             self.steel_summary.total_bars,         "bars"),
            ("Total Steel Weight",     round(self.steel_summary.total_weight_kg, 3), "kg"),
            ("Steel Density",          density,                               "kg/m³"),
            ("Development Length",     dl_value,                              ""),
            ("Cover",                  cover_value,                           "mm" if isinstance(cover_value, (int, float)) else ""),
        ]
        for d in self.steel_summary.diameter_summary:
            totals_data.append((
                f"Y{d.diameter_mm} — Weight",
                round(d.total_weight_kg, 3),
                "kg",
            ))

        for ri, (item, val, unit) in enumerate(totals_data, start=3):
            self.fmt.write_data_row(ws, ri, [item, val, unit], alt=(ri % 2 == 0))

        if review_mode:
            note_row = len(totals_data) + 4
            ws.merge_cells(
                start_row=note_row, start_column=1, end_row=note_row, end_column=3
            )
            _apply_cell(ws, note_row, 1,
                        value="⚠ Engineering Review Workbook — Not for construction use",
                        bold=True, italic=True, fg="FF0000", border=False)
