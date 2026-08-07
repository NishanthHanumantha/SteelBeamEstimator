"""
Write QA.3.1 diagnostic artefacts.
MODEL_VERSION: 10.0.1
"""
from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List

MODEL_VERSION = "10.0.1"
PHASE_ID = "QA.3.1"


def _dump(path: Path, obj: Any) -> None:
    path.write_text(json.dumps(obj, indent=2, default=str), encoding="utf-8")


def write_all(
    out_root: Path,
    beam_records: List[Dict[str, Any]],
    aggregate: Dict[str, Any],
    recommendations: Dict[str, Any],
    meta: Dict[str, Any],
) -> Dict[str, str]:
    out_root.mkdir(parents=True, exist_ok=True)
    paths: Dict[str, str] = {}

    pipeline = {
        "phase_id": PHASE_ID,
        "model_version": MODEL_VERSION,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        **meta,
        "beams": beam_records,
        "aggregate": aggregate,
    }
    p = out_root / "BeamPipelineDiagnostics.json"
    _dump(p, pipeline)
    paths["BeamPipelineDiagnostics.json"] = str(p)

    stage_cmp = {
        "phase_id": PHASE_ID,
        "beams": [
            {
                "beam_id": r["beam_id"],
                "stages": {
                    k: (v or {}).get("status")
                    for k, v in (r.get("stages") or {}).items()
                },
                "first_failing_stage": (r.get("root_cause") or {}).get(
                    "first_failing_stage"
                ),
                "primary_category": (r.get("root_cause") or {}).get("primary_category"),
            }
            for r in beam_records
        ],
    }
    p = out_root / "BeamStageComparison.json"
    _dump(p, stage_cmp)
    paths["BeamStageComparison.json"] = str(p)

    rej = []
    for r in beam_records:
        for e in ((r.get("stages") or {}).get("Ownership") or {}).get(
            "rejected_entities"
        ) or []:
            rej.append({"beam_id": r["beam_id"], **e})
    p = out_root / "OwnershipRejectionLog.json"
    _dump(p, {"phase_id": PHASE_ID, "count": len(rej), "items": rej})
    paths["OwnershipRejectionLog.json"] = str(p)

    crop = {
        "phase_id": PHASE_ID,
        "beams": [
            {"beam_id": r["beam_id"], **((r.get("stages") or {}).get("Crop Window") or {})}
            for r in beam_records
        ],
    }
    p = out_root / "CropDiagnostics.json"
    _dump(p, crop)
    paths["CropDiagnostics.json"] = str(p)

    ann = {
        "phase_id": PHASE_ID,
        "beams": [
            {
                "beam_id": r["beam_id"],
                **((r.get("stages") or {}).get("Annotation Association") or {}),
            }
            for r in beam_records
        ],
    }
    p = out_root / "AnnotationDiagnostics.json"
    _dump(p, ann)
    paths["AnnotationDiagnostics.json"] = str(p)

    rend = {
        "phase_id": PHASE_ID,
        "beams": [
            {"beam_id": r["beam_id"], **((r.get("stages") or {}).get("Rendering") or {})}
            for r in beam_records
        ],
    }
    p = out_root / "RenderDiagnostics.json"
    _dump(p, rend)
    paths["RenderDiagnostics.json"] = str(p)

    root = {
        "phase_id": PHASE_ID,
        "model_version": MODEL_VERSION,
        "hypothesis": aggregate.get("hypothesis"),
        "beams": [
            {"beam_id": r["beam_id"], **(r.get("root_cause") or {})}
            for r in beam_records
        ],
    }
    p = out_root / "RootCauseSummary.json"
    _dump(p, root)
    paths["RootCauseSummary.json"] = str(p)

    p = out_root / "FailureFrequency.json"
    _dump(
        p,
        {
            "phase_id": PHASE_ID,
            "failure_frequency": aggregate.get("failure_frequency"),
            "averages": aggregate.get("averages"),
            "beams_analysed": aggregate.get("beams_analysed"),
            "top_primary_root_cause": aggregate.get("top_primary_root_cause"),
        },
    )
    paths["FailureFrequency.json"] = str(p)

    # Markdown cards
    cards = [
        f"# Beam Diagnostic Cards — {PHASE_ID}",
        "",
        f"MODEL_VERSION: {MODEL_VERSION}",
        "",
    ]
    for r in beam_records:
        rc = r.get("root_cause") or {}
        cards += [
            f"## {r['beam_id']}",
            "",
            f"**First failing stage:** {rc.get('first_failing_stage') or 'None'}",
            f"**Primary root cause:** {rc.get('primary_category')}",
            f"**Confidence:** {rc.get('confidence')}",
            "",
            "| Stage | Status |",
            "|-------|--------|",
        ]
        for k, v in (r.get("stages") or {}).items():
            cards.append(f"| {k} | {(v or {}).get('status')} |")
        cards.append("")
        cards.append("Evidence:")
        for e in rc.get("evidence_summary") or []:
            cards.append(f"- {e}")
        cards.append("")
    p = out_root / "BeamDiagnosticCards.md"
    p.write_text("\n".join(cards), encoding="utf-8")
    paths["BeamDiagnosticCards.md"] = str(p)

    # Heatmap markdown
    freq = aggregate.get("failure_frequency") or {}
    max_f = max(list(freq.values()) or [1])
    heat = [
        f"# Pipeline Stage Heatmap — {PHASE_ID}",
        "",
        "| Stage | Count | Bar |",
        "|-------|------:|-----|",
    ]
    for k, v in freq.items():
        bar = "#" * int(20 * v / max_f) if max_f else ""
        heat.append(f"| {k} | {v} | `{bar}` |")
    heat += ["", f"Top primary root cause: **{aggregate.get('top_primary_root_cause')}**", ""]
    p = out_root / "PipelineStageHeatmap.md"
    p.write_text("\n".join(heat), encoding="utf-8")
    paths["PipelineStageHeatmap.md"] = str(p)

    # Recommendations md
    rec_lines = [
        f"# Engineering Recommendations — {PHASE_ID}",
        "",
        "Based ONLY on diagnostic evidence. No code was modified in this phase.",
        "",
    ]
    hyp = recommendations.get("hypothesis") or {}
    rec_lines += [
        "## Hypothesis",
        "",
        f"- ownership_or_scoping_before_render_is_dominant: "
        f"**{hyp.get('ownership_or_scoping_before_render_is_dominant')}**",
        f"- renderer_mostly_faithful_to_owned_set: "
        f"**{hyp.get('renderer_mostly_faithful_to_owned_set')}**",
        f"- evidence_confidence: **{hyp.get('evidence_confidence')}**",
        "",
        "## Priorities",
        "",
    ]
    for pr in recommendations.get("priorities") or []:
        rec_lines += [
            f"### Priority {pr.get('priority')}: {pr.get('target_stage')}",
            "",
            f"- Frequency: {pr.get('frequency')}",
            f"- Supporting beams: {', '.join(pr.get('supporting_beam_ids') or []) or '-'}",
            f"- Estimated impact: {pr.get('estimated_impact')}",
            f"- Recommendation: {pr.get('recommendation')}",
            f"- Note: {pr.get('deferred_note')}",
            "",
        ]
    p = out_root / "EngineeringRecommendations.md"
    p.write_text("\n".join(rec_lines), encoding="utf-8")
    paths["EngineeringRecommendations.md"] = str(p)

    # Excel
    xlsx_path = out_root / "BeamPipelineDiagnostics.xlsx"
    try:
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Beams"
        ws.append(
            [
                "beam_id",
                "Discovery",
                "Extent",
                "Crop",
                "Ownership",
                "Annotation",
                "Rendering",
                "first_fail",
                "primary",
                "confidence",
            ]
        )
        for r in beam_records:
            st = r.get("stages") or {}
            rc = r.get("root_cause") or {}
            ws.append(
                [
                    r.get("beam_id"),
                    (st.get("Beam Discovery") or {}).get("status"),
                    (st.get("Beam Extent") or {}).get("status"),
                    (st.get("Crop Window") or {}).get("status"),
                    (st.get("Ownership") or {}).get("status"),
                    (st.get("Annotation Association") or {}).get("status"),
                    (st.get("Rendering") or {}).get("status"),
                    rc.get("first_failing_stage"),
                    rc.get("primary_category"),
                    rc.get("confidence"),
                ]
            )
        ws2 = wb.create_sheet("FailureFrequency")
        ws2.append(["stage", "count"])
        for k, v in (aggregate.get("failure_frequency") or {}).items():
            ws2.append([k, v])
        ws3 = wb.create_sheet("Hypothesis")
        for k, v in (aggregate.get("hypothesis") or {}).items():
            ws3.append([k, v])
        wb.save(xlsx_path)
        paths["BeamPipelineDiagnostics.xlsx"] = str(xlsx_path)
    except Exception as exc:
        # Fallback CSV-like json already written; note in execution summary
        paths["BeamPipelineDiagnostics.xlsx"] = f"FAILED: {exc}"

    readme = f"""# Phase QA.3.1 — Pipeline Diagnostics

MODEL_VERSION: {MODEL_VERSION}

Diagnostic-only analysis of QA.3.0 unseen drawing failures.
No engineering / ownership / render logic was modified.

## Run

```
python Run_PY/run_phase_qa31_pipeline_diagnostics.py
```

## Key outputs

- BeamPipelineDiagnostics.json / .xlsx
- BeamDiagnosticCards.md
- RootCauseSummary.json
- FailureFrequency.json
- EngineeringRecommendations.md
- QA31Validation.json
"""
    p = out_root / "README.md"
    p.write_text(readme, encoding="utf-8")
    paths["README.md"] = str(p)

    return paths


def write_execution_summary(
    out_root: Path,
    aggregate: Dict[str, Any],
    recommendations: Dict[str, Any],
    validation: Dict[str, Any],
    elapsed_s: float,
) -> Path:
    hyp = aggregate.get("hypothesis") or {}
    freq = aggregate.get("failure_frequency") or {}
    top = aggregate.get("top_primary_root_cause")
    p1 = (recommendations.get("priorities") or [{}])[0]
    lines = [
        f"# Execution Summary — {PHASE_ID}",
        "",
        f"- MODEL_VERSION: {MODEL_VERSION}",
        f"- Elapsed (s): {elapsed_s}",
        f"- Beams analysed: {aggregate.get('beams_analysed')}",
        f"- Missing-artefact beams: {aggregate.get('beams_with_missing_artefacts')}",
        f"- QA overall_pass: {validation.get('overall_pass')}",
        "",
        "## Failure frequency",
        "",
    ]
    for k, v in freq.items():
        lines.append(f"- {k}: {v}")
    lines += [
        "",
        f"## Top primary root cause: {top}",
        "",
        "## Hypothesis",
        f"- ownership_or_scoping_before_render_is_dominant: "
        f"{hyp.get('ownership_or_scoping_before_render_is_dominant')}",
        f"- renderer_mostly_faithful_to_owned_set: "
        f"{hyp.get('renderer_mostly_faithful_to_owned_set')}",
        "",
        f"## Priority 1: {p1.get('target_stage')} — {p1.get('recommendation')}",
        "",
        "No engineering modules were modified.",
        "",
    ]
    path = out_root / "ExecutionSummary.md"
    path.write_text("\n".join(lines), encoding="utf-8")
    return path
