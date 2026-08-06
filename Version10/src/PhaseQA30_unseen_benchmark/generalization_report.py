"""
QA.3.0 — Generalization report builders.
MODEL_VERSION: 10.0.0
"""
from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

MODEL_VERSION = "10.0.0"
PHASE_ID = "QA.3.0"

# Map QA.2A error types + heuristic buckets to requested engineering categories
_CATEGORY_ALIASES = {
    "Beam Missing": "Missing beams",
    "Beam Mismatch": "Extra beams",
    "Missing Bar": "Missing bars",
    "Extra Bar": "Extra bars",
    "Wrong Diameter": "Diameter mismatch",
    "Wrong Quantity": "Missing bars",
    "Wrong Role": "Ownership issues",
    "Steel Difference": "Rendering mismatch",
}


def _map_freq(freq: Dict[str, int]) -> Dict[str, int]:
    out: Dict[str, int] = {
        "Missing beams": 0,
        "Ownership issues": 0,
        "Diameter mismatch": 0,
        "Missing bars": 0,
        "Extra bars": 0,
        "Rendering mismatch": 0,
        "Stirrup mismatch": 0,
        "Side-face reinforcement mismatch": 0,
        "Development Length mismatch": 0,
    }
    for k, v in (freq or {}).items():
        mapped = _CATEGORY_ALIASES.get(k, k)
        out[mapped] = out.get(mapped, 0) + int(v)
    return dict(sorted(out.items(), key=lambda x: -x[1]))


def build_engineering_error_summary(benchmark: Dict[str, Any]) -> Dict[str, Any]:
    compiled = benchmark.get("compiled") or {}
    freq = (compiled.get("errors") or {}).get("frequency") or {}
    mapped = _map_freq(freq)
    top = list(mapped.items())[:10]
    return {
        "phase_id": PHASE_ID,
        "model_version": MODEL_VERSION,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "raw_error_frequency": freq,
        "engineering_categories": mapped,
        "top_error_categories": [{"category": k, "count": v} for k, v in top],
        "total_errors": (compiled.get("errors") or {}).get("total_errors") or sum(freq.values()),
        "notes": [
            "Categories derived from unmodified QA.2A error_classifier frequency.",
            "Stirrup / side-face / development-length buckets remain 0 unless "
            "future classifiers emit dedicated types (formulas unchanged).",
        ],
    }


def build_generalization_assessment(
    bench: Dict[str, Any], eng: Dict[str, Any]
) -> Dict[str, Any]:
    b = (bench.get("compiled") or {}).get("benchmark") or {}
    top = eng.get("top_error_categories") or []
    strengths = []
    weaknesses = []
    if float(b.get("beam_detection_pct") or 0) >= 85:
        strengths.append("Strong beam detection on unseen drawings")
    else:
        weaknesses.append("Beam detection drops on unseen projects")
    if float(b.get("steel_accuracy_pct") or 0) >= 85:
        strengths.append("Steel tonnage remains relatively stable vs estimator GT")
    else:
        weaknesses.append("Steel accuracy degrades on unseen projects")
    if float(b.get("bar_accuracy_pct") or 0) < 40:
        weaknesses.append("Bar matching is the primary generalization gap")
    if float(b.get("bar_detection_pct") or 0) < 70:
        weaknesses.append("Bar detection incomplete on unseen reinforcement plans")

    failure_modes = [f"{t['category']} ({t['count']})" for t in top[:5]]
    recommendations = [
        "Improve bar-role / diameter matching on unfamiliar annotation styles",
        "Harden beam discovery for novel mark naming / framing conventions",
        "Review ownership / shared-scope edge cases that drive Missing Bar / Extra Bar",
        "Investigate diameter mismatches before changing weight formulas",
        "Keep production DXF-only; use these GT gaps to prioritize Version10 engineering work",
    ]
    if not strengths:
        strengths.append("Pipeline completes end-to-end on completely unseen DXF sets")
    return {
        "engineering_strengths": strengths,
        "engineering_weaknesses": weaknesses,
        "largest_failure_modes": failure_modes,
        "recommended_engineering_improvements": recommendations,
    }


def write_generalization_json(
    out_path: Path,
    production: Dict[str, Any],
    benchmark: Dict[str, Any],
    eng: Dict[str, Any],
    assessment: Dict[str, Any],
) -> Dict[str, Any]:
    compiled = benchmark.get("compiled") or {}
    b = compiled.get("benchmark") or {}
    doc = {
        "phase_id": PHASE_ID,
        "model_version": MODEL_VERSION,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "type": "unseen_drawing_generalization_benchmark",
        "estimator_excel_usage": "benchmark_only",
        "reuse_detected_any": production.get("reuse_detected_any", False),
        "overall_metrics": {
            "beam_detection_pct": b.get("beam_detection_pct"),
            "bar_detection_pct": b.get("bar_detection_pct"),
            "bar_matching_pct": b.get("bar_accuracy_pct"),
            "steel_accuracy_pct": b.get("steel_accuracy_pct"),
            "overall_accuracy_pct": b.get("overall_accuracy_pct"),
            "total_beams": b.get("total_beams"),
            "total_bars": b.get("total_bars"),
            "total_drawing_sets": b.get("total_drawing_sets"),
        },
        "drawing_set_results": (compiled.get("dashboard") or {}).get("drawing_sets") or [],
        "errors": compiled.get("errors"),
        "engineering_error_summary": eng,
        "generalization_assessment": assessment,
        "production": {
            "elapsed_s": production.get("elapsed_s"),
            "sets": [
                {
                    "drawing_set": s.get("drawing_set"),
                    "set_key": s.get("set_key"),
                    "success": s.get("success"),
                    "reuse_detected": s.get("reuse_detected"),
                    "pipeline_elapsed_s": s.get("pipeline_elapsed_s"),
                    "run_id": s.get("run_id"),
                    "model_excel": s.get("model_excel"),
                }
                for s in (production.get("sets") or [])
            ],
        },
        "benchmark": b,
        "statistics": compiled.get("statistics"),
    }
    out_path = Path(out_path)
    out_path.write_text(json.dumps(doc, indent=2, default=str), encoding="utf-8")
    return doc


def write_generalization_summary_md(
    out_path: Path,
    report: Dict[str, Any],
) -> Path:
    om = report.get("overall_metrics") or {}
    assess = report.get("generalization_assessment") or {}
    eng = report.get("engineering_error_summary") or {}
    lines = [
        f"# Generalization Summary — Phase {PHASE_ID}",
        "",
        f"**MODEL_VERSION:** {MODEL_VERSION}",
        f"**Generated:** {report.get('generated_at')}",
        "",
        "Estimator Output Excel used **ONLY during benchmarking** (never during production).",
        "",
        "## Overall Metrics",
        "",
        f"| Metric | Value |",
        f"|--------|------:|",
        f"| Beam Detection | {om.get('beam_detection_pct')}% |",
        f"| Bar Detection | {om.get('bar_detection_pct')}% |",
        f"| Bar Matching | {om.get('bar_matching_pct')}% |",
        f"| Steel Accuracy | {om.get('steel_accuracy_pct')}% |",
        f"| Overall Accuracy | {om.get('overall_accuracy_pct')}% |",
        "",
        "## Drawing Set-wise Results",
        "",
        "| Drawing Set | Beam Det | Bar Det | Bar Match | Steel | Overall |",
        "|-------------|---------:|--------:|----------:|------:|--------:|",
    ]
    for row in report.get("drawing_set_results") or []:
        lines.append(
            f"| {row.get('drawing_set')} | {row.get('beam_detection_pct')}% | "
            f"{row.get('bar_detection_pct')}% | {row.get('bar_accuracy_pct')}% | "
            f"{row.get('steel_accuracy_pct')}% | {row.get('overall_accuracy_pct')}% |"
        )
    lines += ["", "## Engineering Error Summary", ""]
    for item in (eng.get("top_error_categories") or [])[:10]:
        lines.append(f"- {item.get('category')}: {item.get('count')}")
    lines += ["", "## Generalization Assessment", "", "### Engineering strengths"]
    for s in assess.get("engineering_strengths") or []:
        lines.append(f"- {s}")
    lines += ["", "### Engineering weaknesses"]
    for s in assess.get("engineering_weaknesses") or []:
        lines.append(f"- {s}")
    lines += ["", "### Largest failure modes"]
    for s in assess.get("largest_failure_modes") or []:
        lines.append(f"- {s}")
    lines += ["", "### Recommended engineering improvements"]
    for s in assess.get("recommended_engineering_improvements") or []:
        lines.append(f"- {s}")
    lines.append("")
    out_path = Path(out_path)
    out_path.write_text("\n".join(lines), encoding="utf-8")
    return out_path


def write_generalization_xlsx(
    out_path: Path,
    report: Dict[str, Any],
    compiled: Dict[str, Any],
    results: List[Dict[str, Any]],
    engine_root: Path,
) -> Path:
    """Reuse QA.2A ExcelExporter then rename/copy to Generalization report name."""
    import shutil
    import sys

    ExcelExporter = sys.modules["excel_exporter"].ExcelExporter
    tmp = Path(out_path).parent / "_tmp_qa2a_export.xlsx"
    ExcelExporter().export(tmp, compiled, [r for r in results if r.get("compared")])
    dest = Path(out_path)
    shutil.copy2(tmp, dest)
    try:
        tmp.unlink()
    except OSError:
        pass

    # Append a lightweight Generalization sheet if openpyxl available
    try:
        from openpyxl import load_workbook

        wb = load_workbook(dest)
        if "Generalization" in wb.sheetnames:
            del wb["Generalization"]
        ws = wb.create_sheet("Generalization", 0)
        om = report.get("overall_metrics") or {}
        ws.append(["Phase", PHASE_ID])
        ws.append(["MODEL_VERSION", MODEL_VERSION])
        ws.append([])
        ws.append(["Overall Metrics"])
        for k in (
            "beam_detection_pct",
            "bar_detection_pct",
            "bar_matching_pct",
            "steel_accuracy_pct",
            "overall_accuracy_pct",
        ):
            ws.append([k, om.get(k)])
        ws.append([])
        ws.append(["Drawing Set", "Beam Det", "Bar Det", "Bar Match", "Steel", "Overall"])
        for row in report.get("drawing_set_results") or []:
            ws.append(
                [
                    row.get("drawing_set"),
                    row.get("beam_detection_pct"),
                    row.get("bar_detection_pct"),
                    row.get("bar_accuracy_pct"),
                    row.get("steel_accuracy_pct"),
                    row.get("overall_accuracy_pct"),
                ]
            )
        ws.append([])
        ws.append(["Top Engineering Error Categories"])
        for item in (report.get("engineering_error_summary") or {}).get(
            "top_error_categories"
        ) or []:
            ws.append([item.get("category"), item.get("count")])
        wb.save(dest)
    except Exception:
        pass
    return dest
