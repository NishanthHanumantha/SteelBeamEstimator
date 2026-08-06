"""
model_workbook_parser.py — Parse MODEL production workbook structure.
MODEL_VERSION: 8.1.2
"""
from __future__ import annotations

import re
from typing import Any, Dict, List, Optional

import openpyxl

from comparison_models import ModelBeam, ProjectSummary, RoleLine, WorkbookRef
from estimator_workbook_parser import _norm_role, _num


class ModelWorkbookParser:

    def __init__(self, path):
        self._path = path
        self._wb = openpyxl.load_workbook(path, data_only=True)

    def close(self):
        self._wb.close()

    def workbook_ref(self) -> WorkbookRef:
        return WorkbookRef(
            path=str(self._path),
            filename=self._path.name,
            sheet_names=self._wb.sheetnames,
            size_bytes=self._path.stat().st_size,
        )

    def parse_project_summary(self) -> ProjectSummary:
        dia_kg: Dict[int, float] = {}
        total = 0.0

        if "Diameter Summary" in self._wb.sheetnames:
            ws = self._wb["Diameter Summary"]
            for r in range(3, ws.max_row + 1):
                label = ws.cell(r, 1).value
                if not label:
                    continue
                if str(label).upper() == "TOTAL":
                    total = _num(ws.cell(r, 4).value)
                    continue
                m = re.search(r"Y(\d+)", str(label), re.I)
                if m:
                    dia_kg[int(m.group(1))] = round(_num(ws.cell(r, 4).value), 4)

        if total <= 0 and "Project Totals" in self._wb.sheetnames:
            ws = self._wb["Project Totals"]
            for r in range(1, ws.max_row + 1):
                item = ws.cell(r, 1).value
                if item and "total steel" in str(item).lower():
                    total = _num(ws.cell(r, 2).value)

        return ProjectSummary(
            label="MODEL Project Totals",
            concrete_m3=0.0,
            shuttering_m2=0.0,
            diameter_kg=dia_kg,
            total_steel_kg=round(total, 3),
            source_row=0,
        )

    def parse_beams(self) -> List[ModelBeam]:
        beams: Dict[str, ModelBeam] = {}

        if "Steel Summary" in self._wb.sheetnames:
            ws = self._wb["Steel Summary"]
            headers = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
            dia_cols: Dict[int, int] = {}
            total_col = ws.max_column
            for c, h in enumerate(headers, start=1):
                if h and re.search(r"Y(\d+)", str(h), re.I):
                    dia_cols[int(re.search(r"Y(\d+)", str(h), re.I).group(1))] = c
                if h and "total" in str(h).lower():
                    total_col = c

            for r in range(3, ws.max_row + 1):
                bid = ws.cell(r, 1).value
                if not bid:
                    continue
                bid = str(bid).strip().upper()
                if not re.match(r"^B\d+[A-Z]?$", bid, re.I):
                    continue
                dia_kg = {d: round(_num(ws.cell(r, col).value), 4) for d, col in dia_cols.items()}
                total = _num(ws.cell(r, total_col).value)
                if total <= 0:
                    total = sum(dia_kg.values())
                beams[bid] = ModelBeam(
                    beam_id=bid,
                    span_m=0.0,
                    total_bars=0,
                    steel_kg=round(total, 4),
                    diameter_kg={d: v for d, v in dia_kg.items() if v > 0},
                )

        if "Beam Summary" in self._wb.sheetnames:
            ws = self._wb["Beam Summary"]
            for r in range(3, ws.max_row + 1):
                bid = ws.cell(r, 1).value
                if not bid:
                    continue
                bid = str(bid).strip().upper()
                if not re.match(r"^B\d+[A-Z]?$", bid, re.I):
                    continue
                b = beams.get(bid) or ModelBeam(bid, 0, 0, 0, {})
                b.span_m = _num(ws.cell(r, 2).value)
                b.total_bars = int(_num(ws.cell(r, 5).value))
                if b.steel_kg <= 0:
                    b.steel_kg = round(_num(ws.cell(r, 6).value), 4)
                beams[bid] = b

        if "Bar Bending Schedule" in self._wb.sheetnames:
            ws = self._wb["Bar Bending Schedule"]
            headers = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
            weight_cols: Dict[int, int] = {}
            for c, h in enumerate(headers, start=1):
                if h and re.search(r"Y(\d+)", str(h), re.I):
                    weight_cols[int(re.search(r"Y(\d+)", str(h), re.I).group(1))] = c

            current_bid = None
            for r in range(3, ws.max_row + 1):
                desc = ws.cell(r, 3).value
                if not desc:
                    continue
                ds = str(desc).strip()
                if re.match(r"^B\d+[A-Z]?$", ds, re.I):
                    current_bid = ds.upper()
                    if current_bid not in beams:
                        beams[current_bid] = ModelBeam(current_bid, 0, 0, 0, {})
                    continue
                if not current_bid:
                    continue
                role = _norm_role(ds)
                dia = _num(ws.cell(r, 4).value) or None
                line = RoleLine(
                    role=role,
                    description=ds,
                    diameter_mm=dia,
                    spacing_m=_num(ws.cell(r, 5).value) or None,
                    bar_count=_num(ws.cell(r, 6).value) or None,
                    cut_length_m=_num(ws.cell(r, 8).value) or None,
                    total_length_m=_num(ws.cell(r, 9).value) or None,
                    steel_kg=0.0,
                    diameter_kg={},
                )
                for d, col in weight_cols.items():
                    kg = _num(ws.cell(r, col).value)
                    if kg > 0:
                        line.diameter_kg[d] = round(kg, 4)
                        line.steel_kg += kg
                line.steel_kg = round(line.steel_kg, 4)
                beams[current_bid].roles.append(line)

        return list(beams.values())


def discover_model_workbook(v7_root):
    import pathlib
    p = pathlib.Path(v7_root) / "data/output/Production_Output/Estimation_Output.xlsx"
    return p if p.exists() else None
