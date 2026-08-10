"""
Read Estimation_Output.xlsx for totals and B16 engineering rows.
MODEL_VERSION: 10.5.6
"""
from __future__ import annotations

import hashlib
from pathlib import Path
from typing import Any, Dict, List, Optional


def sha256_file(path: Path) -> Optional[str]:
    path = Path(path)
    if not path.exists():
        return None
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def engineering_content_fingerprint(wb: Dict[str, Any]) -> str:
    """
    Deterministic fingerprint of engineering content.

    VB1 embeds run timestamps into workbook/archive metadata, so raw file SHA
    is non-deterministic even when steel quantities are identical.
    """
    import json

    payload = {
        "beam_count": wb.get("beam_count"),
        "bar_count": wb.get("bar_count"),
        "steel_kg": wb.get("steel_kg"),
        "b16": wb.get("b16"),
        "b16_bbs_rows": wb.get("b16_bbs_rows"),
    }
    return hashlib.sha256(
        json.dumps(payload, sort_keys=True, separators=(",", ":"), default=str).encode(
            "utf-8"
        )
    ).hexdigest()


def inspect_workbook(path: Path) -> Dict[str, Any]:
    path = Path(path)
    out: Dict[str, Any] = {
        "path": str(path),
        "exists": path.exists(),
        "sha256": sha256_file(path),
        "content_fingerprint": None,
        "beam_count": 0,
        "bar_count": 0,
        "steel_kg": 0.0,
        "b16": None,
        "b16_bbs_rows": [],
        "ok": False,
        "error": None,
    }
    if not path.exists():
        out["error"] = "missing"
        return out
    try:
        import openpyxl

        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        beam_ids: List[str] = []
        bar_count = 0
        steel = 0.0
        b16_summary = None
        bbs: List[Dict[str, Any]] = []

        if "Beam Summary" in wb.sheetnames:
            ws = wb["Beam Summary"]
            header_seen = False
            headers = []
            for vals in ws.iter_rows(values_only=True):
                cells = list(vals)
                first = str(cells[0]).strip() if cells and cells[0] is not None else ""
                if not header_seen:
                    if first.lower() == "beam id":
                        header_seen = True
                        headers = [str(c).strip() if c is not None else "" for c in cells]
                    continue
                if not first or first.lower().startswith("total"):
                    continue
                beam_ids.append(first)
                row_steel = 0.0
                row_bars = 0
                if len(cells) > 5 and cells[5] is not None:
                    try:
                        row_steel = float(cells[5])
                        steel += row_steel
                    except (TypeError, ValueError):
                        pass
                if len(cells) > 4 and cells[4] is not None:
                    try:
                        row_bars = int(float(cells[4]))
                        bar_count += row_bars
                    except (TypeError, ValueError):
                        pass
                if first == "B16":
                    b16_summary = {
                        "beam_id": "B16",
                        "length_m": cells[1] if len(cells) > 1 else None,
                        "width_m": cells[2] if len(cells) > 2 else None,
                        "depth_m": cells[3] if len(cells) > 3 else None,
                        "bar_count": row_bars,
                        "steel_kg": row_steel,
                        "raw": [cells[i] if i < len(cells) else None for i in range(12)],
                        "headers": headers[:12],
                    }

        if "Bar Bending Schedule" in wb.sheetnames:
            ws = wb["Bar Bending Schedule"]
            header_seen = False
            for vals in ws.iter_rows(values_only=True):
                cells = list(vals)
                # beam id typically column index 2
                bid = None
                for c in cells[:5]:
                    if c is not None and str(c).strip() == "B16":
                        bid = "B16"
                        break
                if bid == "B16":
                    bbs.append(
                        {
                            "cells": [
                                cells[i] if i < len(cells) else None for i in range(12)
                            ]
                        }
                    )

        if "Project Totals" in wb.sheetnames:
            ws = wb["Project Totals"]
            for vals in ws.iter_rows(values_only=True):
                cells = list(vals)
                label = str(cells[0]).strip().lower() if cells and cells[0] else ""
                if "total steel" in label or label == "steel weight (kg)":
                    for c in cells[1:]:
                        if c is not None:
                            try:
                                steel = float(c)
                                break
                            except (TypeError, ValueError):
                                pass

        out.update(
            {
                "beam_count": len(beam_ids),
                "bar_count": bar_count,
                "steel_kg": round(float(steel), 6),
                "b16": b16_summary,
                "b16_bbs_rows": bbs,
                "b16_bbs_row_count": len(bbs),
                "ok": True,
            }
        )
        out["content_fingerprint"] = engineering_content_fingerprint(out)
        wb.close()
    except Exception as exc:  # noqa: BLE001
        out["error"] = str(exc)
    return out
