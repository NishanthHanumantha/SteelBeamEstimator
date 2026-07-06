"""Workbook parsing and comparison helpers — Phase QA.1."""

from __future__ import annotations

import json
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.estimator_validation.audit_types import FLOAT_TOLERANCE, SUMMARY_METADATA_LABELS


@dataclass
class ScheduleRow:
    row_number: int
    description: str
    normalized_description: str
    role_hint: str
    diameter_mm: Optional[float] = None
    spacing_m: Optional[float] = None
    bar_count: Optional[float] = None
    development_length_m: Optional[float] = None
    cut_length_m: Optional[float] = None
    total_length_m: Optional[float] = None
    steel_weight_kg: Optional[float] = None
    fabrication_mark: Optional[str] = None
    shape_code: Optional[str] = None
    display_order: Optional[int] = None
    raw: Dict[str, Any] = field(default_factory=dict)


@dataclass
class BeamBlock:
    beam_mark: str
    header_row: int
    clear_span_m: Optional[float] = None
    width_m: Optional[float] = None
    depth_m: Optional[float] = None
    rows: List[ScheduleRow] = field(default_factory=list)


def normalize_description(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"\s+", " ", text)
    return text


def role_hint_from_description(description: str) -> str:
    text = normalize_description(description)
    mapping = {
        "top bars": "TOP_MAIN",
        "top bars - extra": "TOP_EXTRA",
        "top bars-extra": "TOP_EXTRA",
        "bottom bars": "BOTTOM_MAIN",
        "bottom bars -extra": "BOTTOM_EXTRA",
        "bottom bars - extra": "BOTTOM_EXTRA",
        "stirupps": "STIRRUP",
        "stirrups": "STIRRUP",
        "sfr": "SFR",
        "spacer bars": "SPACER_BAR",
        "side bars": "SIDE_BAR",
    }
    return mapping.get(text, "UNKNOWN")


def values_equal(left: Any, right: Any, tolerance: float = FLOAT_TOLERANCE) -> bool:
    if left is None and right is None:
        return True
    if left is None or right is None:
        return False
    if isinstance(left, (int, float)) and isinstance(right, (int, float)):
        return abs(float(left) - float(right)) <= tolerance
    return str(left).strip() == str(right).strip()


def to_float(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def is_summary_metadata_row(description: str) -> bool:
    return normalize_description(description) in SUMMARY_METADATA_LABELS


def is_beam_header_row(ws: Worksheet, row: int) -> bool:
    serial = ws.cell(row, 1).value
    mark = ws.cell(row, 3).value
    return isinstance(serial, (int, float)) and bool(mark) and str(mark).startswith("B") and str(mark)[1:].isdigit()


def find_schedule_start_row(ws: Worksheet) -> int:
    for row in range(1, ws.max_row + 1):
        description = ws.cell(row, 3).value
        if description and "quantity breakup" in normalize_description(description):
            return row
        if ws.cell(row, 1).value == "SI no" and normalize_description(ws.cell(row, 3).value) == "description":
            return row
    return 1


def parse_schedule_rows(ws: Worksheet, start_row: int) -> Dict[str, BeamBlock]:
    beams: Dict[str, BeamBlock] = {}
    current: Optional[BeamBlock] = None
    for row in range(start_row, ws.max_row + 1):
        if is_beam_header_row(ws, row):
            mark = str(ws.cell(row, 3).value)
            current = BeamBlock(
                beam_mark=mark,
                header_row=row,
                clear_span_m=to_float(ws.cell(row, 5).value),
                width_m=to_float(ws.cell(row, 6).value),
                depth_m=to_float(ws.cell(row, 7).value),
            )
            beams[mark] = current
            continue
        if current is None:
            continue
        description = ws.cell(row, 3).value
        if not description:
            continue
        if str(description).startswith("B") and str(description)[1:].isdigit():
            continue
        if is_summary_metadata_row(str(description)):
            current = None
            continue
        normalized = normalize_description(description)
        schedule_row = ScheduleRow(
            row_number=row,
            description=str(description).strip(),
            normalized_description=normalized,
            role_hint=role_hint_from_description(str(description)),
            diameter_mm=to_float(ws.cell(row, 4).value),
            spacing_m=to_float(ws.cell(row, 5).value),
            bar_count=to_float(ws.cell(row, 6).value),
            development_length_m=to_float(ws.cell(row, 7).value),
            cut_length_m=to_float(ws.cell(row, 8).value),
            total_length_m=to_float(ws.cell(row, 9).value),
            steel_weight_kg=to_float(ws.cell(row, 17).value),
            raw={
                "col_d": ws.cell(row, 4).value,
                "col_e": ws.cell(row, 5).value,
                "col_f": ws.cell(row, 6).value,
                "col_g": ws.cell(row, 7).value,
                "col_h": ws.cell(row, 8).value,
                "col_i": ws.cell(row, 9).value,
                "col_q": ws.cell(row, 17).value,
            },
        )
        current.rows.append(schedule_row)
    return beams


def load_workbook_pair(generated_path: Path, estimator_path: Path) -> tuple[Workbook, Workbook, Worksheet, Worksheet]:
    generated_wb = load_workbook(generated_path, data_only=True)
    estimator_wb = load_workbook(estimator_path, data_only=True)
    generated_ws = generated_wb[generated_wb.sheetnames[0]]
    estimator_ws = estimator_wb[estimator_wb.sheetnames[0]]
    return generated_wb, estimator_wb, generated_ws, estimator_ws


def beam_sort_key(mark: str) -> int:
    return int(str(mark)[1:])


def row_match_key(row: ScheduleRow) -> str:
    return f"{row.role_hint}|{row.normalized_description}|{row.diameter_mm}"


def load_json_if_exists(path: Path) -> Any:
    if not path.exists():
        return None
    return json.loads(path.read_text(encoding="utf-8"))


def index_reports_by_beam(records: List[dict[str, Any]]) -> Dict[str, dict[str, Any]]:
    return {str(item.get("beam_mark") or item.get("beam_id")): item for item in records}


def index_schedules_by_beam(records: List[dict[str, Any]]) -> Dict[str, dict[str, Any]]:
    mapping: Dict[str, dict[str, Any]] = {}
    for item in records:
        mark = str(item.get("beam_mark") or item.get("beam_id"))
        mapping[mark] = item
    return mapping


def workbook_structure_snapshot(wb, ws: Worksheet) -> dict[str, Any]:
    return {
        "workbook_count": 1,
        "worksheet_count": len(wb.sheetnames),
        "worksheet_names": list(wb.sheetnames),
        "active_worksheet": ws.title,
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "merged_cell_count": len(ws.merged_cells.ranges),
        "merged_ranges": [str(item) for item in ws.merged_cells.ranges],
        "freeze_panes": str(ws.freeze_panes),
        "print_area": ws.print_area,
        "page_setup_orientation": ws.page_setup.orientation,
        "page_margins": {
            "left": ws.page_margins.left,
            "right": ws.page_margins.right,
            "top": ws.page_margins.top,
            "bottom": ws.page_margins.bottom,
        },
        "column_widths": {
            col: ws.column_dimensions[col].width
            for col in ws.column_dimensions
            if ws.column_dimensions[col].width is not None
        },
    }
