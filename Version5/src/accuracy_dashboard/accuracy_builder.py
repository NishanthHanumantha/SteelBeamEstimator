"""Build accuracy metrics from existing workbook outputs — Phase QA.ACCURACY.1."""

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List

from src.accuracy_dashboard.accuracy_types import (
    DIAMETER_SUMMARY_SOURCE,
    ENGINEERING_VALUE_FIELDS,
    FLOAT_TOLERANCE,
    STANDARD_DIAMETERS_MM,
    SUMMARY_PARSE_WARNING,
    SUMMARY_TOTAL_TOLERANCE_KG,
)
from src.estimator_validation.comparison_utils import (
    beam_sort_key,
    load_json_if_exists,
    load_workbook_pair,
    parse_schedule_rows,
    row_match_key,
    values_equal,
)
from src.estimator_validation.comparison_utils import find_schedule_start_row
from openpyxl.worksheet.worksheet import Worksheet


class AccuracyBuilder:
    """Compare generated Excel against estimator-validated ground truth."""

    def __init__(self, paths: dict[str, Path]) -> None:
        self._paths = paths

    def build(self) -> dict[str, Any]:
        generated_path = self._paths["generated_workbook"]
        estimator_path = self._paths["estimator_workbook"]
        _, _, generated_ws, estimator_ws = load_workbook_pair(generated_path, estimator_path)
        generated_start = find_schedule_start_row(generated_ws)
        estimator_start = find_schedule_start_row(estimator_ws)
        generated_beams = parse_schedule_rows(generated_ws, generated_start)
        estimator_beams = parse_schedule_rows(estimator_ws, estimator_start)

        estimator_summary = self._parse_official_quantity_summary(estimator_ws)
        generated_summary = self._parse_official_quantity_summary(generated_ws)
        official_quantity_summary = {
            "estimator": self._present_official_summary(estimator_summary),
            "generated": self._present_official_summary(generated_summary),
            "diameter_summary_source": DIAMETER_SUMMARY_SOURCE,
            "warnings": sorted(set(
                (estimator_summary.get("warnings") or [])
                + (generated_summary.get("warnings") or [])
            )),
        }

        excel_accuracy = self._build_excel_accuracy(generated_beams, estimator_beams)
        steel_accuracy = self._build_steel_accuracy_from_summary(estimator_summary, generated_summary)
        pipeline_metadata = self._load_pipeline_metadata()
        beam_table = self._build_beam_coverage_table(generated_beams, estimator_beams)
        diameter_coverage = self._build_diameter_coverage_from_summary(estimator_summary, generated_summary)
        statistics = self._build_statistics(
            generated_beams,
            estimator_beams,
            excel_accuracy,
            steel_accuracy,
            beam_table,
            diameter_coverage,
            official_quantity_summary,
        )

        return {
            "excel_accuracy": excel_accuracy,
            "steel_accuracy": steel_accuracy,
            "diameter_coverage": diameter_coverage,
            "official_quantity_summary": official_quantity_summary,
            "quantity_source": DIAMETER_SUMMARY_SOURCE,
            "schedule_row_aggregation_used": False,
            "beam_coverage_table": beam_table,
            "accuracy_statistics": statistics,
            "pipeline_metadata": pipeline_metadata,
            "generated_beam_marks": sorted(generated_beams.keys(), key=beam_sort_key),
            "estimator_beam_marks": sorted(estimator_beams.keys(), key=beam_sort_key),
        }

    def _load_pipeline_metadata(self) -> dict[str, Any]:
        engineering_report = load_json_if_exists(self._paths["engineering_report_json"])
        beam_schedule = load_json_if_exists(self._paths["beam_schedule_json"])
        report_results = list((engineering_report or {}).get("results") or [])
        schedule_results = list((beam_schedule or {}).get("results") or [])
        return {
            "engineering_report_present": engineering_report is not None,
            "engineering_report_beam_count": len(report_results),
            "beam_schedule_present": beam_schedule is not None,
            "beam_schedule_beam_count": len(schedule_results),
            "engineering_report_path": str(self._paths["engineering_report_json"]),
            "beam_schedule_path": str(self._paths["beam_schedule_json"]),
        }

    @staticmethod
    def _count_rows(beams: dict[str, Any]) -> int:
        return sum(len(block.rows) for block in beams.values())

    @staticmethod
    def _safe_percent(numerator: float, denominator: float) -> float:
        if denominator <= 0:
            return 0.0
        return round((numerator / denominator) * 100.0, 2)

    def _build_excel_accuracy(
        self,
        generated_beams: dict[str, Any],
        estimator_beams: dict[str, Any],
    ) -> dict[str, Any]:
        estimator_marks = set(estimator_beams.keys())
        generated_marks = set(generated_beams.keys())
        matched_beams = estimator_marks & generated_marks
        missing_beam_marks = sorted(estimator_marks - generated_marks, key=beam_sort_key)

        estimator_beam_count = len(estimator_marks)
        generated_beam_count = len(generated_marks)
        beams_present = len(matched_beams)

        estimator_row_count = self._count_rows(estimator_beams)
        generated_row_count = self._count_rows(generated_beams)
        missing_rows, extra_rows = self._count_schedule_row_gaps(
            generated_beams,
            estimator_beams,
            generated_row_count,
            estimator_row_count,
        )
        missing_values = self._count_missing_engineering_values(generated_beams, estimator_beams)

        return {
            "beam_coverage_percent": self._safe_percent(beams_present, estimator_beam_count),
            "beam_coverage": f"{beams_present} / {estimator_beam_count}",
            "generated_beam_count": generated_beam_count,
            "estimator_beam_count": estimator_beam_count,
            "beams_present_in_generated": beams_present,
            "row_coverage_percent": self._safe_percent(generated_row_count, estimator_row_count),
            "row_coverage": f"{generated_row_count} / {estimator_row_count}",
            "generated_row_count": generated_row_count,
            "estimator_row_count": estimator_row_count,
            "missing_beams": len(missing_beam_marks),
            "missing_beam_marks": missing_beam_marks,
            "missing_rows": missing_rows,
            "extra_rows": extra_rows,
            "missing_values": missing_values,
        }

    @staticmethod
    def _present_official_summary(summary: dict[str, Any]) -> dict[str, Any]:
        diameters = summary.get("diameters") or {}
        payload = {str(diameter): round(float(diameters.get(diameter, 0.0)), 3) for diameter in STANDARD_DIAMETERS_MM}
        payload["total"] = round(float(summary.get("total") or 0.0), 3)
        return payload

    @classmethod
    def _parse_official_quantity_summary(cls, worksheet: Worksheet) -> dict[str, Any]:
        header_row, diameter_columns = cls._find_diameter_header_row(worksheet)
        if not header_row or not diameter_columns:
            return {
                "diameters": {diameter: 0.0 for diameter in STANDARD_DIAMETERS_MM},
                "total": 0.0,
                "header_row": None,
                "summary_row": None,
                "total_column": None,
                "aggregation_method": "NOT_FOUND",
                "warnings": [SUMMARY_PARSE_WARNING],
            }

        total_column = cls._find_total_steel_column(worksheet, header_row, diameter_columns)
        summary_row_match = cls._find_official_summary_row(
            worksheet,
            header_row,
            diameter_columns,
            total_column,
        )
        warnings: List[str] = []
        if summary_row_match:
            diameters = summary_row_match["diameters"]
            total = round(float(summary_row_match["total"]), 3)
            aggregation_method = "SUMMARY_ROW"
            summary_row = summary_row_match["row"]
        else:
            diameters, total = cls._aggregate_summary_columns(
                worksheet,
                header_row,
                diameter_columns,
            )
            aggregation_method = "SUMMARY_COLUMN_TOTAL"
            summary_row = None

        diameters = {diameter: round(float(diameters.get(diameter, 0.0)), 3) for diameter in STANDARD_DIAMETERS_MM}
        diameter_sum = round(sum(diameters.values()), 3)
        if total > 0 and abs(diameter_sum - total) > SUMMARY_TOTAL_TOLERANCE_KG:
            warnings.append(SUMMARY_PARSE_WARNING)

        return {
            "diameters": diameters,
            "total": total,
            "header_row": header_row,
            "summary_row": summary_row,
            "total_column": total_column,
            "diameter_columns": diameter_columns,
            "aggregation_method": aggregation_method,
            "warnings": warnings,
        }

    @staticmethod
    def _find_diameter_header_row(worksheet: Worksheet) -> tuple[int | None, dict[int, int]]:
        targets = list(STANDARD_DIAMETERS_MM)
        best_row: int | None = None
        best_columns: dict[int, int] = {}
        for row in range(1, worksheet.max_row + 1):
            matches: List[tuple[int, int]] = []
            for col in range(1, worksheet.max_column + 1):
                value = worksheet.cell(row, col).value
                if isinstance(value, (int, float)) and int(value) in targets:
                    matches.append((col, int(value)))
            if len(matches) < len(targets):
                continue
            ordered = sorted(matches, key=lambda item: item[0])
            values = [diameter for _, diameter in ordered]
            if values[: len(targets)] != targets:
                continue
            selected = {diameter: col for col, diameter in ordered[: len(targets)]}
            if best_row is None or row > best_row:
                best_row = row
                best_columns = selected
        return best_row, best_columns

    @staticmethod
    def _find_total_steel_column(
        worksheet: Worksheet,
        header_row: int,
        diameter_columns: dict[int, int],
    ) -> int:
        max_diameter_col = max(diameter_columns.values())
        for col in range(max_diameter_col + 1, worksheet.max_column + 1):
            header = str(worksheet.cell(header_row, col).value or "").lower()
            if "steel" in header and "kg" in header:
                return col
        return max_diameter_col + 1

    @classmethod
    def _find_official_summary_row(
        cls,
        worksheet: Worksheet,
        header_row: int,
        diameter_columns: dict[int, int],
        total_column: int,
    ) -> dict[str, Any] | None:
        best: dict[str, Any] | None = None
        for row in range(header_row + 1, worksheet.max_row + 1):
            diameters: dict[int, float] = {}
            numeric_count = 0
            for diameter, col in diameter_columns.items():
                value = worksheet.cell(row, col).value
                if isinstance(value, (int, float)):
                    diameters[diameter] = float(value)
                    numeric_count += 1
                else:
                    diameters[diameter] = 0.0
            if numeric_count < len(STANDARD_DIAMETERS_MM):
                continue
            total_value = worksheet.cell(row, total_column).value
            if not isinstance(total_value, (int, float)):
                continue
            total = float(total_value)
            diameter_sum = sum(diameters.values())
            if total <= 0 or diameter_sum <= 0:
                continue
            if abs(diameter_sum - total) > SUMMARY_TOTAL_TOLERANCE_KG:
                continue
            if best is None or total > float(best["total"]):
                best = {"row": row, "diameters": diameters, "total": total}
        return best

    @classmethod
    def _aggregate_summary_columns(
        cls,
        worksheet: Worksheet,
        header_row: int,
        diameter_columns: dict[int, int],
    ) -> tuple[dict[int, float], float]:
        totals = {diameter: 0.0 for diameter in STANDARD_DIAMETERS_MM}
        for row in range(header_row + 1, worksheet.max_row + 1):
            for diameter, col in diameter_columns.items():
                value = worksheet.cell(row, col).value
                if isinstance(value, (int, float)):
                    totals[diameter] += float(value)
        total = round(sum(totals.values()), 3)
        rounded_totals = {diameter: round(value, 3) for diameter, value in totals.items()}
        return rounded_totals, total

    def _build_steel_accuracy_from_summary(
        self,
        estimator_summary: dict[str, Any],
        generated_summary: dict[str, Any],
    ) -> dict[str, Any]:
        generated_steel = float(generated_summary.get("total") or 0.0)
        estimator_steel = float(estimator_summary.get("total") or 0.0)
        difference_kg = round(estimator_steel - generated_steel, 3)
        accuracy_percent = self._safe_percent(generated_steel, estimator_steel)
        difference_percent = round(100.0 - accuracy_percent, 2) if estimator_steel > 0 else 0.0
        return {
            "generated_steel_kg": round(generated_steel, 3),
            "estimator_steel_kg": round(estimator_steel, 3),
            "accuracy_percent": accuracy_percent,
            "difference_kg": difference_kg,
            "difference_percent": difference_percent,
            "quantity_source": DIAMETER_SUMMARY_SOURCE,
        }

    @classmethod
    def _build_diameter_coverage_from_summary(
        cls,
        estimator_summary: dict[str, Any],
        generated_summary: dict[str, Any],
    ) -> dict[str, Any]:
        estimator_diameters = estimator_summary.get("diameters") or {}
        generated_diameters = generated_summary.get("diameters") or {}
        diameters: List[dict[str, Any]] = []
        calculable: List[dict[str, Any]] = []
        for diameter in STANDARD_DIAMETERS_MM:
            estimator_kg = float(estimator_diameters.get(diameter, 0.0))
            generated_kg = float(generated_diameters.get(diameter, 0.0))
            entry = cls._coverage_entry(diameter, estimator_kg, generated_kg, [])
            diameters.append(entry)
            if isinstance(entry["coverage_percent"], (int, float)):
                calculable.append(entry)

        missing_diameters = [
            entry["diameter_mm"]
            for entry in diameters
            if entry["estimator_steel_kg"] > 0 and entry["generated_steel_kg"] <= 0
        ]
        coverage_values = [float(entry["coverage_percent"]) for entry in calculable]
        average_coverage = round(sum(coverage_values) / len(coverage_values), 2) if coverage_values else None
        median_coverage = None
        if coverage_values:
            ordered = sorted(coverage_values)
            mid = len(ordered) // 2
            if len(ordered) % 2 == 0:
                median_coverage = round((ordered[mid - 1] + ordered[mid]) / 2.0, 2)
            else:
                median_coverage = round(ordered[mid], 2)

        best_entry = max(calculable, key=lambda item: float(item["coverage_percent"])) if calculable else None
        worst_entry = min(calculable, key=lambda item: float(item["coverage_percent"])) if calculable else None
        largest_gap_entry = max(
            diameters,
            key=lambda item: float(item["difference_kg"]) if item["estimator_steel_kg"] > 0 else -1.0,
        ) if any(item["estimator_steel_kg"] > 0 for item in diameters) else None

        distribution = cls._coverage_distribution(calculable)
        return {
            "diameters": diameters,
            "quantity_source": DIAMETER_SUMMARY_SOURCE,
            "summary": {
                "best_performing_diameter_mm": best_entry["diameter_mm"] if best_entry else None,
                "best_performing_coverage_percent": best_entry["coverage_percent"] if best_entry else None,
                "worst_performing_diameter_mm": worst_entry["diameter_mm"] if worst_entry else None,
                "worst_performing_coverage_percent": worst_entry["coverage_percent"] if worst_entry else None,
                "missing_diameters_mm": missing_diameters,
                "largest_quantity_gap": {
                    "diameter_mm": largest_gap_entry["diameter_mm"] if largest_gap_entry else None,
                    "difference_kg": largest_gap_entry["difference_kg"] if largest_gap_entry else None,
                },
                "largest_coverage": {
                    "diameter_mm": best_entry["diameter_mm"] if best_entry else None,
                    "coverage_percent": best_entry["coverage_percent"] if best_entry else None,
                },
                "average_diameter_coverage_percent": average_coverage,
                "median_diameter_coverage_percent": median_coverage,
                "coverage_distribution": distribution,
            },
        }

    @staticmethod
    def _count_schedule_row_gaps(
        generated_beams: dict[str, Any],
        estimator_beams: dict[str, Any],
        generated_row_count: int,
        estimator_row_count: int,
    ) -> tuple[int, int]:
        """Count missing/extra schedule rows using positional beam-level comparison."""
        missing_rows = 0
        extra_rows = 0
        for mark in sorted(estimator_beams.keys(), key=beam_sort_key):
            estimator_rows = estimator_beams[mark].rows
            generated_block = generated_beams.get(mark)
            generated_rows = generated_block.rows if generated_block else []
            max_len = max(len(estimator_rows), len(generated_rows))
            for index in range(max_len):
                estimator_row = estimator_rows[index] if index < len(estimator_rows) else None
                generated_row = generated_rows[index] if index < len(generated_rows) else None
                if estimator_row and not generated_row:
                    missing_rows += 1
                elif generated_row and not estimator_row:
                    extra_rows += 1
        if extra_rows == 0 and missing_rows == 0 and estimator_row_count > generated_row_count:
            missing_rows = estimator_row_count - generated_row_count
        return missing_rows, extra_rows

    @staticmethod
    def _count_missing_engineering_values(
        generated_beams: dict[str, Any],
        estimator_beams: dict[str, Any],
    ) -> int:
        missing_values = 0
        row_fields = (
            "diameter_mm",
            "spacing_m",
            "bar_count",
            "cut_length_m",
            "total_length_m",
            "steel_weight_kg",
            "fabrication_mark",
            "shape_code",
        )
        for mark in sorted(estimator_beams.keys(), key=beam_sort_key):
            estimator_block = estimator_beams[mark]
            generated_block = generated_beams.get(mark)
            if generated_block:
                if not values_equal(
                    estimator_block.clear_span_m,
                    generated_block.clear_span_m,
                    tolerance=FLOAT_TOLERANCE,
                ):
                    missing_values += 1
            estimator_map = {row_match_key(row): row for row in estimator_block.rows}
            generated_map = {
                row_match_key(row): row for row in (generated_block.rows if generated_block else [])
            }
            for key, estimator_row in estimator_map.items():
                generated_row = generated_map.get(key)
                if not generated_row:
                    continue
                if estimator_row.role_hint != generated_row.role_hint:
                    missing_values += 1
                for field in row_fields:
                    left = getattr(estimator_row, field)
                    right = getattr(generated_row, field)
                    if not values_equal(left, right, tolerance=FLOAT_TOLERANCE):
                        missing_values += 1
        return missing_values

    @staticmethod
    def _build_beam_coverage_table(
        generated_beams: dict[str, Any],
        estimator_beams: dict[str, Any],
    ) -> List[dict[str, Any]]:
        table: List[dict[str, Any]] = []
        for mark in sorted(estimator_beams.keys(), key=beam_sort_key):
            estimator_block = estimator_beams[mark]
            generated_block = generated_beams.get(mark)
            estimator_rows = len(estimator_block.rows)
            generated_rows = len(generated_block.rows) if generated_block else 0
            generated_keys = {
                row_match_key(row) for row in (generated_block.rows if generated_block else [])
            }
            matched_rows = sum(
                1 for row in estimator_block.rows if row_match_key(row) in generated_keys
            )
            table.append({
                "beam_mark": mark,
                "estimator_rows": estimator_rows,
                "generated_rows": generated_rows,
                "matched_rows": matched_rows,
                "missing_rows": estimator_rows - matched_rows,
                "beam_present": generated_block is not None,
                "row_coverage_percent": AccuracyBuilder._safe_percent(generated_rows, estimator_rows),
            })
        return table

    @staticmethod
    def _build_statistics(
        generated_beams: dict[str, Any],
        estimator_beams: dict[str, Any],
        excel_accuracy: dict[str, Any],
        steel_accuracy: dict[str, Any],
        beam_table: List[dict[str, Any]],
        diameter_coverage: dict[str, Any],
        official_quantity_summary: dict[str, Any],
    ) -> dict[str, Any]:
        return {
            "total_beams_estimator": len(estimator_beams),
            "total_beams_generated": len(generated_beams),
            "total_rows_estimator": excel_accuracy["estimator_row_count"],
            "total_rows_generated": excel_accuracy["generated_row_count"],
            "beam_coverage_percent": excel_accuracy["beam_coverage_percent"],
            "row_coverage_percent": excel_accuracy["row_coverage_percent"],
            "missing_beams": excel_accuracy["missing_beams"],
            "missing_rows": excel_accuracy["missing_rows"],
            "extra_rows": excel_accuracy["extra_rows"],
            "missing_values": excel_accuracy["missing_values"],
            "generated_steel_kg": steel_accuracy["generated_steel_kg"],
            "estimator_steel_kg": steel_accuracy["estimator_steel_kg"],
            "steel_accuracy_percent": steel_accuracy["accuracy_percent"],
            "steel_difference_kg": steel_accuracy["difference_kg"],
            "steel_difference_percent": steel_accuracy["difference_percent"],
            "quantity_source": DIAMETER_SUMMARY_SOURCE,
            "official_total_steel_estimator": official_quantity_summary.get("estimator", {}).get("total"),
            "official_total_steel_generated": official_quantity_summary.get("generated", {}).get("total"),
            "beam_table": beam_table,
            "diameter_coverage_summary": diameter_coverage.get("summary", {}),
        }

    @staticmethod
    def _coverage_entry(
        diameter_mm: int,
        estimator_kg: float,
        generated_kg: float,
        roles: List[str],
    ) -> dict[str, Any]:
        if estimator_kg <= 0:
            return {
                "diameter_mm": diameter_mm,
                "estimator_steel_kg": round(estimator_kg, 3),
                "generated_steel_kg": round(generated_kg, 3),
                "coverage_percent": "N/A",
                "difference_kg": round(estimator_kg - generated_kg, 3),
                "difference_percent": "N/A",
                "roles_present": roles,
            }
        coverage_percent = AccuracyBuilder._safe_percent(generated_kg, estimator_kg)
        difference_kg = round(estimator_kg - generated_kg, 3)
        difference_percent = round(100.0 - coverage_percent, 2)
        return {
            "diameter_mm": diameter_mm,
            "estimator_steel_kg": round(estimator_kg, 3),
            "generated_steel_kg": round(generated_kg, 3),
            "coverage_percent": coverage_percent,
            "difference_kg": difference_kg,
            "difference_percent": difference_percent,
            "roles_present": roles,
        }

    @staticmethod
    def _coverage_distribution(calculable: List[dict[str, Any]]) -> dict[str, List[int]]:
        bands = {
            "high_coverage_gte_90": [],
            "moderate_coverage_70_to_90": [],
            "low_coverage_30_to_70": [],
            "very_low_coverage_lt_30": [],
        }
        for entry in calculable:
            diameter = entry["diameter_mm"]
            value = float(entry["coverage_percent"])
            if value >= 90.0:
                bands["high_coverage_gte_90"].append(diameter)
            elif value >= 70.0:
                bands["moderate_coverage_70_to_90"].append(diameter)
            elif value >= 30.0:
                bands["low_coverage_30_to_70"].append(diameter)
            else:
                bands["very_low_coverage_lt_30"].append(diameter)
        return bands
