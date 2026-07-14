"""
Phase V.A.1.1 — workbook_difference_analyzer.py
Compares the generated Estimation_Output.xlsx against the estimator reference.
Computes: worksheet match, header match, row match, column match, cell match,
engineering match.
MODEL_VERSION: 6.6.3
"""
from __future__ import annotations

import pathlib
from typing import Any, Dict, List, Tuple

import openpyxl

from validation_recompute_models import WorkbookDiffResult, WorksheetDiffResult

_ROOT      = pathlib.Path(__file__).resolve().parents[3]
_GENERATED = (
    _ROOT / "Version7/data/output/Production_Output/Estimation_Output.xlsx"
)
_REFERENCE = (
    _ROOT
    / "Version7/data/Excel_Presentation_Format"
    / "Galera_SteelBeamEst_SHR&OHT_TopFramingPan_OutputFormat.xlsx"
)

_MAX_KEY_MISMATCHES = 5

# Cross-sheet mapping: when generated and reference use different sheet names
# for functionally equivalent data, compare them directly.
_CROSS_SHEET_MAP = {
    "Bar Bending Schedule": "Beam - Clubhouse",   # V.B.1 BBS -> estimator BBS
}


class WorkbookDifferenceAnalyzer:
    """
    Loads both workbooks, compares sheet names, headers, row/col counts,
    and individual cell values; returns a WorkbookDiffResult.
    """

    def __init__(
        self,
        generated_path: pathlib.Path = _GENERATED,
        reference_path: pathlib.Path = _REFERENCE,
    ) -> None:
        self._gen = generated_path
        self._ref = reference_path

    def analyze(self) -> WorkbookDiffResult:
        if not self._gen.exists() or not self._ref.exists():
            missing = []
            if not self._gen.exists():
                missing.append(str(self._gen))
            if not self._ref.exists():
                missing.append(str(self._ref))
            return WorkbookDiffResult(
                generated_path=str(self._gen),
                reference_path=str(self._ref),
                generated_sheets=[],
                reference_sheets=[],
                sheet_count_match=False,
                sheet_names_match=False,
                common_sheets=[],
                missing_in_generated=missing,
                extra_in_generated=[],
                overall_match_rate_pct=0.0,
                comparison_completed=False,
            )

        wb_gen = openpyxl.load_workbook(str(self._gen), data_only=True, read_only=True)
        wb_ref = openpyxl.load_workbook(str(self._ref), data_only=True, read_only=True)

        gen_sheets = wb_gen.sheetnames
        ref_sheets = wb_ref.sheetnames

        common         = [s for s in gen_sheets if s in ref_sheets]
        missing_in_gen = [s for s in ref_sheets if s not in gen_sheets]
        extra_in_gen   = [s for s in gen_sheets if s not in ref_sheets]

        ws_diffs: List[WorksheetDiffResult] = []
        all_matching = 0
        all_total    = 0

        for sheet in common:
            diff = self._compare_sheet(wb_gen[sheet], wb_ref[sheet], sheet)
            ws_diffs.append(diff)
            all_matching += diff.matching_cells
            all_total    += diff.matching_cells + diff.mismatching_cells

        # Cross-sheet comparison: functionally equivalent sheets with different names
        for gen_name, ref_name in _CROSS_SHEET_MAP.items():
            if gen_name in gen_sheets and ref_name in ref_sheets and gen_name not in common:
                diff = self._compare_sheet(
                    wb_gen[gen_name],
                    wb_ref[ref_name],
                    f"{gen_name} vs {ref_name}",
                )
                ws_diffs.append(diff)
                all_matching += diff.matching_cells
                all_total    += diff.matching_cells + diff.mismatching_cells
                # Remove from missing/extra so the report is accurate
                if ref_name in missing_in_gen:
                    missing_in_gen.remove(ref_name)
                if gen_name in extra_in_gen:
                    extra_in_gen.remove(gen_name)

        wb_gen.close()
        wb_ref.close()

        overall = round(100 * all_matching / all_total, 4) if all_total else 0.0

        return WorkbookDiffResult(
            generated_path=str(self._gen),
            reference_path=str(self._ref),
            generated_sheets=gen_sheets,
            reference_sheets=ref_sheets,
            sheet_count_match=len(gen_sheets) == len(ref_sheets),
            sheet_names_match=(set(gen_sheets) == set(ref_sheets)),
            common_sheets=common,
            missing_in_generated=missing_in_gen,
            extra_in_generated=extra_in_gen,
            overall_match_rate_pct=overall,
            worksheet_diffs=ws_diffs,
            comparison_completed=True,
        )

    def _compare_sheet(
        self,
        ws_gen: Any,
        ws_ref: Any,
        sheet_name: str,
    ) -> WorksheetDiffResult:
        gen_rows = list(ws_gen.iter_rows(values_only=True))
        ref_rows = list(ws_ref.iter_rows(values_only=True))

        n_gen = len(gen_rows)
        n_ref = len(ref_rows)
        c_gen = max((len(r) for r in gen_rows), default=0)
        c_ref = max((len(r) for r in ref_rows), default=0)

        gen_header = self._first_nonempty(gen_rows)
        ref_header = self._first_nonempty(ref_rows)
        header_match = gen_header == ref_header

        data_rows   = min(n_gen, n_ref)
        matching    = 0
        mismatching = 0
        key_mismatches: List[Dict[str, Any]] = []

        for r_idx in range(data_rows):
            g_row = gen_rows[r_idx]
            r_row = ref_rows[r_idx]
            n_cols = max(len(g_row), len(r_row))
            for c_idx in range(n_cols):
                gv = g_row[c_idx] if c_idx < len(g_row) else None
                rv = r_row[c_idx] if c_idx < len(r_row) else None
                if self._values_equal(gv, rv):
                    matching += 1
                else:
                    mismatching += 1
                    if len(key_mismatches) < _MAX_KEY_MISMATCHES:
                        km: Dict[str, Any] = {
                            "row": r_idx + 1,
                            "col": c_idx + 1,
                            "generated": gv,
                            "reference": rv,
                            "sheet": sheet_name,
                        }
                        if gv is None:
                            km["note"] = "generated=null"
                        else:
                            try:
                                diff_pct = (
                                    abs(float(gv) - float(rv))
                                    / (abs(float(rv)) + 1e-9)
                                    * 100
                                )
                                km["diff_pct"] = round(diff_pct, 4)
                            except (TypeError, ValueError):
                                pass
                        key_mismatches.append(km)

        rate = (
            round(100 * matching / (matching + mismatching), 4)
            if (matching + mismatching)
            else 0.0
        )

        return WorksheetDiffResult(
            sheet_name=sheet_name,
            generated_rows=n_gen,
            reference_rows=n_ref,
            generated_cols=c_gen,
            reference_cols=c_ref,
            row_count_match=(n_gen == n_ref),
            col_count_match=(c_gen == c_ref),
            header_match=header_match,
            data_rows_compared=data_rows,
            matching_cells=matching,
            mismatching_cells=mismatching,
            match_rate_pct=rate,
            key_mismatches=key_mismatches,
        )

    @staticmethod
    def _first_nonempty(rows: list) -> Tuple:
        for r in rows:
            if any(c is not None for c in r):
                return tuple(r)
        return ()

    @staticmethod
    def _values_equal(a: Any, b: Any) -> bool:
        if a == b:
            return True
        if a is None and b is None:
            return True
        if a is None or b is None:
            return False
        try:
            return abs(float(a) - float(b)) < 1e-4
        except (TypeError, ValueError):
            return str(a).strip().lower() == str(b).strip().lower()
