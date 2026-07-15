"""
stage_snapshot_collector.py — Reads every pipeline stage output artefact
and builds a StageSnapshot per stage.
MODEL_VERSION: 7.1.2  |  READ-ONLY — no files written or modified.
"""

from __future__ import annotations
import json
import pathlib
from typing import Any, Dict, List, Optional

import openpyxl

from .engineering_trace_models import StageSnapshot


WORKSPACE = pathlib.Path(r"C:\Users\nishanth.h\SteelBeamEstimator")


def _load_json(path: pathlib.Path) -> Optional[dict]:
    try:
        return json.loads(path.read_text("utf-8"))
    except Exception:
        return None


def _extract_beam_ids_from_dict_keys(data: dict, path_keys: List[str]) -> List[str]:
    """Navigate nested dict via path_keys; return keys of the final dict."""
    node = data
    for k in path_keys:
        if isinstance(node, dict) and k in node:
            node = node[k]
        else:
            return []
    if isinstance(node, dict):
        return list(node.keys())
    if isinstance(node, list):
        # Try to extract beam_id field from list items
        ids = []
        for item in node:
            if isinstance(item, dict):
                bid = item.get("beam_id") or item.get("id") or item.get("beam_mark")
                if bid:
                    ids.append(str(bid))
        return ids
    return []


def _extract_scalar(data: dict, path_keys: List[str]) -> Optional[Any]:
    node = data
    for k in path_keys:
        if isinstance(node, dict) and k in node:
            node = node[k]
        else:
            return None
    return node


class StageSnapshotCollector:
    """
    Reads the primary output artefact of every configured pipeline stage
    and returns a StageSnapshot for each.
    """

    def __init__(self, stage_configs: List[dict]):
        self._stages = stage_configs

    def collect_all(self) -> Dict[str, StageSnapshot]:
        snapshots: Dict[str, StageSnapshot] = {}
        for cfg in self._stages:
            snap = self._collect_stage(cfg)
            snapshots[cfg["id"]] = snap
        return snapshots

    def _collect_stage(self, cfg: dict) -> StageSnapshot:
        stage_id   = cfg["id"]
        stage_name = cfg["name"]
        out_dir    = WORKSPACE / cfg["output_dir"]
        artefact   = cfg["primary_artefact"]
        art_path   = out_dir / artefact

        if not art_path.exists():
            # Check alternate case (Excel files)
            return StageSnapshot(
                stage_id=stage_id,
                stage_name=stage_name,
                beam_count=0,
                beam_ids=[],
                beam_uuids={},
                input_files=[],
                output_file=str(art_path),
                artefact_exists=False,
                timestamp=None,
                raw_metadata={},
                notes=[f"Artefact not found: {art_path}"],
            )

        if art_path.suffix.lower() == ".xlsx":
            return self._collect_excel(cfg, art_path)
        else:
            return self._collect_json(cfg, art_path)

    def _collect_json(self, cfg: dict, art_path: pathlib.Path) -> StageSnapshot:
        stage_id   = cfg["id"]
        stage_name = cfg["name"]
        data       = _load_json(art_path)
        notes      = []

        if data is None:
            return StageSnapshot(
                stage_id=stage_id,
                stage_name=stage_name,
                beam_count=0,
                beam_ids=[],
                beam_uuids={},
                input_files=[],
                output_file=str(art_path),
                artefact_exists=True,
                timestamp=None,
                raw_metadata={},
                notes=["Failed to parse JSON"],
            )

        beam_id_path    = cfg.get("beam_id_path", [])
        beam_count_path = cfg.get("beam_count_path", [])

        beam_ids    = _extract_beam_ids_from_dict_keys(data, beam_id_path)
        count_field = _extract_scalar(data, beam_count_path) if beam_count_path else None

        if count_field is not None and not beam_ids:
            notes.append(f"beam_count={count_field} from scalar field; beam IDs not enumerable at this stage.")

        # Try to extract UUIDs (best effort)
        beam_uuids: Dict[str, str] = {}
        node = data
        for k in beam_id_path:
            if isinstance(node, dict) and k in node:
                node = node[k]
            else:
                node = None
                break
        if isinstance(node, dict):
            for bid, bval in node.items():
                if isinstance(bval, dict):
                    uuid = (bval.get("beam_uuid")
                            or bval.get("uuid")
                            or bval.get("id", ""))
                    if uuid:
                        beam_uuids[bid] = str(uuid)

        # Determine timestamp from data
        ts = (data.get("generated_at")
              or data.get("run_timestamp")
              or data.get("timestamp")
              or data.get("discovered_at"))

        # Raw metadata (top-level scalar fields only)
        raw_meta = {k: v for k, v in data.items()
                    if not isinstance(v, (dict, list)) or k in ("beam_count", "model_count", "total_beams")}

        beam_count = len(beam_ids) if beam_ids else (int(count_field) if count_field is not None else 0)

        # Stage-specific beam ID normalisation for stages that use scalar counts
        if not beam_ids and stage_id == "L21":
            # L.2.1 uses feature database — doesn't enumerate beam IDs but has total_features
            # Pull beam IDs from features list
            features = data.get("features", [])
            seen: Dict[str, bool] = {}
            for f in features:
                bid = f.get("beam_id") or f.get("beam_mark")
                if bid:
                    seen[str(bid)] = True
            beam_ids   = sorted(seen.keys())
            beam_count = len(beam_ids) if beam_ids else (int(count_field) if count_field is not None else 0)

        if not beam_ids and stage_id == "VB1_JSON":
            notes.append("No JSON artefact — beam count sourced from Excel.")

        return StageSnapshot(
            stage_id=stage_id,
            stage_name=stage_name,
            beam_count=beam_count,
            beam_ids=beam_ids,
            beam_uuids=beam_uuids,
            input_files=[],
            output_file=str(art_path),
            artefact_exists=True,
            timestamp=str(ts) if ts else None,
            raw_metadata=raw_meta,
            notes=notes,
        )

    def _collect_excel(self, cfg: dict, art_path: pathlib.Path) -> StageSnapshot:
        stage_id   = cfg["id"]
        stage_name = cfg["name"]
        notes      = []
        beam_ids   = []

        try:
            wb = openpyxl.load_workbook(str(art_path), read_only=True, data_only=True)
            # Look for a "Beams" or summary sheet
            sheet_names = wb.sheetnames
            # Try BBS or Beam Schedule sheet first
            candidate_sheets = [s for s in sheet_names
                                 if any(k in s.upper() for k in
                                        ("BEAM", "BBS", "SCHEDULE", "ESTIM"))]
            target_sheet = candidate_sheets[0] if candidate_sheets else sheet_names[0]
            ws = wb[target_sheet]
            # Scan first column for beam IDs
            seen: Dict[str, bool] = {}
            import re
            for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
                for cell in row:
                    if cell and isinstance(cell, str):
                        if re.match(r"B\d+[A-Z]?$|BR\d+$", str(cell).strip().upper()):
                            seen[str(cell).strip().upper()] = True
            beam_ids = sorted(seen.keys())
            wb.close()
            notes.append(f"Excel sheet used: '{target_sheet}'. Beam IDs extracted from first 3 columns.")
        except Exception as exc:
            notes.append(f"Excel read error: {exc}")

        return StageSnapshot(
            stage_id=stage_id,
            stage_name=stage_name,
            beam_count=len(beam_ids),
            beam_ids=beam_ids,
            beam_uuids={},
            input_files=[],
            output_file=str(art_path),
            artefact_exists=True,
            timestamp=None,
            raw_metadata={"excel_sheets": []},
            notes=notes,
        )
