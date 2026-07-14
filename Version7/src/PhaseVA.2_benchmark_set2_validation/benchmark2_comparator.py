"""
Phase V.A.2 -- benchmark2_comparator.py
Compare the generated workbook against an estimator reference (if available).
MODEL_VERSION: 7.0.0
"""
from __future__ import annotations

import pathlib
from typing import Any, Dict, List, Optional

from benchmark2_models import WorkbookComparison, WorksheetComparison

_ROOT           = pathlib.Path(__file__).resolve().parents[3]
_V7             = _ROOT / "Version7"
_PRODUCTION_OUT = _V7 / "data/output/Production_Output"
_GENERATED_WB   = _PRODUCTION_OUT / "Estimation_Output.xlsx"

# No estimator Excel for Benchmark Set 2 -- documented as a finding
_ESTIMATOR_WB: Optional[pathlib.Path] = None


def _safe_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _compare_sheets(
    ws_gen: Any,
    ws_ref: Any,
    sheet_name: str,
) -> WorksheetComparison:
    """Cell-by-cell comparison of two openpyxl worksheets."""
    gen_rows = ws_gen.max_row or 0
    ref_rows = ws_ref.max_row or 0
    gen_cols = ws_gen.max_column or 0
    ref_cols = ws_ref.max_column or 0
    max_rows = max(gen_rows, ref_rows, 1)
    max_cols = max(gen_cols, ref_cols, 1)

    matching   = 0
    mismatching = 0
    compared   = 0

    for r in range(1, min(max_rows, 300) + 1):
        for c in range(1, min(max_cols, 30) + 1):
            gv = _safe_str(ws_gen.cell(r, c).value)
            rv = _safe_str(ws_ref.cell(r, c).value)
            compared += 1
            if gv == rv:
                matching += 1
            else:
                mismatching += 1

    match_rate = round(100 * matching / compared, 2) if compared else 0.0
    hdr_match  = True
    if gen_rows >= 1 and ref_rows >= 1:
        gen_hdr = [_safe_str(ws_gen.cell(1, c).value) for c in range(1, gen_cols + 1)]
        ref_hdr = [_safe_str(ws_ref.cell(1, c).value) for c in range(1, ref_cols + 1)]
        hdr_match = gen_hdr == ref_hdr

    return WorksheetComparison(
        sheet_name=sheet_name,
        generated_rows=gen_rows,
        reference_rows=ref_rows,
        generated_cols=gen_cols,
        reference_cols=ref_cols,
        row_count_match=(gen_rows == ref_rows),
        col_count_match=(gen_cols == ref_cols),
        header_match=hdr_match,
        data_rows_compared=compared,
        matching_cells=matching,
        mismatching_cells=mismatching,
        match_rate_pct=match_rate,
    )


class Benchmark2Comparator:
    """
    Compare the generated Estimation_Output.xlsx against the estimator reference.
    For Benchmark Set 2, the estimator reference does not exist.
    This class documents the comparison status accurately.
    """

    def __init__(
        self,
        generated_path:  Optional[pathlib.Path] = None,
        reference_path:  Optional[pathlib.Path] = _ESTIMATOR_WB,
    ) -> None:
        self._gen = generated_path or _GENERATED_WB
        self._ref = reference_path

    def compare(self) -> WorkbookComparison:
        gen_exists = self._gen.exists()
        ref_exists = self._ref is not None and self._ref.exists()

        if not ref_exists:
            # Benchmark Set 2 has no estimator reference -- document this
            gen_sheets: List[str] = []
            if gen_exists:
                try:
                    import openpyxl  # type: ignore
                    wb = openpyxl.load_workbook(str(self._gen), data_only=True)
                    gen_sheets = wb.sheetnames
                except Exception:
                    pass
            return WorkbookComparison(
                generated_path=str(self._gen),
                reference_path=str(self._ref) if self._ref else "NOT_PROVIDED",
                reference_exists=False,
                generated_sheets=gen_sheets,
                reference_sheets=[],
                common_sheets=[],
                overall_match_rate_pct=0.0,
                worksheet_comparisons=[],
                comparison_completed=True,
                note=(
                    "No estimator workbook available for Benchmark Set 2 (Galera GF drawings). "
                    "Workbook comparison is SKIPPED. "
                    "This is documented as a generalization finding: "
                    "an estimator reference workbook is required for full validation."
                ),
            )

        if not gen_exists:
            return WorkbookComparison(
                generated_path=str(self._gen),
                reference_path=str(self._ref),
                reference_exists=True,
                generated_sheets=[],
                reference_sheets=[],
                common_sheets=[],
                overall_match_rate_pct=0.0,
                worksheet_comparisons=[],
                comparison_completed=False,
                note="Generated workbook does not exist -- pipeline did not produce output",
            )

        try:
            import openpyxl  # type: ignore
            wb_gen = openpyxl.load_workbook(str(self._gen), data_only=True)
            wb_ref = openpyxl.load_workbook(str(self._ref), data_only=True)
        except Exception as exc:
            return WorkbookComparison(
                generated_path=str(self._gen),
                reference_path=str(self._ref),
                reference_exists=True,
                generated_sheets=[],
                reference_sheets=[],
                common_sheets=[],
                overall_match_rate_pct=0.0,
                worksheet_comparisons=[],
                comparison_completed=False,
                note=f"Failed to load workbooks: {exc}",
            )

        gen_sheets = wb_gen.sheetnames
        ref_sheets = wb_ref.sheetnames
        common     = [s for s in gen_sheets if s in ref_sheets]

        ws_comparisons: List[WorksheetComparison] = []
        all_rates: List[float] = []

        for sname in common:
            wsc = _compare_sheets(wb_gen[sname], wb_ref[sname], sname)
            ws_comparisons.append(wsc)
            all_rates.append(wsc.match_rate_pct)

        overall = round(sum(all_rates) / len(all_rates), 2) if all_rates else 0.0

        return WorkbookComparison(
            generated_path=str(self._gen),
            reference_path=str(self._ref),
            reference_exists=True,
            generated_sheets=gen_sheets,
            reference_sheets=ref_sheets,
            common_sheets=common,
            overall_match_rate_pct=overall,
            worksheet_comparisons=ws_comparisons,
            comparison_completed=True,
            note="",
        )
