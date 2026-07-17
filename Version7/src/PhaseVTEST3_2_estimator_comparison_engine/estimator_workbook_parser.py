"""
estimator_workbook_parser.py — Parse official estimator workbook structure.
MODEL_VERSION: 8.1.3

Structure recognised:
  STEP 1: Reinforcement Total (Pink) ABSTRACT table — row 29 headers (TOTAL-MT + kg)
  STEP 2: Beam-wise blocks below detail header ~36, beam IDs in column C
"""
from __future__ import annotations

import re
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

import openpyxl

from comparison_models import DIAMETER_MM, ROLE_PATTERNS, BeamBlock, ProjectSummary, RoleLine, WorkbookRef

_BEAM_ID_RE = re.compile(r"^B\d+[A-Z]?$", re.I)
_SUMMARY_LABELS = ("total - ch- tf", "total - ch-tf", "total - tf", "total ch tf", "terrace floor")

# Detail-section diameter kg columns (header row 36) — beam role lines only
_DETAIL_DIA_COLS = {8: 17, 10: 18, 12: 19, 16: 20, 20: 21, 25: 22, 32: 23}
_DETAIL_STEEL_COL = 24

_STEEL_ROUNDING_TOLERANCE_KG = 1.0
_MT_KG_TOLERANCE_PCT = 0.001


@dataclass
class PinkColumnMap:
    header_row: int
    concrete_col: Optional[int]
    shuttering_col: Optional[int]
    dia_mt_cols: Dict[int, int]
    total_mt_col: Optional[int]
    kg_col: Optional[int]


def _num(v) -> float:
    try:
        if v is None or v == "":
            return 0.0
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _norm_role(desc: str) -> str:
    d = (desc or "").strip().lower()
    for role, patterns in ROLE_PATTERNS.items():
        for pat in patterns:
            if re.search(pat, d, re.I):
                return role
    if d:
        return "Unknown"
    return "Unknown"


def _dia_mm_from_header(text: str) -> Optional[int]:
    t = (text or "").strip().lower()
    m = re.search(r"(\d+)\s*mm", t)
    if m:
        d = int(m.group(1))
        return d if d in DIAMETER_MM else None
    if t.isdigit():
        d = int(t)
        return d if d in DIAMETER_MM else None
    return None


def _find_pink_header_row(ws) -> Optional[PinkColumnMap]:
    """Locate ABSTRACT pink table header row containing TOTAL-MT and kg."""
    for r in range(1, min(60, ws.max_row + 1)):
        total_mt_col: Optional[int] = None
        kg_col: Optional[int] = None
        for c in range(1, min(30, ws.max_column + 1)):
            val = str(ws.cell(r, c).value or "").strip().upper()
            if val == "TOTAL-MT":
                total_mt_col = c
            elif val == "KG":
                kg_col = c
        if not (total_mt_col and kg_col):
            continue

        concrete_col: Optional[int] = None
        shuttering_col: Optional[int] = None
        dia_mt_cols: Dict[int, int] = {}
        for c in range(1, min(30, ws.max_column + 1)):
            hdr = ws.cell(r, c).value
            if not hdr:
                continue
            hs = str(hdr).strip().lower()
            if hs == "concrete":
                concrete_col = c
            elif hs == "shuttering":
                shuttering_col = c
            else:
                dia = _dia_mm_from_header(str(hdr))
                if dia:
                    dia_mt_cols[dia] = c

        return PinkColumnMap(
            header_row=r,
            concrete_col=concrete_col,
            shuttering_col=shuttering_col,
            dia_mt_cols=dia_mt_cols,
            total_mt_col=total_mt_col,
            kg_col=kg_col,
        )
    return None


def _find_pink_data_row(ws, col_map: PinkColumnMap) -> Optional[Tuple[int, str]]:
    """Find the primary pink summary data row (highest kg in ABSTRACT section)."""
    end_row = col_map.header_row + 15
    for r in range(col_map.header_row + 1, min(end_row, ws.max_row + 1)):
        c1 = str(ws.cell(r, 1).value or "").strip().lower()
        if "breakup" in c1 or "quantity breakup" in c1:
            end_row = r
            break

    candidates: List[Tuple[int, str, float, int]] = []
    for r in range(col_map.header_row + 1, end_row):
        label = str(ws.cell(r, 3).value or ws.cell(r, 4).value or "").strip()
        if not label or label.lower() in ("description", "clubhouse"):
            continue
        kg = _num(ws.cell(r, col_map.kg_col).value) if col_map.kg_col else 0.0
        if kg <= 0:
            continue
        priority = 0
        ll = label.lower()
        if "terrace floor" in ll:
            priority = 3
        elif any(lbl in ll for lbl in _SUMMARY_LABELS):
            priority = 2
        candidates.append((r, label, kg, priority))

    if not candidates:
        return None

    candidates.sort(key=lambda x: (x[3], x[2]), reverse=True)
    row, label, _, _ = candidates[0]
    return row, label


def _resolve_project_steel(
    total_mt: float, kg: float, has_kg_col: bool
) -> Tuple[float, str, List[str]]:
    """Resolve canonical project steel — never sum TOTAL-MT×1000 and kg."""
    warnings: List[str] = []
    if has_kg_col and kg > 0:
        canonical = round(kg, 3)
        source = "kg_column"
    elif total_mt > 0:
        canonical = round(total_mt * 1000.0, 3)
        source = "total_mt_converted"
    else:
        return 0.0, "none", ["No kg or TOTAL-MT value found in pink summary row."]

    if total_mt > 0 and kg > 0:
        expected_kg = total_mt * 1000.0
        diff = abs(expected_kg - kg)
        tol = max(_STEEL_ROUNDING_TOLERANCE_KG, kg * _MT_KG_TOLERANCE_PCT)
        if diff > tol:
            warnings.append(
                f"TOTAL-MT×1000 ({expected_kg:.2f} kg) differs from kg column "
                f"({kg:.2f} kg) by {diff:.2f} kg — using kg column as canonical."
            )
    return canonical, source, warnings


class EstimatorWorkbookParser:

    def __init__(self, path):
        self._path = path
        self._wb = openpyxl.load_workbook(path, data_only=True)
        self._ws = self._wb.active
        self._pink_map: Optional[PinkColumnMap] = None
        self._summary_validation: Dict[str, Any] = {}

    def close(self):
        self._wb.close()

    def workbook_ref(self) -> WorkbookRef:
        return WorkbookRef(
            path=str(self._path),
            filename=self._path.name,
            sheet_names=self._wb.sheetnames,
            size_bytes=self._path.stat().st_size,
        )

    @property
    def summary_validation(self) -> Dict[str, Any]:
        return self._summary_validation

    def find_summary_table(self) -> Optional[ProjectSummary]:
        """Locate Reinforcement Total (Pink) summary using TOTAL-MT / kg columns."""
        self._pink_map = _find_pink_header_row(self._ws)
        if not self._pink_map:
            return self._fallback_summary_from_detail()

        col_map = self._pink_map
        found = _find_pink_data_row(self._ws, col_map)
        if not found:
            return self._fallback_summary_from_detail()

        data_row, label = found
        warnings: List[str] = []

        dia_kg: Dict[int, float] = {}
        dia_mt: Dict[int, float] = {}
        for dia, col in col_map.dia_mt_cols.items():
            mt = _num(self._ws.cell(data_row, col).value)
            dia_mt[dia] = round(mt, 6)
            dia_kg[dia] = round(mt * 1000.0, 3)

        total_mt = _num(self._ws.cell(data_row, col_map.total_mt_col).value) if col_map.total_mt_col else 0.0
        kg = _num(self._ws.cell(data_row, col_map.kg_col).value) if col_map.kg_col else 0.0

        total_steel_kg, source, steel_warnings = _resolve_project_steel(
            total_mt, kg, col_map.kg_col is not None
        )
        warnings.extend(steel_warnings)

        sum_dia_kg = round(sum(dia_kg.values()), 3)
        if total_steel_kg > 0 and abs(sum_dia_kg - total_steel_kg) > max(1.0, total_steel_kg * 0.01):
            warnings.append(
                f"Sum of diameter MT×1000 ({sum_dia_kg:.2f} kg) differs from project total "
                f"({total_steel_kg:.2f} kg) by {abs(sum_dia_kg - total_steel_kg):.2f} kg."
            )

        self._summary_validation = self._build_summary_validation(
            data_row, col_map, total_mt, kg, total_steel_kg, source, dia_kg, warnings
        )

        return ProjectSummary(
            label=label,
            concrete_m3=round(_num(self._ws.cell(data_row, col_map.concrete_col).value), 4)
            if col_map.concrete_col else 0.0,
            shuttering_m2=round(_num(self._ws.cell(data_row, col_map.shuttering_col).value), 4)
            if col_map.shuttering_col else 0.0,
            diameter_kg=dia_kg,
            total_steel_kg=total_steel_kg,
            source_row=data_row,
            total_steel_mt=round(total_mt, 6),
            total_steel_source=source,
            diameter_mt=dia_mt,
            parser_warnings=warnings,
        )

    def _fallback_summary_from_detail(self) -> Optional[ProjectSummary]:
        """Fallback: Total - CH- TF row in detail section — diameters in MT at C17-C23."""
        for r in range(1, self._ws.max_row + 1):
            c3 = self._ws.cell(r, 3).value
            if not c3:
                continue
            label = str(c3).strip().lower()
            if not (any(lbl in label for lbl in _SUMMARY_LABELS) or label.startswith("total - ch")):
                continue

            dia_kg: Dict[int, float] = {}
            dia_mt: Dict[int, float] = {}
            for dia, col in _DETAIL_DIA_COLS.items():
                mt = _num(self._ws.cell(r, col).value)
                dia_mt[dia] = round(mt, 6)
                dia_kg[dia] = round(mt * 1000.0, 3)

            sum_dia_kg = round(sum(dia_kg.values()), 3)
            warnings = [
                "Pink ABSTRACT table not found — using detail-section diameter MT columns; "
                "total steel derived from sum(diameter MT)×1000 (C24/detail Steel column ignored)."
            ]
            self._summary_validation = {
                "fallback_used": True,
                "source_row": r,
                "total_steel_kg": sum_dia_kg,
                "total_steel_source": "diameter_mt_sum",
                "warnings": warnings,
            }
            return ProjectSummary(
                label=str(c3).strip(),
                concrete_m3=round(_num(self._ws.cell(r, 9).value), 4),
                shuttering_m2=round(_num(self._ws.cell(r, 10).value), 4),
                diameter_kg=dia_kg,
                total_steel_kg=sum_dia_kg,
                source_row=r,
                total_steel_mt=round(sum_dia_kg / 1000.0, 6),
                total_steel_source="diameter_mt_sum",
                diameter_mt=dia_mt,
                parser_warnings=warnings,
            )
        return None

    def _build_summary_validation(
        self,
        data_row: int,
        col_map: PinkColumnMap,
        total_mt: float,
        kg: float,
        total_steel_kg: float,
        source: str,
        dia_kg: Dict[int, float],
        warnings: List[str],
    ) -> Dict[str, Any]:
        mt_x1000 = round(total_mt * 1000.0, 3)
        kg_diff = abs(mt_x1000 - kg) if kg > 0 and total_mt > 0 else 0.0
        return {
            "pink_header_row": col_map.header_row,
            "data_row": data_row,
            "total_mt": round(total_mt, 6),
            "kg_column": round(kg, 3),
            "total_steel_kg": total_steel_kg,
            "total_steel_source": source,
            "mt_x1000": mt_x1000,
            "mt_kg_difference": round(kg_diff, 3),
            "mt_kg_consistent": kg_diff <= max(_STEEL_ROUNDING_TOLERANCE_KG, kg * _MT_KG_TOLERANCE_PCT)
            if kg > 0 and total_mt > 0 else True,
            "diameter_kg": dia_kg,
            "diameter_parsed_once": True,
            "duplicate_total_avoided": source == "kg_column" or total_mt > 0,
            "warnings": warnings,
        }

    def find_beam_section_start(self) -> int:
        for r in range(1, min(80, self._ws.max_row + 1)):
            c3 = self._ws.cell(r, 3).value
            c16 = self._ws.cell(r, 16).value
            if c3 and "description" in str(c3).lower() and c16 and "total length" in str(c16).lower():
                return r
        return 36

    def parse_beam_blocks(self) -> List[BeamBlock]:
        header_row = self.find_beam_section_start()
        blocks: List[BeamBlock] = []
        r = header_row + 2
        current: Optional[BeamBlock] = None

        while r <= self._ws.max_row:
            c3 = self._ws.cell(r, 3).value
            if c3 is None:
                r += 1
                continue

            c3s = str(c3).strip()
            if _BEAM_ID_RE.match(c3s):
                if current:
                    self._finalize_block(current)
                    blocks.append(current)
                current = BeamBlock(
                    beam_id=c3s.upper(),
                    start_row=r,
                    end_row=r,
                    concrete_m3=round(_num(self._ws.cell(r, 9).value), 4),
                    shuttering_m2=round(_num(self._ws.cell(r, 10).value), 4),
                )
                r += 1
                continue

            if current is None:
                r += 1
                continue

            desc = c3s
            role = _norm_role(desc)
            # Detail-section role lines: diameter columns are kg, not MT
            dia_kg = {
                d: round(_num(self._ws.cell(r, col).value), 4)
                for d, col in _DETAIL_DIA_COLS.items()
            }
            steel = _num(self._ws.cell(r, _DETAIL_STEEL_COL).value)
            if steel <= 0:
                steel = sum(dia_kg.values())

            line = RoleLine(
                role=role,
                description=desc,
                diameter_mm=_num(self._ws.cell(r, 4).value) or None,
                spacing_m=_num(self._ws.cell(r, 5).value) or None,
                bar_count=_num(self._ws.cell(r, 6).value) or None,
                cut_length_m=_num(self._ws.cell(r, 15).value) or None,
                total_length_m=_num(self._ws.cell(r, 16).value) or None,
                steel_kg=round(steel, 4),
                diameter_kg=dia_kg,
            )
            current.lines.append(line)
            current.end_row = r
            r += 1

        if current:
            self._finalize_block(current)
            blocks.append(current)

        return blocks

    @staticmethod
    def _finalize_block(block: BeamBlock):
        dia_totals: Dict[int, float] = {d: 0.0 for d in DIAMETER_MM}
        for line in block.lines:
            block.total_steel_kg += line.steel_kg
            for d, kg in line.diameter_kg.items():
                dia_totals[d] = dia_totals.get(d, 0.0) + kg
        block.diameter_kg = {d: round(v, 4) for d, v in dia_totals.items() if v > 0}
        block.total_steel_kg = round(block.total_steel_kg, 4)


def discover_estimator_workbook(folder) -> Optional[Any]:
    import pathlib
    folder = pathlib.Path(folder)
    if not folder.exists():
        return None
    candidates = sorted(
        [p for p in folder.glob("*.xlsx") if not p.name.startswith("~$")],
        key=lambda p: p.stat().st_size,
        reverse=True,
    )
    return candidates[0] if candidates else None
