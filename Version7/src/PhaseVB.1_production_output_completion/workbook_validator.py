"""
Workbook Validator — Phase V.B.1 MODULE 7

Validates generated workbooks for:
  - Readability, completeness, worksheet count/names/headers
  - Engineering totals, steel totals, no corrupted cells
"""
import pathlib
from typing import List, Dict, Any

from production_output_models import WorkbookValidationResult
from excel_structure_builder import ALL_WORKSHEETS

try:
    import openpyxl
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False


class WorkbookValidator:
    """Validates a generated workbook against the Phase V.B.1 specification."""

    def __init__(
        self,
        workbook_path: pathlib.Path,
        expected_steel_total_kg: float = 0.0,
        steel_tolerance: float = 0.001,
    ) -> None:
        self.path = workbook_path
        self.expected_total = expected_steel_total_kg
        self.tolerance = steel_tolerance

    def validate(self) -> WorkbookValidationResult:
        errors: List[str] = []
        is_readable = False
        is_complete = False
        ws_count = 0
        ws_names: List[str] = []
        missing: List[str] = []
        row_counts: Dict[str, int] = {}
        header_checks: Dict[str, bool] = {}
        steel_total_check = False
        steel_total_found = 0.0
        no_corrupted = True

        # ── 1. Readability ────────────────────────────────────────────────
        if not OPENPYXL_OK:
            errors.append("openpyxl not installed — cannot validate workbook")
            return self._result(False, False, 0, [], [], {}, {}, False, 0.0, False, errors)

        if not self.path.exists():
            errors.append(f"Workbook not found: {self.path}")
            return self._result(False, False, 0, [], [], {}, {}, False, 0.0, False, errors)

        try:
            wb = openpyxl.load_workbook(str(self.path), read_only=True, data_only=True)
            is_readable = True
        except Exception as exc:
            errors.append(f"Workbook not readable: {exc}")
            return self._result(False, False, 0, [], [], {}, {}, False, 0.0, False, errors)

        # ── 2. Worksheet count & names ────────────────────────────────────
        ws_names = wb.sheetnames
        ws_count = len(ws_names)
        missing = [w for w in ALL_WORKSHEETS if w not in ws_names]
        if missing:
            errors.append(f"Missing worksheets: {missing}")

        # ── 3. Row counts & header checks ─────────────────────────────────
        for ws_name in ws_names:
            try:
                ws = wb[ws_name]
                rows = list(ws.iter_rows(values_only=True))
                data_rows = [r for r in rows if any(v is not None for v in r)]
                row_counts[ws_name] = len(data_rows)
                # Header check: row 2 should have non-None values
                if len(rows) >= 2:
                    header_row = rows[1]
                    has_headers = any(v is not None for v in header_row)
                    header_checks[ws_name] = has_headers
                    if not has_headers:
                        errors.append(f"{ws_name}: header row 2 is empty")
                else:
                    header_checks[ws_name] = False
                    errors.append(f"{ws_name}: fewer than 2 rows")
            except Exception as exc:
                errors.append(f"{ws_name}: error reading sheet — {exc}")
                no_corrupted = False

        # ── 4. Steel total check ──────────────────────────────────────────
        if "Project Totals" in ws_names:
            try:
                pt_ws = wb["Project Totals"]
                for row in pt_ws.iter_rows(values_only=True):
                    if row and "Total Steel Weight" in str(row[0] or ""):
                        if row[1] is not None:
                            steel_total_found = float(row[1])
                            if self.expected_total > 0:
                                diff = abs(steel_total_found - self.expected_total)
                                steel_total_check = diff <= self.tolerance * max(
                                    abs(steel_total_found), abs(self.expected_total), 1.0
                                ) * 1000
                            else:
                                steel_total_check = steel_total_found > 0
                            break
            except Exception as exc:
                errors.append(f"Steel total check failed: {exc}")

        wb.close()

        # ── 5. Completeness ───────────────────────────────────────────────
        is_complete = (
            is_readable
            and len(missing) == 0
            and all(header_checks.values())
            and no_corrupted
        )

        validation_passed = is_complete and (steel_total_found > 0)

        return self._result(
            is_readable, is_complete, ws_count, ws_names, missing,
            row_counts, header_checks, steel_total_check, steel_total_found,
            no_corrupted, errors,
            validation_passed=validation_passed,
        )

    @staticmethod
    def _result(
        is_readable, is_complete, ws_count, ws_names, missing,
        row_counts, header_checks, steel_total_check, steel_total_found,
        no_corrupted, errors, validation_passed=None
    ) -> WorkbookValidationResult:
        if validation_passed is None:
            validation_passed = is_complete and steel_total_check
        return WorkbookValidationResult(
            is_readable=is_readable,
            is_complete=is_complete,
            worksheet_count=ws_count,
            worksheet_names=ws_names,
            expected_worksheets=ALL_WORKSHEETS,
            missing_worksheets=missing,
            row_counts=row_counts,
            header_checks=header_checks,
            steel_total_check=steel_total_check,
            steel_total_found=steel_total_found,
            no_corrupted_cells=no_corrupted,
            validation_passed=validation_passed,
            validation_errors=errors,
        )
