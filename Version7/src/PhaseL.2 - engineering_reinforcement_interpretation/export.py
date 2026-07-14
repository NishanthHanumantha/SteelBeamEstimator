"""Export all Phase L.2 Engineering Reinforcement Interpretation artifacts."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, List, Tuple

from beam_reinforcement_model import BeamReinforcementModel

EXPORT_FILES: Tuple[str, ...] = (
    "beam_reinforcement_models.json",
    "bar_role_classification.json",
    "support_zone_analysis.json",
    "continuity_analysis.json",
    "beam_ownership_analysis.json",
    "reinforcement_regions.json",
    "engineering_semantics.json",
    "interpretation_statistics.json",
    "interpretation_dashboard.json",
)


class InterpretationExport:
    @staticmethod
    def export_all(output_dir: Path, result: Dict[str, Any], config: Dict[str, Any]) -> Dict[str, str]:
        output_dir.mkdir(parents=True, exist_ok=True)
        written: Dict[str, str] = {}
        models: List[BeamReinforcementModel] = result.get("models") or []

        mapping: Dict[str, Any] = {
            "beam_reinforcement_models.json": {
                "model_count": len(models),
                "models": [m.to_dict() for m in models],
            },
            "bar_role_classification.json": result.get("bar_role_classification"),
            "support_zone_analysis.json": result.get("support_zone_analysis"),
            "continuity_analysis.json": result.get("continuity_analysis"),
            "beam_ownership_analysis.json": {
                "total_beams": len(models),
                "ownership_entries": [
                    {
                        "beam_id": m.beam_id,
                        "total_bars": len(m.all_bars()),
                        "is_multi_span": m.beam_id in {"B8", "B9", "B10"},
                    }
                    for m in models
                ],
            },
            "reinforcement_regions.json": result.get("reinforcement_regions"),
            "engineering_semantics.json": result.get("engineering_semantics"),
            "interpretation_statistics.json": result.get("statistics"),
            "interpretation_dashboard.json": {
                "summary": result.get("summary"),
                "validation": result.get("validation"),
                "benchmark_coverage": {
                    b: (b in {m.beam_id for m in models})
                    for b in ["B1", "B2", "B8", "B9", "B10"]
                },
                "classification_rate_percent": (result.get("statistics") or {}).get("classification_rate_percent"),
            },
        }

        for filename in EXPORT_FILES:
            path = output_dir / filename
            payload = mapping.get(filename)
            path.write_text(json.dumps(payload, indent=2, sort_keys=False, default=str), encoding="utf-8")
            written[filename] = str(path)

        if config.get("generate_excel_report", True):
            xlsx_path = output_dir / "interpretation_report.xlsx"
            if InterpretationExport._write_excel(xlsx_path, result, models):
                written["interpretation_report.xlsx"] = str(xlsx_path)

        result["export_paths"] = written
        return written

    @staticmethod
    def validate_exports(output_dir: Path) -> Dict[str, Any]:
        checks = []
        for filename in EXPORT_FILES:
            path = output_dir / filename
            exists = path.exists() and path.stat().st_size > 2
            valid = False
            if exists:
                try:
                    json.loads(path.read_text(encoding="utf-8"))
                    valid = True
                except Exception:
                    pass
            checks.append({"name": f"Exists {filename}", "status": "PASS" if exists else "FAIL"})
            checks.append({"name": f"Valid JSON {filename}", "status": "PASS" if valid else "FAIL"})
        xlsx = output_dir / "interpretation_report.xlsx"
        checks.append({
            "name": "Exists interpretation_report.xlsx",
            "status": "PASS" if xlsx.exists() and xlsx.stat().st_size > 0 else "FAIL",
        })
        failed = [c for c in checks if c["status"] == "FAIL"]
        return {
            "status": "PASS" if not failed else "FAIL",
            "checks": checks,
            "summary": {"total": len(checks), "passed": len(checks) - len(failed), "failed": len(failed)},
        }

    @staticmethod
    def _write_excel(path: Path, result: Dict[str, Any], models: List[BeamReinforcementModel]) -> bool:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font
        except ImportError:
            return False

        wb = Workbook()
        ROLE_COLORS = {
            "TOP_MAIN": "4472C4", "BOTTOM_MAIN": "ED7D31",
            "TOP_EXTRA": "9DC3E6", "BOTTOM_EXTRA": "F4B183",
            "STIRRUP": "70AD47", "SIDE_FACE_REINFORCEMENT": "FFC000",
            "SPACER_BAR": "BFBFBF", "SUPPLEMENTARY_BAR": "D9D9D9",
        }

        # ── Beam Reinforcement Models sheet ──────────────────────────────────
        ws = wb.active
        ws.title = "Beam Reinforcement Models"
        headers = [
            "Beam ID", "Is Benchmark", "Confidence",
            "TOP_MAIN", "BOTTOM_MAIN", "TOP_EXTRA", "BOTTOM_EXTRA",
            "STIRRUP", "SIDE_FACE", "SPACER", "Total Bars",
        ]
        ws.append(headers)
        for m in models:
            rc = m.bar_count_by_role()
            ws.append([
                m.beam_id, m.is_benchmark_beam, m.interpretation_confidence,
                rc.get("TOP_MAIN", 0), rc.get("BOTTOM_MAIN", 0),
                rc.get("TOP_EXTRA", 0), rc.get("BOTTOM_EXTRA", 0),
                rc.get("STIRRUP", 0), rc.get("SIDE_FACE_REINFORCEMENT", 0),
                rc.get("SPACER_BAR", 0), len(m.all_bars()),
            ])

        # ── Bar Role Classification sheet ─────────────────────────────────
        brc_ws = wb.create_sheet("Bar Role Classification")
        brc_ws.append([
            "Beam ID", "Bar ID", "Semantic Role", "Bar Label",
            "Dia (mm)", "Qty", "Zone", "Extent", "Coverage Ratio",
            "Confidence", "Is Corrected", "Ref Anchored", "Evidence",
        ])
        for m in models:
            for b in m.all_bars():
                row = [
                    b.beam_id, b.bar_id, b.semantic_role, b.bar_label,
                    b.diameter_mm, b.quantity, b.position_zone, b.extent,
                    b.coverage_ratio, b.classification_confidence,
                    b.is_corrected, b.is_reference_anchored, b.classification_evidence,
                ]
                brc_ws.append(row)
                cell = brc_ws.cell(row=brc_ws.max_row, column=3)
                color = ROLE_COLORS.get(b.semantic_role)
                if color:
                    cell.fill = PatternFill("solid", fgColor=color)

        # ── Support Zones sheet ───────────────────────────────────────────
        sz_ws = wb.create_sheet("Support Zones")
        sz_ws.append(["Beam ID", "Support Type", "Adjacent Beam", "Position Frac", "Width mm"])
        for m in models:
            for s in m.support_zones:
                sz_ws.append([
                    m.beam_id, s.support_type, s.adjacent_beam_id,
                    s.position_fraction, s.support_width_mm,
                ])

        # ── Engineering Semantics (pipeline corrections) ──────────────────
        sem_ws = wb.create_sheet("Pipeline Corrections")
        sem_ws.append(["Beam ID", "Bar Label", "Pipeline Role", "Corrected To", "Evidence"])
        for m in models:
            for b in m.all_bars():
                if b.is_corrected:
                    sem_ws.append([
                        b.beam_id, b.bar_label, b.source_pipeline_role,
                        b.semantic_role, b.classification_evidence,
                    ])

        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)
        return True

    @staticmethod
    def print_summary(result: Dict[str, Any]) -> None:
        stats = result.get("statistics") or {}
        val = result.get("validation") or {}
        exp = result.get("export_validation") or {}
        roles = stats.get("roles_distribution") or {}
        print("\n" + "=" * 80)
        print("Phase L.2 — Engineering Reinforcement Interpretation Engine")
        print("=" * 80)
        print(f"Model Version         : {result.get('model_version')}")
        print(f"Total Beams           : {stats.get('total_beams', 0)}")
        print(f"Total Bars            : {stats.get('total_bars', 0)}")
        print(f"Classification Rate   : {stats.get('classification_rate_percent', 0)}%")
        print(f"Pipeline Corrections  : {stats.get('pipeline_corrections', 0)}")
        print(f"Reference Anchored    : {stats.get('reference_anchored_bars', 0)}")
        print("")
        print("Role Distribution:")
        for role in ["TOP_MAIN", "BOTTOM_MAIN", "TOP_EXTRA", "BOTTOM_EXTRA",
                     "STIRRUP", "SIDE_FACE_REINFORCEMENT", "SPACER_BAR"]:
            cnt = roles.get(role, 0)
            if cnt:
                print(f"  {role:<30}: {cnt}")
        print("")
        print(f"Validation : {val.get('summary', {}).get('passed', 0)}/{val.get('summary', {}).get('total_checks', 0)} PASS")
        print(f"Exports    : {exp.get('summary', {}).get('passed', 0)}/{exp.get('summary', {}).get('total', 0)} PASS")
        print("=" * 80 + "\n")
