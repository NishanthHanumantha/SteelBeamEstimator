"""Excel output consumption trace — READ-ONLY."""
from __future__ import annotations
from typing import Any, Dict, List, Optional

from .engineering_bar_loader import EngineeringBarLoader

WS_BEAM = "Beam Summary"
WS_DIAM = "Diameter Summary"
WS_TOTALS = "Project Totals"
WS_BBS = "Bar Bending Schedule"


class ExcelTrace:

    def trace(self, loader: EngineeringBarLoader) -> Dict[str, Any]:
        generated = self._read_workbook(loader.workbook_path)
        reference = self._read_workbook(loader.reference_workbook_path)

        comparison = {}
        if generated and reference:
            comparison = self._compare_workbooks(generated, reference)

        pipeline_consistency = {
            "workbook_exists": loader.workbook_path is not None,
            "bbs_rows_in_json": loader.bbs_summary_json.get("total_bbs_rows", 0),
            "bbs_rows_computed": len(loader.bbs_rows_computed or []),
            "steel_total_json": loader.steel_summary_json.get("total_weight_kg", 0),
            "excel_project_total_kg": generated.get("project_total_kg"),
            "excel_diameter_totals": generated.get("diameter_totals", {}),
        }

        return {
            "generated_workbook": str(loader.workbook_path) if loader.workbook_path else None,
            "reference_workbook": str(loader.reference_workbook_path)
            if loader.reference_workbook_path else None,
            "generated": generated,
            "reference": reference,
            "reference_comparison": comparison,
            "pipeline_consistency": pipeline_consistency,
            "trace_chain": "EngineeringBar -> Steel Weight -> BBS -> Excel",
        }

    def _read_workbook(self, path: Optional[Any]) -> Dict[str, Any]:
        if not path or not path.exists():
            return {}
        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
        except Exception as exc:
            return {"error": str(exc)}

        result: Dict[str, Any] = {"sheets": list(wb.sheetnames)}

        if WS_BEAM in wb.sheetnames:
            ws = wb[WS_BEAM]
            beams = []
            for row in ws.iter_rows(min_row=3, values_only=True):
                if row and row[0]:
                    beams.append({
                        "beam_id": str(row[0]),
                        "steel_kg": float(row[5] or 0) if len(row) > 5 else 0,
                    })
            result["beam_summary"] = beams
            result["beam_total_kg"] = round(sum(b["steel_kg"] for b in beams), 3)

        if WS_DIAM in wb.sheetnames:
            ws = wb[WS_DIAM]
            dia_totals = {}
            for row in ws.iter_rows(min_row=3, values_only=True):
                if row and row[0] and str(row[0]).startswith("Y"):
                    try:
                        dia = int(str(row[0]).replace("Y", ""))
                        dia_totals[dia] = {
                            "bars": int(row[1] or 0),
                            "weight_kg": float(row[3] or 0),
                        }
                    except (ValueError, TypeError):
                        pass
            result["diameter_totals"] = dia_totals

        if WS_TOTALS in wb.sheetnames:
            ws = wb[WS_TOTALS]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0] and "steel" in str(row[0]).lower():
                    try:
                        result["project_total_kg"] = float(row[1] or 0)
                    except (ValueError, TypeError):
                        pass

        if WS_BBS in wb.sheetnames:
            ws = wb[WS_BBS]
            eng_rows = 0
            for row in ws.iter_rows(min_row=3, values_only=True):
                if row and row[2] and row[2] not in ("Description",):
                    desc = str(row[2])
                    if not desc.startswith("B") or len(desc) > 4:
                        eng_rows += 1
            result["bbs_engineering_rows"] = eng_rows

        wb.close()
        return result

    def _compare_workbooks(
        self, generated: Dict[str, Any], reference: Dict[str, Any]
    ) -> Dict[str, Any]:
        gen_total = generated.get("project_total_kg", 0)
        ref_total = reference.get("project_total_kg", 0)
        gen_dia = generated.get("diameter_totals", {})
        ref_dia = reference.get("diameter_totals", {})

        dia_mismatches = []
        all_dias = sorted(set(gen_dia) | set(ref_dia))
        for dia in all_dias:
            g = gen_dia.get(dia, {})
            r = ref_dia.get(dia, {})
            g_bars = g.get("bars", 0)
            r_bars = r.get("bars", 0)
            g_w = g.get("weight_kg", 0)
            r_w = r.get("weight_kg", 0)
            if g_bars != r_bars or abs(g_w - r_w) > 0.5:
                dia_mismatches.append({
                    "diameter_mm": dia,
                    "generated_bars": g_bars,
                    "reference_bars": r_bars,
                    "bar_delta": g_bars - r_bars,
                    "generated_weight_kg": g_w,
                    "reference_weight_kg": r_w,
                    "weight_delta_kg": round(g_w - r_w, 3),
                })

        gen_beams = {b["beam_id"]: b["steel_kg"] for b in generated.get("beam_summary", [])}
        ref_beams = {b["beam_id"]: b["steel_kg"] for b in reference.get("beam_summary", [])}
        beam_mismatches = []
        for bid in sorted(set(gen_beams) | set(ref_beams)):
            gw = gen_beams.get(bid, 0)
            rw = ref_beams.get(bid, 0)
            if abs(gw - rw) > 0.5:
                beam_mismatches.append({
                    "beam_id": bid,
                    "generated_kg": gw,
                    "reference_kg": rw,
                    "delta_kg": round(gw - rw, 3),
                })

        return {
            "project_total_delta_kg": round((gen_total or 0) - (ref_total or 0), 3),
            "generated_total_kg": gen_total,
            "reference_total_kg": ref_total,
            "diameter_mismatches": dia_mismatches,
            "beam_mismatches": beam_mismatches,
            "diameter_mismatch_count": len(dia_mismatches),
            "beam_mismatch_count": len(beam_mismatches),
        }
