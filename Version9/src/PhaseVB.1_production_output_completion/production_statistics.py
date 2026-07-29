"""
Production Statistics — Phase V.B.1 MODULE 8

Collects and aggregates production metrics:
  - Execution time, rows/columns generated, engineering rows
  - Steel totals, beam totals, BBS rows, worksheet statistics
"""
import time
from typing import List, Dict, Any, Optional
import pathlib

from production_output_models import BBSRow, ProductionStatistics, ProjectSteelSummary


class ProductionStatisticsCollector:
    """Gathers production metrics from workbook generation results."""

    def __init__(self) -> None:
        self._start = time.time()

    def collect(
        self,
        bbs_rows: List[BBSRow],
        steel_summary: ProjectSteelSummary,
        workbook_paths: Dict[str, pathlib.Path],
    ) -> ProductionStatistics:

        elapsed = time.time() - self._start

        # BBS row types
        header_rows = [r for r in bbs_rows if r.is_beam_header]
        eng_rows    = [r for r in bbs_rows if not r.is_beam_header]

        # Column count (BBS sheet has the most columns)
        from excel_structure_builder import BBS_COLUMNS
        ncols = len(BBS_COLUMNS)

        # Worksheet statistics per file
        ws_stats: Dict[str, Dict[str, Any]] = {}
        for name, path in workbook_paths.items():
            try:
                import openpyxl
                wb = openpyxl.load_workbook(str(path), read_only=True)
                ws_info: Dict[str, Any] = {"sheets": {}}
                for ws_name in wb.sheetnames:
                    ws = wb[ws_name]
                    ws_info["sheets"][ws_name] = {
                        "rows": ws.max_row,
                        "cols": ws.max_column,
                    }
                ws_info["sheet_count"] = len(wb.sheetnames)
                wb.close()
                ws_stats[name] = ws_info
            except Exception:
                ws_stats[name] = {"error": "could not open"}

        # Diameter totals
        diam_totals: Dict[int, float] = {}
        for ds in steel_summary.diameter_summary:
            diam_totals[ds.diameter_mm] = round(ds.total_weight_kg, 3)

        return ProductionStatistics(
            execution_time_sec=round(elapsed, 2),
            total_beams=steel_summary.total_beams,
            total_bbs_rows=len(bbs_rows),
            total_engineering_rows=len(eng_rows),
            total_rows_generated=len(bbs_rows),
            total_columns=ncols,
            steel_total_kg=round(steel_summary.total_weight_kg, 3),
            workbook_files_generated=len(workbook_paths),
            worksheet_statistics=ws_stats,
            diameter_summary=diam_totals,
        )
