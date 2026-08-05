"""
QA.2B.1 — Workbook path / hash / content helpers.
MODEL_VERSION: 9.6.1
"""
from __future__ import annotations

import hashlib
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

MODEL_VERSION = "9.6.1"
VB1_REL = Path("data/output/Production_Output/Estimation_Output.xlsx")


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def iso_mtime(path: Path) -> str:
    return datetime.fromtimestamp(path.stat().st_mtime, tz=timezone.utc).isoformat()


def resolve_latest_workbook(web_runs: Path, set_key: str) -> Optional[Path]:
    """set_key: First / Second / Third (matched in qa2_* folder names)."""
    web_runs = Path(web_runs)
    if not web_runs.exists():
        return None
    key = set_key.lower().replace(" ", "_")
    cands = [
        p
        for p in web_runs.iterdir()
        if p.is_dir()
        and p.name.lower().startswith("qa2_")
        and key in p.name.lower()
    ]
    cands = sorted(cands, key=lambda p: p.stat().st_mtime)
    for run in reversed(cands):
        xlsx = run / VB1_REL
        if xlsx.exists():
            return xlsx
    return None


def snapshot_workbook(path: Optional[Path], *, label: str = "") -> Dict[str, Any]:
    if path is None or not Path(path).exists():
        return {
            "label": label,
            "path": None,
            "exists": False,
            "sha256": None,
            "mtime_utc": None,
            "size_bytes": None,
        }
    path = Path(path)
    return {
        "label": label,
        "path": str(path),
        "run_root": str(path.parents[3]) if len(path.parts) > 3 else None,
        "exists": True,
        "sha256": sha256_file(path),
        "mtime_utc": iso_mtime(path),
        "size_bytes": path.stat().st_size,
    }


def inspect_workbook_contents(path: Path) -> Dict[str, Any]:
    """Read Estimation_Output.xlsx for beam/bar/steel counts (read-only)."""
    path = Path(path)
    out: Dict[str, Any] = {
        "path": str(path),
        "row_count": 0,
        "beam_count": 0,
        "bar_count": 0,
        "steel_kg": 0.0,
        "sheets": [],
        "ok": False,
        "error": None,
    }
    try:
        import openpyxl

        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        out["sheets"] = list(wb.sheetnames)
        beam_ids: List[str] = []
        bar_count = 0
        steel = 0.0
        rows = 0

        if "Beam Summary" in wb.sheetnames:
            ws = wb["Beam Summary"]
            header_seen = False
            for vals in ws.iter_rows(values_only=True):
                cells = list(vals)
                first = str(cells[0]).strip() if cells and cells[0] is not None else ""
                if not header_seen:
                    if first.lower() == "beam id":
                        header_seen = True
                    continue
                if not first or first.lower().startswith("total"):
                    continue
                beam_ids.append(first)
                rows += 1
                # Steel Weight (kg) column index 5 in current VB1 layout
                if len(cells) > 5 and cells[5] is not None:
                    try:
                        steel += float(cells[5])
                    except (TypeError, ValueError):
                        pass
                if len(cells) > 4 and cells[4] is not None:
                    try:
                        bar_count += int(float(cells[4]))
                    except (TypeError, ValueError):
                        pass

        if "Project Totals" in wb.sheetnames:
            ws = wb["Project Totals"]
            for vals in ws.iter_rows(values_only=True):
                cells = list(vals)
                if not cells or cells[0] is None:
                    continue
                key = str(cells[0]).strip().lower()
                val = cells[1] if len(cells) > 1 else None
                try:
                    if key == "total beams" and val is not None:
                        out["beam_count"] = int(float(val))
                    elif key == "total bars" and val is not None:
                        bar_count = int(float(val))
                    elif key == "total steel weight" and val is not None:
                        steel = float(val)
                except (TypeError, ValueError):
                    pass

        wb.close()
        if not out["beam_count"]:
            out["beam_count"] = len(beam_ids)
        out["row_count"] = rows or out["beam_count"]
        out["bar_count"] = bar_count
        out["steel_kg"] = round(steel, 3)
        out["ok"] = out["beam_count"] > 0 and out["row_count"] > 0
    except Exception as exc:  # noqa: BLE001
        out["error"] = str(exc)
        out["ok"] = path.stat().st_size > 0
    return out


def set_key_from_drawing_name(name: str) -> str:
    low = name.lower()
    if "first" in low:
        return "First"
    if "second" in low:
        return "Second"
    if "third" in low:
        return "Third"
    return name
