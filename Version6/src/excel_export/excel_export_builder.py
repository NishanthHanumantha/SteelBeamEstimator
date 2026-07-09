"""Excel export builder — Phase I.17."""

from __future__ import annotations

import shutil
from copy import copy
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.excel_export.excel_export_types import (
    CREATED_PHASE,
    DEFAULT_LOCATION_CODE,
    DETERMINATION_METHOD,
    MODEL_VERSION,
    OUTPUT_WORKBOOK_FILENAME,
    ExportState,
    default_template_path,
)


class TemplateMapper:
    """Centralized EngineeringReport → Excel cell mapping configuration."""

    WORKSHEET_NAME = "Beam - Clubhouse"
    PRESENTATION_HEADER_END_ROW = 8
    DATA_START_ROW = 9
    TEMPLATE_BEAM_HEADER_ROW = 9
    TEMPLATE_DETAIL_ROW = 10

    COLUMNS = {
        "si_no": 1,
        "location_code": 2,
        "description": 3,
        "diameter": 4,
        "spacing": 5,
        "bar_count": 6,
        "development_length": 7,
        "cut_length": 8,
        "total_length": 9,
        "weight_start": 10,
        "weight_end": 16,
        "steel_total_kg": 17,
    }

    DIAMETER_WEIGHT_COLUMNS = {
        8: 10,
        10: 11,
        12: 12,
        16: 13,
        20: 14,
        25: 15,
        32: 16,
    }

    SUMMARY_LABELS = {
        "total_bars": "Total Bars",
        "total_cut_length_mm": "Total Cut Length (mm)",
        "total_steel_weight_kg": "Total Steel Weight (kg)",
        "validation_status": "Validation Status",
        "generation_timestamp": "Generation Timestamp",
        "model_version": "Model Version",
    }

    @classmethod
    def weight_column_for_diameter(cls, diameter_mm: Any) -> int:
        try:
            diameter = int(diameter_mm)
        except (TypeError, ValueError):
            return cls.COLUMNS["weight_start"]
        return cls.DIAMETER_WEIGHT_COLUMNS.get(diameter, cls.COLUMNS["weight_start"])


class ExcelWorkbookBuilder:
    """Load presentation template, populate EngineeringReport values, save workbook."""

    FORBIDDEN_FORMULA_PREFIX = "="

    @staticmethod
    def _presentation_meters(value_mm: Any) -> Optional[float]:
        if value_mm is None:
            return None
        try:
            return round(float(value_mm) / 1000.0, 3)
        except (TypeError, ValueError):
            return None

    @staticmethod
    def _copy_cell_style(source, target) -> None:
        if source.has_style:
            target._style = copy(source._style)
            target.number_format = source.number_format
            target.font = copy(source.font)
            target.border = copy(source.border)
            target.fill = copy(source.fill)
            target.protection = copy(source.protection)
            target.alignment = copy(source.alignment)

    @classmethod
    def _copy_row_styles(
        cls,
        source_ws: Worksheet,
        source_row: int,
        target_ws: Worksheet,
        target_row: int,
        max_col: int = 17,
    ) -> int:
        copied = 0
        if source_row in source_ws.row_dimensions:
            target_ws.row_dimensions[target_row].height = source_ws.row_dimensions[source_row].height
        for col in range(1, max_col + 1):
            cls._copy_cell_style(source_ws.cell(source_row, col), target_ws.cell(target_row, col))
            copied += 1
        return copied

    @classmethod
    def _write_value(cls, ws: Worksheet, row: int, col: int, value: Any) -> int:
        if value is None:
            return 0
        ws.cell(row, col, value)
        return 1

    @classmethod
    def _clear_formula_cells(cls, ws: Worksheet, min_row: int, max_row: int, max_col: int = 17) -> int:
        cleared = 0
        for row in range(min_row, max_row + 1):
            for col in range(1, max_col + 1):
                cell = ws.cell(row, col)
                if isinstance(cell.value, str) and cell.value.startswith(cls.FORBIDDEN_FORMULA_PREFIX):
                    cell.value = None
                    cleared += 1
        return cleared

    @classmethod
    def _strip_header_formulas(cls, ws: Worksheet, value_ws: Optional[Worksheet]) -> None:
        for row in range(1, TemplateMapper.PRESENTATION_HEADER_END_ROW + 1):
            for col in range(1, 18):
                cell = ws.cell(row, col)
                if isinstance(cell.value, str) and cell.value.startswith(cls.FORBIDDEN_FORMULA_PREFIX):
                    if value_ws is not None:
                        resolved = value_ws.cell(row, col).value
                        cell.value = resolved if not (
                            isinstance(resolved, str) and resolved.startswith(cls.FORBIDDEN_FORMULA_PREFIX)
                        ) else None
                    else:
                        cell.value = None

    @classmethod
    def _write_beam_header(
        cls,
        ws: Worksheet,
        row: int,
        si_no: int,
        report: dict[str, Any],
        header: dict[str, Any],
    ) -> int:
        cells = 0
        section = header.get("beam_section") or {}
        cells += cls._write_value(ws, row, TemplateMapper.COLUMNS["si_no"], si_no)
        cells += cls._write_value(ws, row, TemplateMapper.COLUMNS["location_code"], DEFAULT_LOCATION_CODE)
        cells += cls._write_value(ws, row, TemplateMapper.COLUMNS["description"], report.get("beam_mark"))
        cells += cls._write_value(ws, row, TemplateMapper.COLUMNS["diameter"], 1)
        cells += cls._write_value(
            ws,
            row,
            TemplateMapper.COLUMNS["spacing"],
            cls._presentation_meters(header.get("clear_span_mm")),
        )
        cells += cls._write_value(
            ws,
            row,
            TemplateMapper.COLUMNS["bar_count"],
            cls._presentation_meters(section.get("width")),
        )
        cells += cls._write_value(
            ws,
            row,
            TemplateMapper.COLUMNS["development_length"],
            cls._presentation_meters(section.get("depth")),
        )
        return cells

    @classmethod
    def _write_schedule_row(cls, ws: Worksheet, row: int, schedule_row: dict[str, Any]) -> int:
        cells = 0
        cells += cls._write_value(ws, row, TemplateMapper.COLUMNS["location_code"], DEFAULT_LOCATION_CODE)
        cells += cls._write_value(ws, row, TemplateMapper.COLUMNS["description"], schedule_row.get("description"))
        cells += cls._write_value(ws, row, TemplateMapper.COLUMNS["diameter"], schedule_row.get("diameter_mm"))
        cells += cls._write_value(
            ws,
            row,
            TemplateMapper.COLUMNS["spacing"],
            cls._presentation_meters(schedule_row.get("spacing_mm")),
        )
        cells += cls._write_value(ws, row, TemplateMapper.COLUMNS["bar_count"], schedule_row.get("bar_count"))
        cells += cls._write_value(
            ws,
            row,
            TemplateMapper.COLUMNS["development_length"],
            cls._presentation_meters(schedule_row.get("development_length_mm")),
        )
        cells += cls._write_value(
            ws,
            row,
            TemplateMapper.COLUMNS["cut_length"],
            cls._presentation_meters(schedule_row.get("cut_length_mm")),
        )
        cells += cls._write_value(
            ws,
            row,
            TemplateMapper.COLUMNS["total_length"],
            cls._presentation_meters(schedule_row.get("total_length_mm")),
        )

        weight_col = TemplateMapper.weight_column_for_diameter(schedule_row.get("diameter_mm"))
        for col in range(TemplateMapper.COLUMNS["weight_start"], TemplateMapper.COLUMNS["weight_end"] + 1):
            ws.cell(row, col, None)
        cells += cls._write_value(ws, row, weight_col, schedule_row.get("steel_weight_kg"))
        cells += cls._write_value(ws, row, TemplateMapper.COLUMNS["steel_total_kg"], schedule_row.get("steel_weight_kg"))
        return cells

    @classmethod
    def _write_report_summary(cls, ws: Worksheet, row: int, report: dict[str, Any]) -> tuple[int, int]:
        sections = report.get("sections") or {}
        summary = sections.get("summary") or {}
        validation = sections.get("validation") or {}
        footer = sections.get("footer") or {}
        rows_written = 0
        cells_written = 0
        entries = [
            (TemplateMapper.SUMMARY_LABELS["total_bars"], summary.get("total_bars")),
            (TemplateMapper.SUMMARY_LABELS["total_cut_length_mm"], summary.get("total_cut_length_mm")),
            (TemplateMapper.SUMMARY_LABELS["total_steel_weight_kg"], summary.get("total_steel_weight_kg")),
            (
                TemplateMapper.SUMMARY_LABELS["validation_status"],
                validation.get("schedule_state") or report.get("report_state"),
            ),
            (
                TemplateMapper.SUMMARY_LABELS["generation_timestamp"],
                footer.get("generation_timestamp"),
            ),
            (
                TemplateMapper.SUMMARY_LABELS["model_version"],
                footer.get("model_version") or MODEL_VERSION,
            ),
        ]
        for label, value in entries:
            cls._write_value(ws, row, TemplateMapper.COLUMNS["location_code"], report.get("beam_mark"))
            cells_written += cls._write_value(ws, row, TemplateMapper.COLUMNS["description"], label)
            cells_written += cls._write_value(ws, row, TemplateMapper.COLUMNS["diameter"], value)
            row += 1
            rows_written += 1
        return rows_written, cells_written

    @classmethod
    def _create_fallback_workbook(cls) -> Workbook:
        wb = Workbook()
        ws = wb.active
        ws.title = TemplateMapper.WORKSHEET_NAME
        headers = [
            "SI no",
            "Loc",
            "Description",
            "No./Dia.",
            "L/ Spcng (m)",
            "B(m)/ No.",
            "D/Dvlp L (m)",
            "Cutting Length",
            "Total Length",
            "Steel 8",
            "Steel 10",
            "Steel 12",
            "Steel 16",
            "Steel 20",
            "Steel 25",
            "Steel 32",
            "Steel in KG",
        ]
        for col, header in enumerate(headers, start=1):
            ws.cell(TemplateMapper.PRESENTATION_HEADER_END_ROW, col, header)
        return wb

    @classmethod
    def build_workbook(
        cls,
        report_records: List[dict[str, Any]],
        output_path: Path,
        template_path: Path | None = None,
        generation_timestamp: str | None = None,
    ) -> dict[str, Any]:
        timestamp = generation_timestamp or datetime.now(timezone.utc).isoformat()
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        resolved_template = Path(template_path) if template_path else default_template_path()
        template_used = False
        fallback = False
        warnings: List[str] = []
        errors: List[str] = []

        style_ws: Optional[Worksheet] = None
        value_ws: Optional[Worksheet] = None

        if resolved_template.exists():
            try:
                shutil.copy2(resolved_template, output_path)
                wb = load_workbook(output_path)
                style_wb = load_workbook(resolved_template).active
                value_wb = load_workbook(resolved_template, data_only=True)
                value_ws = value_wb.active
                template_used = True
            except Exception as exc:
                warnings.append(f"Template open failed: {exc}")
                fallback = True
                wb = cls._create_fallback_workbook()
        else:
            warnings.append(f"Template not found: {resolved_template}")
            fallback = True
            wb = cls._create_fallback_workbook()

        ws = wb[TemplateMapper.WORKSHEET_NAME] if TemplateMapper.WORKSHEET_NAME in wb.sheetnames else wb.active
        if template_used and style_ws is not None:
            cls._strip_header_formulas(ws, value_ws)

        if ws.max_row >= TemplateMapper.DATA_START_ROW:
            ws.delete_rows(TemplateMapper.DATA_START_ROW, ws.max_row - TemplateMapper.DATA_START_ROW + 1)

        sorted_reports = sorted(report_records, key=lambda item: str(item.get("beam_id", "")))
        report_references = [str(item.get("report_id", "")) for item in sorted_reports]

        current_row = TemplateMapper.DATA_START_ROW
        rows_written = 0
        cells_written = 0
        inserted_rows = 0
        copied_styles = 0
        exported_schedule_rows = 0

        for si_no, report in enumerate(sorted_reports, start=1):
            sections = report.get("sections") or {}
            header = sections.get("header") or {}
            schedule_table = list(sections.get("schedule_table") or [])

            if style_ws is not None:
                copied_styles += cls._copy_row_styles(
                    style_ws,
                    TemplateMapper.TEMPLATE_BEAM_HEADER_ROW,
                    ws,
                    current_row,
                )
            cells_written += cls._write_beam_header(ws, current_row, si_no, report, header)
            current_row += 1
            rows_written += 1

            for schedule_row in schedule_table:
                if style_ws is not None:
                    copied_styles += cls._copy_row_styles(
                        style_ws,
                        TemplateMapper.TEMPLATE_DETAIL_ROW,
                        ws,
                        current_row,
                    )
                elif current_row > TemplateMapper.DATA_START_ROW:
                    ws.insert_rows(current_row)
                    inserted_rows += 1
                    copied_styles += cls._copy_row_styles(ws, current_row - 1, ws, current_row)
                cells_written += cls._write_schedule_row(ws, current_row, schedule_row)
                current_row += 1
                rows_written += 1
                exported_schedule_rows += 1

            summary_rows, summary_cells = cls._write_report_summary(ws, current_row + 1, report)
            current_row += summary_rows + 1
            rows_written += summary_rows
            cells_written += summary_cells

        cls._clear_formula_cells(ws, TemplateMapper.DATA_START_ROW, ws.max_row)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=20):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith(cls.FORBIDDEN_FORMULA_PREFIX):
                    cell.value = None
        wb.save(output_path)

        status = ExportState.FALLBACK.value if fallback else ExportState.SUCCESS.value
        if not sorted_reports:
            status = ExportState.SUCCESS.value
            warnings.append("No engineering reports supplied; workbook contains presentation header only.")

        return {
            "export_id": None,
            "report_references": report_references,
            "report_reference": report_references,
            "template_used": template_used,
            "template_path": str(resolved_template),
            "output_path": str(output_path),
            "output_filename": OUTPUT_WORKBOOK_FILENAME,
            "worksheet_count": len(wb.sheetnames),
            "worksheet_name": ws.title,
            "rows_written": rows_written,
            "cells_written": cells_written,
            "inserted_rows": inserted_rows,
            "copied_styles": copied_styles,
            "exported_schedule_rows": exported_schedule_rows,
            "generation_time": timestamp,
            "generation_timestamp": timestamp,
            "status": status,
            "validation_status": "PENDING",
            "determination_method": DETERMINATION_METHOD,
            "generation_phase": CREATED_PHASE,
            "model_version": MODEL_VERSION,
            "fallback_used": fallback,
            "warnings": warnings,
            "errors": errors,
        }


class ExcelExportBuilder:
    """Facade for workbook construction from EngineeringReport records."""

    @staticmethod
    def build_export(
        report_records: List[dict[str, Any]],
        output_path: Path,
        template_path: Path | None = None,
        generation_timestamp: str | None = None,
    ) -> dict[str, Any]:
        return ExcelWorkbookBuilder.build_workbook(
            report_records,
            output_path,
            template_path=template_path,
            generation_timestamp=generation_timestamp,
        )
