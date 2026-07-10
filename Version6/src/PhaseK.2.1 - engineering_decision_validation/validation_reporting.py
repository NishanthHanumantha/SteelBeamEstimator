"""Reporting helpers for Phase K.2.1."""

from __future__ import annotations

from typing import Any, List

from decision_loader import MODEL_VERSION, PHASE
from decision_validation_types import RULE_CATALOG


class ValidationReporting:
    """Build report, matrix, errors, warnings, and Excel payloads."""

    @staticmethod
    def build_report(result: dict[str, Any]) -> dict[str, Any]:
        return {
            "phase": PHASE,
            "model_version": MODEL_VERSION,
            "run_timestamp": result.get("run_timestamp"),
            "statistics": result.get("statistics"),
            "health": result.get("health"),
            "summary": result.get("summary"),
            "validation_status": (result.get("validation") or {}).get("status"),
            "idempotent": result.get("idempotent"),
            "config": result.get("config"),
            "execution_gate": result.get("execution_gate"),
        }

    @staticmethod
    def build_errors(validations: List[dict[str, Any]]) -> dict[str, Any]:
        rows = []
        for item in validations:
            for error in item.get("validation_errors") or []:
                rows.append(
                    {
                        "decision_id": item.get("decision_id"),
                        "decision_key": item.get("decision_key"),
                        "validation_id": item.get("validation_id"),
                        "group": error.get("group"),
                        "code": error.get("code"),
                        "message": error.get("message"),
                    }
                )
        return {"error_count": len(rows), "errors": rows}

    @staticmethod
    def build_warnings(validations: List[dict[str, Any]]) -> dict[str, Any]:
        rows = []
        for item in validations:
            for warning in item.get("validation_warnings") or []:
                rows.append(
                    {
                        "decision_id": item.get("decision_id"),
                        "decision_key": item.get("decision_key"),
                        "validation_id": item.get("validation_id"),
                        "group": warning.get("group"),
                        "code": warning.get("code"),
                        "message": warning.get("message"),
                    }
                )
        return {"warning_count": len(rows), "warnings": rows}

    @staticmethod
    def build_traceability(validations: List[dict[str, Any]]) -> dict[str, Any]:
        chains = [
            {
                "decision_id": item.get("decision_id"),
                "decision_key": item.get("decision_key"),
                "validation_id": item.get("validation_id"),
                "execution_allowed": item.get("execution_allowed"),
                "traceability": item.get("traceability") or {},
            }
            for item in validations
        ]
        return {"chain_count": len(chains), "chains": chains}

    @staticmethod
    def build_matrix(validations: List[dict[str, Any]]) -> dict[str, Any]:
        rows = []
        for item in validations:
            breakdown = item.get("score_breakdown") or {}
            rows.append(
                {
                    "decision_id": item.get("decision_id"),
                    "decision_key": item.get("decision_key"),
                    "validation_status": item.get("validation_status"),
                    "validation_score": item.get("validation_score"),
                    "identity": breakdown.get("IDENTITY", 0),
                    "traceability": breakdown.get("TRACEABILITY", 0),
                    "execution": breakdown.get("EXECUTION", 0),
                    "engineering": breakdown.get("ENGINEERING", 0),
                    "production_safety": breakdown.get("PRODUCTION_SAFETY", 0),
                    "execution_allowed": item.get("execution_allowed"),
                    "error_count": len(item.get("validation_errors") or []),
                    "warning_count": len(item.get("validation_warnings") or []),
                }
            )
        return {"row_count": len(rows), "rows": rows}

    @staticmethod
    def build_rules_catalog() -> dict[str, Any]:
        return {
            "phase": PHASE,
            "model_version": MODEL_VERSION,
            "rule_count": len(RULE_CATALOG),
            "rules": list(RULE_CATALOG),
        }

    @staticmethod
    def write_excel(path, validations: List[dict[str, Any]], statistics: dict[str, Any]) -> bool:
        try:
            from openpyxl import Workbook
        except ImportError:
            return False

        workbook = Workbook()
        summary = workbook.active
        summary.title = "Summary"
        summary.append(["Metric", "Value"])
        for key, value in statistics.items():
            summary.append([key, value])

        sheet = workbook.create_sheet("Validations")
        sheet.append(
            [
                "validation_id",
                "decision_id",
                "decision_key",
                "validation_status",
                "validation_score",
                "execution_allowed",
                "error_count",
                "warning_count",
            ]
        )
        for item in validations:
            sheet.append(
                [
                    item.get("validation_id"),
                    item.get("decision_id"),
                    item.get("decision_key"),
                    item.get("validation_status"),
                    item.get("validation_score"),
                    item.get("execution_allowed"),
                    len(item.get("validation_errors") or []),
                    len(item.get("validation_warnings") or []),
                ]
            )

        errors = workbook.create_sheet("Errors")
        errors.append(["decision_id", "group", "code", "message"])
        for item in validations:
            for error in item.get("validation_errors") or []:
                errors.append(
                    [
                        item.get("decision_id"),
                        error.get("group"),
                        error.get("code"),
                        error.get("message"),
                    ]
                )

        warnings = workbook.create_sheet("Warnings")
        warnings.append(["decision_id", "group", "code", "message"])
        for item in validations:
            for warning in item.get("validation_warnings") or []:
                warnings.append(
                    [
                        item.get("decision_id"),
                        warning.get("group"),
                        warning.get("code"),
                        warning.get("message"),
                    ]
                )

        path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(path)
        return True
