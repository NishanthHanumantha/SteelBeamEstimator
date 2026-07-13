"""Export Phase L.1 accuracy sprint artifacts."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

EXPORT_FILES: Tuple[str, ...] = (
    "engineering_gap_report.json",
    "engineering_gap_statistics.json",
    "engineering_gap_summary.json",
    "engineering_gap_matrix.json",
    "root_cause_analysis.json",
    "reinforcement_role_gap_analysis.json",
    "engineering_rule_gap_analysis.json",
    "coverage_analysis.json",
    "priority_ranking.json",
    "improvement_tracker.json",
    "accuracy_dashboard.json",
)


class AccuracyExport:
    """Write all L.1 JSON and Excel outputs."""

    @staticmethod
    def export_all(
        output_dir: Path,
        result: Dict[str, Any],
        config: Dict[str, Any],
    ) -> Dict[str, str]:
        output_dir.mkdir(parents=True, exist_ok=True)
        written: Dict[str, str] = {}

        mapping: Dict[str, Any] = {
            "engineering_gap_report.json": result.get("report"),
            "engineering_gap_statistics.json": result.get("statistics"),
            "engineering_gap_summary.json": result.get("summary"),
            "engineering_gap_matrix.json": result.get("gap_matrix"),
            "root_cause_analysis.json": {"gaps": result.get("classified_gaps")},
            "reinforcement_role_gap_analysis.json": {
                "phase": "Phase L.1", "rows": result.get("reinforcement_role_gaps")
            },
            "engineering_rule_gap_analysis.json": result.get("rule_gap_analysis"),
            "coverage_analysis.json": result.get("coverage"),
            "priority_ranking.json": {"ranked_gaps": result.get("priority_backlog")},
            "improvement_tracker.json": result.get("improvement_tracker"),
            "accuracy_dashboard.json": result.get("dashboard"),
        }

        for filename in EXPORT_FILES:
            path = output_dir / filename
            payload = mapping.get(filename)
            AccuracyExport._write_json(path, payload)
            written[filename] = str(path)

        if config.get("generate_excel_report", True):
            xlsx_path = output_dir / "accuracy_sprint_report.xlsx"
            ok = AccuracyExport._write_excel(xlsx_path, result)
            if ok:
                written["accuracy_sprint_report.xlsx"] = str(xlsx_path)

        result["export_paths"] = written
        return written

    @staticmethod
    def validate_exports(output_dir: Path) -> Dict[str, Any]:
        checks = []
        for filename in EXPORT_FILES:
            path = output_dir / filename
            exists = path.exists() and path.stat().st_size > 2
            parsed = False
            if exists:
                try:
                    json.loads(path.read_text(encoding="utf-8"))
                    parsed = True
                except Exception:
                    pass
            checks.append({"name": f"Export Exists {filename}", "status": "PASS" if exists else "FAIL"})
            checks.append({"name": f"Export Valid JSON {filename}", "status": "PASS" if parsed else "FAIL"})
        xlsx = output_dir / "accuracy_sprint_report.xlsx"
        checks.append({
            "name": "Export Exists accuracy_sprint_report.xlsx",
            "status": "PASS" if xlsx.exists() and xlsx.stat().st_size > 0 else "FAIL",
        })
        failed = [c for c in checks if c["status"] == "FAIL"]
        return {
            "status": "PASS" if not failed else "FAIL",
            "checks": checks,
            "summary": {"total_checks": len(checks), "passed": len(checks) - len(failed), "failed": len(failed)},
        }

    @staticmethod
    def print_summary(result: Dict[str, Any]) -> None:
        stats = result.get("statistics") or {}
        health = result.get("health") or {}
        export_paths = result.get("export_paths") or {}
        print("\n" + "=" * 80)
        print("Phase L.1 — Accuracy Sprint 1: Estimator Gap Closure")
        print("=" * 80)
        print(f"Model Version : {result.get('model_version')}")
        print(f"Benchmark     : Sobha Galera Clubhouse")
        print("")
        print(f"Beam Coverage                   : {stats.get('beam_coverage_percent', 0.0)}%")
        print(f"Geometry Coverage               : {stats.get('geometry_coverage_percent', 0.0)}%")
        print(f"Steel Weight Coverage           : {stats.get('steel_coverage_percent', 0.0)}%")
        print(f"Diameter Coverage               : {stats.get('diameter_coverage_percent', 0.0)}%")
        print(f"Reinforcement Role Coverage     : {stats.get('reinforcement_role_coverage_percent', 0.0)}%")
        print(f"Engineering Rule Coverage       : {stats.get('engineering_rule_coverage_percent', 0.0)}%")
        print(f"Decision Coverage               : {stats.get('decision_coverage_percent', 0.0)}%")
        print(f"Row Coverage                    : {stats.get('row_coverage_percent', 0.0)}%")
        print(f"Estimator Equivalence           : {stats.get('estimator_equivalence_percent', 0.0)}%")
        print(f"Overall Estimator Accuracy      : {stats.get('overall_estimator_accuracy_percent', 0.0)}%")
        print("")
        print(f"Total Gaps                      : {stats.get('total_gaps', 0)}")
        print(f"Critical Gaps                   : {stats.get('critical_gaps', 0)}")
        print(f"High Gaps                       : {stats.get('high_gaps', 0)}")
        print(f"Overall Accuracy Health         : {health.get('overall_accuracy_health')}")
        print("")
        val = result.get("validation") or {}
        exp = result.get("export_validation") or {}
        print(f"Validation : {val.get('summary', {}).get('passed', 0)}/{val.get('summary', {}).get('total_checks', 0)} PASS")
        print(f"Exports    : {exp.get('summary', {}).get('passed', 0)}/{exp.get('summary', {}).get('total_checks', 0)} PASS")
        print("")
        print("Export Locations")
        print("-" * 80)
        for fname, fpath in export_paths.items():
            print(f"{fname}: {fpath}")
        print("=" * 80 + "\n")

    @staticmethod
    def _write_json(path: Path, payload: Any) -> None:
        path.write_text(json.dumps(payload, indent=2, sort_keys=False, default=str), encoding="utf-8")

    @staticmethod
    def _write_excel(path: Path, result: Dict[str, Any]) -> bool:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return False

        wb = Workbook()

        # ---- Summary sheet ----
        ws = wb.active
        ws.title = "Summary"
        stats = result.get("statistics") or {}
        ws.append(["Phase L.1 Accuracy Sprint 1 — Engineering Accuracy Dashboard"])
        ws.append(["Model Version", result.get("model_version")])
        ws.append(["Benchmark", "Sobha Galera Clubhouse"])
        ws.append([])
        ws.append(["KPI", "Value"])
        for k, v in stats.items():
            if not isinstance(v, dict):
                ws.append([k, v])

        # ---- Gap Matrix ----
        gap_ws = wb.create_sheet("Gap Matrix")
        gap_ws.append([
            "Gap ID", "Category", "Priority", "Rank", "Title",
            "Affected Beams", "Affected Roles", "Steel Impact (kg)", "Future Phase",
        ])
        for g in (result.get("priority_backlog") or []):
            gap_ws.append([
                g.get("gap_id"),
                g.get("gap_category"),
                g.get("priority"),
                g.get("priority_rank"),
                g.get("title"),
                ", ".join(g.get("affected_beams") or []) or "ALL",
                ", ".join(g.get("affected_roles") or []),
                g.get("estimated_steel_impact_kg", 0.0),
                g.get("future_phase"),
            ])

        # ---- Role Coverage ----
        role_ws = wb.create_sheet("Role Coverage")
        role_ws.append([
            "Role", "Estimator Bars", "Estimator Weight (kg)", "Model Decisions",
            "Difference", "Coverage %", "Priority", "Root Cause",
        ])
        for row in (result.get("reinforcement_role_gaps") or []):
            role_ws.append([
                row.get("role"),
                row.get("estimator_bar_count"),
                row.get("estimator_weight_kg"),
                row.get("model_decision_count"),
                row.get("difference"),
                row.get("coverage_percent"),
                row.get("priority"),
                row.get("root_cause"),
            ])

        # ---- Per Beam ----
        beam_ws = wb.create_sheet("Per Beam")
        beam_ws.append([
            "Beam", "Est Rows", "Model Rows", "Missing Rows",
            "Row Coverage %", "Est Weight kg", "Model Weight kg", "Weight Gap kg",
        ])
        for b in ((result.get("comparison") or {}).get("per_beam") or []):
            beam_ws.append([
                b.get("beam_mark"),
                b.get("estimator_rows"),
                b.get("model_rows"),
                b.get("missing_rows"),
                b.get("row_coverage_percent"),
                b.get("estimator_weight_kg"),
                b.get("model_weight_kg"),
                b.get("weight_difference_kg"),
            ])

        # ---- Rule Gaps ----
        rule_ws = wb.create_sheet("Rule Gaps")
        rule_ws.append(["Rule ID", "Rule Name", "Category", "Status", "Priority", "Steel Impact %", "Future Phase"])
        for r in ((result.get("rule_gap_analysis") or {}).get("rules") or []):
            rule_ws.append([
                r.get("rule_id"),
                r.get("rule_name"),
                r.get("rule_category"),
                r.get("status"),
                r.get("priority"),
                r.get("estimated_steel_impact_percent"),
                r.get("future_phase"),
            ])

        # ---- Improvement Tracker ----
        imp_ws = wb.create_sheet("Improvement Tracker")
        imp_ws.append(["Improvement ID", "Gap Category", "Priority", "Title", "Status", "Future Phase", "Created Version"])
        for imp in ((result.get("improvement_tracker") or {}).get("improvements") or []):
            imp_ws.append([
                imp.get("improvement_id"),
                imp.get("gap_category"),
                imp.get("priority"),
                imp.get("title"),
                imp.get("status"),
                imp.get("future_phase"),
                imp.get("created_version"),
            ])

        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)
        return True
