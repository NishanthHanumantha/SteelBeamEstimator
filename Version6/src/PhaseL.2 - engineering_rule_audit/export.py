"""Export all Phase L.2 Engineering Rule Audit artifacts."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, Tuple

EXPORT_FILES: Tuple[str, ...] = (
    "engineering_rule_inventory.json",
    "pipeline_trace.json",
    "execution_breaks.json",
    "dependency_graph.json",
    "role_audit.json",
    "beam_audit.json",
    "implementation_matrix.json",
    "dead_rules.json",
    "unused_rules.json",
    "coverage_statistics.json",
    "engineering_rule_audit_summary.json",
)


class AuditExport:
    """Write all L.2 JSON and Excel outputs."""

    @staticmethod
    def export_all(
        output_dir: Path,
        result: Dict[str, Any],
        config: Dict[str, Any],
    ) -> Dict[str, str]:
        output_dir.mkdir(parents=True, exist_ok=True)
        written: Dict[str, str] = {}

        rule_inv = result.get("rule_inventory") or {}
        all_rules = rule_inv.get("inventory", {}).get("rules") or []
        dead_rules = [r for r in all_rules if r.get("dead_code_candidate")]
        unused_rules = [r for r in all_rules if "NOT_IMPLEMENTED" in str(result.get("coverage_statistics", {}))]

        mapping: Dict[str, Any] = {
            "engineering_rule_inventory.json": rule_inv,
            "pipeline_trace.json": result.get("pipeline_trace"),
            "execution_breaks.json": {"breaks": result.get("execution_breaks")},
            "dependency_graph.json": result.get("dependency_graph"),
            "role_audit.json": {"rows": result.get("role_audit")},
            "beam_audit.json": result.get("beam_audit"),
            "implementation_matrix.json": result.get("implementation_matrix"),
            "dead_rules.json": {"dead_rule_candidates": dead_rules},
            "unused_rules.json": {
                "unused_rules": [
                    r for r in all_rules
                    if not r.get("roles_referenced")
                ],
            },
            "coverage_statistics.json": result.get("coverage_statistics"),
            "engineering_rule_audit_summary.json": result.get("summary"),
        }

        for filename in EXPORT_FILES:
            path = output_dir / filename
            payload = mapping.get(filename)
            AuditExport._write_json(path, payload)
            written[filename] = str(path)

        if config.get("generate_excel_report", True):
            xlsx_path = output_dir / "engineering_rule_audit_report.xlsx"
            ok = AuditExport._write_excel(xlsx_path, result)
            if ok:
                written["engineering_rule_audit_report.xlsx"] = str(xlsx_path)

        result["export_paths"] = written
        return written

    @staticmethod
    def validate_exports(output_dir: Path) -> Dict[str, Any]:
        checks = []
        for filename in EXPORT_FILES:
            path = output_dir / filename
            exists = path.exists() and path.stat().st_size > 2
            parsed = False
            if exists:
                try:
                    json.loads(path.read_text(encoding="utf-8"))
                    parsed = True
                except Exception:
                    pass
            checks.append({"name": f"Exists {filename}", "status": "PASS" if exists else "FAIL"})
            checks.append({"name": f"Valid JSON {filename}", "status": "PASS" if parsed else "FAIL"})
        xlsx = output_dir / "engineering_rule_audit_report.xlsx"
        checks.append({
            "name": "Exists engineering_rule_audit_report.xlsx",
            "status": "PASS" if xlsx.exists() and xlsx.stat().st_size > 0 else "FAIL",
        })
        failed = [c for c in checks if c["status"] == "FAIL"]
        return {
            "status": "PASS" if not failed else "FAIL",
            "checks": checks,
            "summary": {"total_checks": len(checks), "passed": len(checks) - len(failed), "failed": len(failed)},
        }

    @staticmethod
    def print_summary(result: Dict[str, Any]) -> None:
        cov = result.get("coverage_statistics") or {}
        val = result.get("validation") or {}
        exp = result.get("export_validation") or {}
        export_paths = result.get("export_paths") or {}
        print("\n" + "=" * 80)
        print("Phase L.2 — Engineering Rule Audit Engine")
        print("=" * 80)
        print(f"Model Version         : {result.get('model_version')}")
        print(f"Data Source           : {result.get('data_source', 'V5_REFERENCE')}")
        print("")
        print(f"Rules Discovered      : {cov.get('total_engineering_rules_discovered', 0)}")
        print(f"Roles Audited         : {cov.get('total_roles_audited', 0)}")
        print(f"Implemented %         : {cov.get('implemented_percent', 0.0)}%")
        print(f"Executed %            : {cov.get('executed_percent', 0.0)}%")
        print(f"Reachable %           : {cov.get('reachable_percent', 0.0)}%")
        print(f"Exported %            : {cov.get('exported_percent', 0.0)}%")
        print(f"Estimator Coverage %  : {cov.get('estimator_coverage_percent', 0.0)}%")
        print(f"Dead Code Candidates  : {cov.get('dead_code_rule_candidates', 0)}")
        print(f"Pipeline Completion % : {cov.get('pipeline_completion_percent', 0.0)}%")
        print("")
        print(f"Validation : {val.get('summary', {}).get('passed', 0)}/{val.get('summary', {}).get('total_checks', 0)} PASS")
        print(f"Exports    : {exp.get('summary', {}).get('passed', 0)}/{exp.get('summary', {}).get('total_checks', 0)} PASS")
        print("")
        print("Exports:")
        for fname in list(export_paths.keys())[:5]:
            print(f"  {fname}")
        print("=" * 80 + "\n")

    @staticmethod
    def _write_json(path: Path, payload: Any) -> None:
        path.write_text(json.dumps(payload, indent=2, sort_keys=False, default=str), encoding="utf-8")

    @staticmethod
    def _write_excel(path: Path, result: Dict[str, Any]) -> bool:
        try:
            from openpyxl import Workbook
        except ImportError:
            return False

        wb = Workbook()

        # ---- Implementation Matrix ----
        ws = wb.active
        ws.title = "Implementation Matrix"
        headers = [
            "Role", "Detected", "Parsed", "Geometry", "Ownership", "Context",
            "Rule Exists", "Rule Executed", "Quantity", "Exported", "Estimator Match",
            "Final Status", "Break Stage", "Break Category",
        ]
        ws.append(headers)
        for row in ((result.get("implementation_matrix") or {}).get("rows") or []):
            ws.append([row.get(h.lower().replace(" ", "_")) or row.get(h.lower().replace(" ", "")) or "" for h in headers])

        # ---- Coverage Statistics ----
        cs_ws = wb.create_sheet("Coverage Statistics")
        cs_ws.append(["Metric", "Value"])
        for k, v in (result.get("coverage_statistics") or {}).items():
            if not isinstance(v, dict):
                cs_ws.append([k, v])

        # ---- Execution Breaks ----
        brk_ws = wb.create_sheet("Execution Breaks")
        brk_ws.append(["Role", "Break Category", "Break Stage", "Object Count", "Success Count", "Estimator Gap"])
        for b in (result.get("execution_breaks") or []):
            brk_ws.append([
                b.get("role"), b.get("break_category"), b.get("break_stage"),
                b.get("pipeline_object_count"), b.get("pipeline_success_count"),
                b.get("estimator_gap"),
            ])

        # ---- Dependency Graph ----
        dep_ws = wb.create_sheet("Dependency Graph")
        dep_ws.append(["Role", "Total Deps", "Satisfied", "Missing", "First Missing", "First Missing Phase"])
        for e in ((result.get("dependency_graph") or {}).get("entries") or []):
            dep_ws.append([
                e.get("role"), e.get("total_dependencies"), e.get("satisfied_dependencies"),
                e.get("missing_dependencies"), e.get("first_missing_dependency"),
                e.get("first_missing_phase"),
            ])

        # ---- Rule Inventory ----
        inv_ws = wb.create_sheet("Rule Inventory")
        inv_ws.append(["Rule ID", "Class/Func", "Module", "Phase", "Roles Referenced", "Dead Code"])
        for rule in ((result.get("rule_inventory") or {}).get("inventory", {}).get("rules") or [])[:200]:
            inv_ws.append([
                rule.get("rule_id"),
                rule.get("class_or_function"),
                rule.get("module"),
                rule.get("phase_introduced"),
                ", ".join(rule.get("roles_referenced") or []),
                rule.get("dead_code_candidate", False),
            ])

        # ---- Estimator Trace ----
        est_ws = wb.create_sheet("Estimator Trace")
        est_ws.append(["Role", "Est Bars", "Est Weight", "V6 Bars", "V6 Schedule Rows", "Match", "Absence Reason"])
        for t in ((result.get("estimator_trace") or {}).get("traces") or []):
            est_ws.append([
                t.get("role"), t.get("estimator_bar_count"), t.get("estimator_weight_kg"),
                t.get("v6_bar_count"), t.get("v6_schedule_rows"), t.get("match_quality"),
                t.get("absence_reason"),
            ])

        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)
        return True
