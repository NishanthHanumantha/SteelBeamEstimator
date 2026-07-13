"""Export all Phase L.2.1 Engineering Feature artifacts."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, List, Tuple

from engineering_feature_model import EngineeringFeatureModel

EXPORT_FILES: Tuple[str, ...] = (
    "engineering_feature_database.json",
    "geometry_features.json",
    "position_features.json",
    "continuity_features.json",
    "support_features.json",
    "extent_features.json",
    "orientation_features.json",
    "annotation_features.json",
    "topology_features.json",
    "feature_statistics.json",
    "feature_dashboard.json",
)


class FeatureExport:
    @staticmethod
    def export_all(output_dir: Path, result: Dict[str, Any], config: Dict[str, Any]) -> Dict[str, str]:
        output_dir.mkdir(parents=True, exist_ok=True)
        written: Dict[str, str] = {}
        features: List[EngineeringFeatureModel] = result.get("features") or []

        mapping: Dict[str, Any] = {
            "engineering_feature_database.json": {
                "phase": result.get("phase"),
                "model_version": result.get("model_version"),
                "total_features": len(features),
                "features": [f.to_dict() for f in features],
            },
            "geometry_features.json": result.get("geometry_features"),
            "position_features.json": result.get("position_features"),
            "continuity_features.json": result.get("continuity_features"),
            "support_features.json": result.get("support_features"),
            "extent_features.json": result.get("extent_features"),
            "orientation_features.json": result.get("orientation_features"),
            "annotation_features.json": result.get("annotation_features"),
            "topology_features.json": result.get("topology_features"),
            "feature_statistics.json": result.get("statistics"),
            "feature_dashboard.json": {
                "summary": result.get("summary"),
                "validation": result.get("validation"),
                "load_status": result.get("load_status"),
            },
        }

        for filename in EXPORT_FILES:
            path = output_dir / filename
            payload = mapping.get(filename)
            path.write_text(
                json.dumps(payload, indent=2, sort_keys=False, default=str),
                encoding="utf-8",
            )
            written[filename] = str(path)

        if config.get("generate_excel_report", True):
            xlsx_path = output_dir / "engineering_feature_report.xlsx"
            if FeatureExport._write_excel(xlsx_path, features):
                written["engineering_feature_report.xlsx"] = str(xlsx_path)

        result["export_paths"] = written
        return written

    @staticmethod
    def validate_exports(output_dir: Path) -> Dict[str, Any]:
        checks = []
        for filename in EXPORT_FILES:
            path = output_dir / filename
            exists = path.exists() and path.stat().st_size > 2
            valid = False
            if exists:
                try:
                    json.loads(path.read_text(encoding="utf-8"))
                    valid = True
                except Exception:
                    pass
            checks.append({"name": f"Exists {filename}", "status": "PASS" if exists else "FAIL"})
            checks.append({"name": f"Valid JSON {filename}", "status": "PASS" if valid else "FAIL"})
        xlsx = output_dir / "engineering_feature_report.xlsx"
        checks.append({
            "name": "Exists engineering_feature_report.xlsx",
            "status": "PASS" if xlsx.exists() and xlsx.stat().st_size > 0 else "FAIL",
        })
        failed = [c for c in checks if c["status"] == "FAIL"]
        return {
            "status": "PASS" if not failed else "FAIL",
            "checks": checks,
            "summary": {"total": len(checks), "passed": len(checks) - len(failed), "failed": len(failed)},
        }

    @staticmethod
    def _write_excel(path: Path, features: List[EngineeringFeatureModel]) -> bool:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font
        except ImportError:
            return False

        wb = Workbook()

        # ── Feature Database ──────────────────────────────────────────────
        ws = wb.active
        ws.title = "Feature Database"
        ws.append([
            "Feature ID", "Bar ID", "Beam ID",
            "Position Zone", "Orientation", "Extent Type",
            "Coverage Ratio", "Is Continuous", "Continuity Type",
            "Support Region", "Diameter mm", "Quantity",
            "Vertical Rank", "Depth Ratio", "Completeness Score",
        ])
        for f in features:
            ws.append([
                f.feature_id, f.bar_id, f.beam_id,
                f.position.position_zone, f.orientation.orientation, f.extent.extent_type,
                f.extent.coverage_ratio, f.continuity.is_continuous, f.continuity.continuity_type,
                f.support.support_region_type, f.annotation.diameter_mm, f.annotation.quantity,
                f.position.vertical_rank, f.position.beam_depth_ratio,
                f.feature_completeness_score,
            ])

        # ── Geometry Features ─────────────────────────────────────────────
        g_ws = wb.create_sheet("Geometry Features")
        g_ws.append([
            "Feature ID", "Beam ID", "Length mm", "Relative Length",
            "Angle deg", "Is Closed", "Touches Support", "Crosses Beam Axis",
        ])
        for f in features:
            g = f.geometry
            g_ws.append([
                f.feature_id, f.beam_id, g.length_mm, g.relative_length,
                g.orientation_angle_deg, g.is_closed, g.touches_support, g.crosses_beam_axis,
            ])

        # ── Position Features ─────────────────────────────────────────────
        p_ws = wb.create_sheet("Position Features")
        p_ws.append([
            "Feature ID", "Beam ID", "Zone", "V-Rank",
            "Dist Top mm", "Dist Bottom mm", "Dist Left mm", "Dist Right mm", "Depth Ratio",
        ])
        for f in features:
            p = f.position
            p_ws.append([
                f.feature_id, f.beam_id, p.position_zone, p.vertical_rank,
                p.distance_from_top_face_mm, p.distance_from_bottom_face_mm,
                p.distance_from_left_support_mm, p.distance_from_right_support_mm,
                p.beam_depth_ratio,
            ])

        # ── Continuity Features ───────────────────────────────────────────
        c_ws = wb.create_sheet("Continuity Features")
        c_ws.append([
            "Feature ID", "Beam ID", "Is Continuous", "Cont Type",
            "Is Multi-Span", "N Beams Crossed", "Beam Sequence",
        ])
        for f in features:
            c = f.continuity
            c_ws.append([
                f.feature_id, f.beam_id, c.is_continuous, c.continuity_type,
                c.is_multi_span, c.number_of_beams_crossed, ",".join(c.beam_sequence),
            ])

        # ── Annotation Features ───────────────────────────────────────────
        a_ws = wb.create_sheet("Annotation Features")
        a_ws.append([
            "Feature ID", "Beam ID", "Callout", "Dia mm", "Qty", "Spacing mm", "Priority",
        ])
        for f in features:
            a = f.annotation
            a_ws.append([
                f.feature_id, f.beam_id, a.callout, a.diameter_mm, a.quantity,
                a.spacing_mm, a.annotation_priority,
            ])

        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)
        return True

    @staticmethod
    def print_summary(result: Dict[str, Any]) -> None:
        stats = result.get("statistics") or {}
        val = result.get("validation") or {}
        exp = result.get("export_validation") or {}
        zones = stats.get("zone_distribution") or {}
        oris = stats.get("orientation_distribution") or {}
        print("\n" + "=" * 80)
        print("Phase L.2.1 - Engineering Feature Extraction Engine")
        print("=" * 80)
        print(f"Model Version       : {result.get('model_version')}")
        print(f"Total Features      : {stats.get('total_features', 0)}")
        print(f"Total Beams         : {stats.get('total_beams', 0)}")
        print(f"Completeness Rate   : {stats.get('completeness_rate_percent', 0)}%")
        print(f"Multi-Span Bars     : {stats.get('multi_span_bars', 0)}")
        print("")
        print("Position Zone Distribution:")
        for z, cnt in sorted(zones.items()):
            if cnt:
                print(f"  {z:<20}: {cnt}")
        print("")
        print("Orientation Distribution:")
        for o, cnt in sorted(oris.items()):
            if cnt:
                print(f"  {o:<20}: {cnt}")
        print("")
        print(f"Validation  : {val.get('summary', {}).get('passed', 0)}/{val.get('summary', {}).get('total_checks', 0)} PASS")
        print(f"Exports     : {exp.get('summary', {}).get('passed', 0)}/{exp.get('summary', {}).get('total', 0)} PASS")
        print("=" * 80 + "\n")
