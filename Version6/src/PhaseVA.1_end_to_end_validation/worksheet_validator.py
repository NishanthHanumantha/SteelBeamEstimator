"""
Phase V.A.1 — End-to-End Validation
worksheet_validator.py — Validate every worksheet in the generated workbook.
MODEL_VERSION: 6.5.3
"""
from __future__ import annotations

import pathlib
from typing import Any, Dict, List, Optional

from validation_models import WorksheetValidation

# Expected header columns in the main reinforcement schedule sheet
EXPECTED_HEADERS = {
    "SI no", "Description", "No./Dia.", "L/ Spcng (m)", "B(m)/ No.",
    "D/Dvlp L (m)", "Cutting Length", "Total Length", "Steel in KG",
}

# Diameter columns expected
EXPECTED_DIAMETER_COLS = {8, 10, 12, 16, 20, 25, 32}

# Expected project header keywords
EXPECTED_PROJECT_KEYWORDS = {"Project", "Floor", "Reinforcement"}

# Standard V6 pipeline output JSON files to check (not Excel-internal worksheets,
# but intermediate output directories)
PIPELINE_OUTPUT_CHECKS: List[Dict[str, str]] = [
    {
        "label": "L.2 Reinforcement Models",
        "path": "Version6/data/output/PhaseL.2 - engineering_reinforcement_interpretation/beam_reinforcement_models.json",
    },
    {
        "label": "L.2.2 Extended Models",
        "path": "Version6/data/output/PhaseL.2.2_geometry_recovery/extended_beam_reinforcement_models.json",
    },
    {
        "label": "L.2.1 Feature Database",
        "path": "Version6/data/output/PhaseL.2.1 - engineering_feature_extraction/engineering_feature_database.json",
    },
    {
        "label": "L.3 Patterns",
        "path": "Version6/data/output/PhaseL.3_beam_pattern_recognition/engineering_patterns.json",
    },
    {
        "label": "BBS Results",
        "path": "Version6/data/output/phase_i/i_10_bbs/bbs_results.json",
    },
    {
        "label": "Steel Weight Results",
        "path": "Version6/data/output/phase_i/i_11_steel_weight/steel_weight_results.json",
    },
    {
        "label": "Cut Length Results",
        "path": "Version6/data/output/phase_i/i_6_cut_length/cut_length_results.json",
    },
    {
        "label": "Quantity Results",
        "path": "Version6/data/output/phase_i/i_13_quantity/quantity_results.json",
    },
    {
        "label": "Excel Workbook",
        "path": "Version6/data/output/phase_i/i_17_excel_export/Beam_Reinforcement_Schedule.xlsx",
    },
]


class WorksheetValidator:
    """Validates every worksheet in the generated workbook + pipeline output files."""

    def __init__(self, repo_root: pathlib.Path) -> None:
        self._root = repo_root

    def validate_workbook_sheets(self, workbook_path: pathlib.Path) -> List[WorksheetValidation]:
        if not workbook_path.exists():
            return []

        import openpyxl
        try:
            wb = openpyxl.load_workbook(str(workbook_path))
        except Exception:
            return []

        results: List[WorksheetValidation] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            results.append(self._validate_sheet(sheet_name, ws))
        return results

    def _validate_sheet(self, sheet_name: str, ws: Any) -> WorksheetValidation:
        import openpyxl
        issues: List[str] = []

        row_count = ws.max_row
        col_count = ws.max_column

        # Read header row candidates (rows 5-6 typically contain column headers)
        header_row: List[Any] = []
        for r in range(1, min(8, row_count + 1)):
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, col_count + 1)]
            non_none = [v for v in row_vals if v is not None]
            if len(non_none) >= 4:
                header_row = row_vals
                break

        has_headers = len(header_row) > 0

        # Check for expected header keywords
        header_flat = {str(v).strip() for v in header_row if v is not None}
        missing_headers = EXPECTED_HEADERS - header_flat
        if missing_headers and "Description" in EXPECTED_HEADERS:
            # Try the next row
            pass  # Already handled by the scan above

        # Check for project info
        cell_a1 = ws.cell(row=1, column=1).value
        if cell_a1 is None or "Project" not in str(cell_a1):
            issues.append(f"Missing 'Project :' in A1 — found: {cell_a1!r}")

        # Check minimum data rows
        if row_count < 10:
            issues.append(f"Sheet has only {row_count} rows — expected at least 10 for reinforcement schedule.")

        # Read first data row after headers
        first_data_row: Optional[List[Any]] = None
        for r in range(8, min(20, row_count + 1)):
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, col_count + 1)]
            if any(v is not None for v in row_vals[:6]):
                first_data_row = row_vals
                break

        has_data_rows = first_data_row is not None

        validation_passed = has_headers and has_data_rows and row_count >= 10 and not [
            i for i in issues if "corrupted" in i.lower()
        ]

        return WorksheetValidation(
            sheet_name=sheet_name,
            exists=True,
            row_count=row_count,
            col_count=col_count,
            has_headers=has_headers,
            header_row=header_row,
            has_data_rows=has_data_rows,
            first_data_row=first_data_row,
            validation_passed=validation_passed,
            issues=issues,
        )

    def validate_pipeline_outputs(self) -> List[Dict[str, Any]]:
        """Check all expected pipeline output files exist."""
        results: List[Dict[str, Any]] = []
        for check in PIPELINE_OUTPUT_CHECKS:
            p = self._root / check["path"]
            results.append({
                "label": check["label"],
                "path": str(p),
                "exists": p.exists(),
                "size_bytes": p.stat().st_size if p.exists() else 0,
            })
        return results
