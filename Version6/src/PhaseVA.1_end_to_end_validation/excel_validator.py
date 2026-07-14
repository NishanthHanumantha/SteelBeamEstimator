"""
Phase V.A.1 — End-to-End Validation
excel_validator.py — Validate the generated Excel workbook.
MODEL_VERSION: 6.5.3
"""
from __future__ import annotations

import pathlib
from typing import List

from validation_models import WorkbookValidation


class ExcelValidator:
    """Validates existence, readability, and integrity of the generated workbook."""

    # Expected minimum workbook size (100 KB)
    MIN_SIZE_BYTES = 100_000

    def validate(self, workbook_path: pathlib.Path) -> WorkbookValidation:
        issues: List[str] = []

        if not workbook_path.exists():
            return WorkbookValidation(
                workbook_path=str(workbook_path),
                exists=False,
                readable=False,
                corrupted=True,
                size_bytes=0,
                sheet_names=[],
                total_sheets=0,
                total_rows=0,
                total_columns=0,
                has_data=False,
                validation_passed=False,
                issues=[f"Workbook not found: {workbook_path}"],
            )

        size_bytes = workbook_path.stat().st_size
        if size_bytes < self.MIN_SIZE_BYTES:
            issues.append(
                f"Workbook size {size_bytes:,} bytes is below minimum "
                f"{self.MIN_SIZE_BYTES:,} bytes — may be empty or corrupted."
            )

        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(workbook_path))
        except Exception as exc:
            return WorkbookValidation(
                workbook_path=str(workbook_path),
                exists=True,
                readable=False,
                corrupted=True,
                size_bytes=size_bytes,
                sheet_names=[],
                total_sheets=0,
                total_rows=0,
                total_columns=0,
                has_data=False,
                validation_passed=False,
                issues=[f"Workbook could not be opened: {exc}"],
            )

        sheet_names = wb.sheetnames
        if not sheet_names:
            issues.append("Workbook contains no sheets.")

        total_rows = 0
        total_cols = 0
        has_data = False
        for sn in sheet_names:
            ws = wb[sn]
            total_rows += ws.max_row
            total_cols = max(total_cols, ws.max_column)
            if ws.max_row > 1:
                has_data = True

        if not has_data:
            issues.append("Workbook has no data rows.")

        # Check for project header (sanity)
        if sheet_names:
            first_ws = wb[sheet_names[0]]
            first_cell = first_ws.cell(row=1, column=1).value
            if first_cell and "Project" not in str(first_cell):
                issues.append(
                    f"Row 1 cell A1 = '{first_cell}' — expected 'Project :' header."
                )

        validation_passed = (
            size_bytes >= self.MIN_SIZE_BYTES
            and len(sheet_names) > 0
            and has_data
            and len([i for i in issues if "corrupted" in i.lower() or "not found" in i.lower()]) == 0
        )

        return WorkbookValidation(
            workbook_path=str(workbook_path),
            exists=True,
            readable=True,
            corrupted=False,
            size_bytes=size_bytes,
            sheet_names=sheet_names,
            total_sheets=len(sheet_names),
            total_rows=total_rows,
            total_columns=total_cols,
            has_data=has_data,
            validation_passed=validation_passed,
            issues=issues,
        )
