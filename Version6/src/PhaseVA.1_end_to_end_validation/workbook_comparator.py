"""
Phase V.A.1 — End-to-End Validation
workbook_comparator.py — Compare generated workbook vs estimator reference.
No modifications. Comparison only.
MODEL_VERSION: 6.5.3
"""
from __future__ import annotations

import pathlib
from typing import Any, Dict, List, Optional, Tuple

from validation_models import CellComparison, WorkbookComparison, WorksheetComparison


def _numeric(v: Any) -> Optional[float]:
    """Try to convert a cell value to float."""
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _pct_diff(a: float, b: float) -> Optional[float]:
    if b == 0:
        return None
    return round(abs(a - b) / abs(b) * 100, 4)


class WorkbookComparator:
    """
    Compares the generated workbook against the estimator reference.
    Pure read-only comparison. No modifications to either file.
    """

    # Maximum data rows to compare per sheet (for performance)
    MAX_ROWS_TO_COMPARE = 500

    def compare(
        self,
        generated_path: pathlib.Path,
        reference_path: pathlib.Path,
    ) -> WorkbookComparison:
        import openpyxl

        gen_sheets: List[str] = []
        ref_sheets: List[str] = []
        worksheet_comparisons: List[WorksheetComparison] = []

        gen_ok = generated_path.exists()
        ref_ok = reference_path.exists()

        if gen_ok:
            wb_gen = openpyxl.load_workbook(str(generated_path))
            gen_sheets = wb_gen.sheetnames
        else:
            wb_gen = None

        if ref_ok:
            wb_ref = openpyxl.load_workbook(str(reference_path), data_only=True)
            ref_sheets = wb_ref.sheetnames
        else:
            wb_ref = None

        common = list(set(gen_sheets) & set(ref_sheets))
        missing_in_gen = [s for s in ref_sheets if s not in gen_sheets]
        extra_in_gen = [s for s in gen_sheets if s not in ref_sheets]

        for sheet_name in common:
            ws_gen = wb_gen[sheet_name] if wb_gen else None
            ws_ref = wb_ref[sheet_name] if wb_ref else None
            comp = self._compare_sheet(sheet_name, ws_gen, ws_ref)
            worksheet_comparisons.append(comp)

        # Overall match rate (weighted by rows compared)
        total_cells = sum(c.data_rows_compared * min(c.gen_cols, c.ref_cols)
                          for c in worksheet_comparisons)
        total_match = sum(c.matching_cells for c in worksheet_comparisons)
        overall = round(100 * total_match / total_cells, 4) if total_cells > 0 else 0.0

        # Steel weight and quantity comparison
        steel_comp = {}
        qty_comp = {}
        if worksheet_comparisons:
            sheet_comp = worksheet_comparisons[0]
            steel_comp = self._compare_totals(
                wb_gen, wb_ref, common[0] if common else None, "Steel in KG"
            )
            qty_comp = self._compare_totals(
                wb_gen, wb_ref, common[0] if common else None, "Total Length"
            )

        return WorkbookComparison(
            generated_path=str(generated_path),
            reference_path=str(reference_path),
            generated_sheets=gen_sheets,
            reference_sheets=ref_sheets,
            sheet_count_match=len(gen_sheets) == len(ref_sheets),
            sheet_names_match=set(gen_sheets) == set(ref_sheets),
            common_sheets=common,
            missing_in_generated=missing_in_gen,
            extra_in_generated=extra_in_gen,
            worksheet_comparisons=worksheet_comparisons,
            overall_match_rate_pct=overall,
            totals_match=steel_comp.get("match", False),
            steel_weight_comparison=steel_comp,
            quantity_comparison=qty_comp,
        )

    def _compare_sheet(
        self,
        sheet_name: str,
        ws_gen: Any,
        ws_ref: Any,
    ) -> WorksheetComparison:
        if ws_gen is None or ws_ref is None:
            return WorksheetComparison(
                sheet_name=sheet_name,
                gen_rows=0, ref_rows=0, gen_cols=0, ref_cols=0,
                row_count_match=False, col_count_match=False,
                header_match=False, data_rows_compared=0,
                matching_cells=0, mismatching_cells=0, match_rate_pct=0.0,
            )

        gen_rows = ws_gen.max_row
        ref_rows = ws_ref.max_row
        gen_cols = ws_gen.max_column
        ref_cols = ws_ref.max_column

        # Header comparison (row 6 — the main column header row)
        HEADER_ROW = 6
        gen_header = [ws_gen.cell(row=HEADER_ROW, column=c).value
                      for c in range(1, min(gen_cols, ref_cols) + 1)]
        ref_header = [ws_ref.cell(row=HEADER_ROW, column=c).value
                      for c in range(1, min(gen_cols, ref_cols) + 1)]
        header_match = _headers_compatible(gen_header, ref_header)

        # Data rows comparison (start from row 9)
        START_ROW = 9
        compare_rows = min(gen_rows, ref_rows, self.MAX_ROWS_TO_COMPARE + START_ROW)
        compare_cols = min(gen_cols, ref_cols, 17)  # first 17 cols (core data)

        matching = 0
        mismatching = 0
        key_diffs: List[Dict[str, Any]] = []

        for r in range(START_ROW, compare_rows + 1):
            for c in range(1, compare_cols + 1):
                gen_val = ws_gen.cell(row=r, column=c).value
                ref_val = ws_ref.cell(row=r, column=c).value

                if gen_val is None and ref_val is None:
                    matching += 1
                    continue

                gen_num = _numeric(gen_val)
                ref_num = _numeric(ref_val)

                if gen_num is not None and ref_num is not None:
                    if abs(gen_num - ref_num) < 0.01:
                        matching += 1
                    else:
                        mismatching += 1
                        if len(key_diffs) < 50:
                            key_diffs.append({
                                "row": r, "col": c,
                                "generated": gen_num, "reference": ref_num,
                                "diff_pct": _pct_diff(gen_num, ref_num),
                            })
                elif str(gen_val).strip() == str(ref_val).strip():
                    matching += 1
                else:
                    # One is non-numeric string or None
                    if gen_val is None and ref_val is not None:
                        mismatching += 1
                        if len(key_diffs) < 50:
                            key_diffs.append({
                                "row": r, "col": c,
                                "generated": None, "reference": str(ref_val)[:50],
                                "note": "generated=null",
                            })
                    elif str(gen_val).strip() == str(ref_val).strip():
                        matching += 1
                    else:
                        mismatching += 1

        total_cells = (compare_rows - START_ROW + 1) * compare_cols
        match_rate = round(100 * matching / total_cells, 4) if total_cells > 0 else 0.0

        return WorksheetComparison(
            sheet_name=sheet_name,
            gen_rows=gen_rows,
            ref_rows=ref_rows,
            gen_cols=gen_cols,
            ref_cols=ref_cols,
            row_count_match=gen_rows == ref_rows,
            col_count_match=gen_cols == ref_cols,
            header_match=header_match,
            data_rows_compared=compare_rows - START_ROW + 1,
            matching_cells=matching,
            mismatching_cells=mismatching,
            match_rate_pct=match_rate,
            key_differences=key_diffs[:20],
        )

    def _compare_totals(
        self,
        wb_gen: Any,
        wb_ref: Any,
        sheet_name: Optional[str],
        column_header: str,
    ) -> Dict[str, Any]:
        if not sheet_name or wb_gen is None or wb_ref is None:
            return {"match": False, "note": "workbook not available"}
        try:
            ws_gen = wb_gen[sheet_name]
            ws_ref = wb_ref[sheet_name]

            # Find the steel KG column in generated
            HEADER_ROW = 6
            gen_col = None
            for c in range(1, ws_gen.max_column + 1):
                v = ws_gen.cell(row=HEADER_ROW, column=c).value
                if v and column_header in str(v):
                    gen_col = c
                    break

            ref_col = None
            for c in range(1, ws_ref.max_column + 1):
                v = ws_ref.cell(row=HEADER_ROW, column=c).value
                if v and column_header in str(v):
                    ref_col = c
                    break

            if gen_col is None or ref_col is None:
                return {"match": False, "note": f"Column '{column_header}' not found"}

            # Sum all numeric values in that column
            gen_total = 0.0
            for r in range(9, ws_gen.max_row + 1):
                v = _numeric(ws_gen.cell(row=r, column=gen_col).value)
                if v:
                    gen_total += v

            ref_total = 0.0
            for r in range(9, ws_ref.max_row + 1):
                v = _numeric(ws_ref.cell(row=r, column=ref_col).value)
                if v:
                    ref_total += v

            diff = abs(gen_total - ref_total)
            diff_pct = _pct_diff(gen_total, ref_total)
            match = diff_pct is not None and diff_pct < 5.0

            return {
                "column": column_header,
                "generated_total": round(gen_total, 3),
                "reference_total": round(ref_total, 3),
                "difference": round(diff, 3),
                "difference_pct": diff_pct,
                "match": match,
                "within_5pct_tolerance": match,
            }
        except Exception as exc:
            return {"match": False, "note": str(exc)}


def _headers_compatible(gen: List[Any], ref: List[Any]) -> bool:
    """True if at least 50% of non-None header cells match."""
    pairs = [(g, r) for g, r in zip(gen, ref) if g is not None or r is not None]
    if not pairs:
        return False
    matches = sum(1 for g, r in pairs
                  if g is not None and r is not None and str(g).strip() == str(r).strip())
    return matches / len(pairs) >= 0.5
